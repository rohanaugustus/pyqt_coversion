"""
Tkinter-based GUI with a modern Windows 11 tab layout.
Tab 1: Timeline Dashboard with embedded interactive chart
Tab 2: Project Manager
Tab 3: Milestone & Phase Editor
Tab 4: Admin Panel (admin only)
"""

from __future__ import annotations

import datetime
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
import pathlib

import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.patches as mpatches
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk

from timeline_tool import database as db
from timeline_tool import auth
from timeline_tool import config as cfg
from timeline_tool.models import Project, ReferenceLine
from timeline_tool.utils import date_range_padded
from timeline_tool.renderer import render_timeline

from timeline_tool.export_report import (
    generate_pdf_report, generate_excel_report, 
    check_dependencies, REPORTLAB_AVAILABLE, OPENPYXL_AVAILABLE
)

from timeline_tool.resources import (
    Resource, init_resource_tables, add_resource, update_resource, 
    delete_resource, get_all_resources, assign_resource_to_project,
    remove_assignment, get_project_assignments, get_resource_assignments,
    calculate_resource_utilization, get_team_utilization_summary,
    find_available_resources
)

from timeline_tool.audit_viewer import (
    get_audit_log, get_unique_users, get_unique_actions,
    get_activity_summary, export_audit_log_csv, get_entity_history,
    get_action_description, get_action_icon, ACTION_DESCRIPTIONS
)

from timeline_tool.backup import (
    create_backup, restore_backup, list_backups, delete_backup,
    export_to_json, import_from_json, get_backup_dir
)

# ─────────────────────────────────────────────────────────────────────────
# Phase colours
# ─────────────────────────────────────────────────────────────────────────
PHASE_PALETTE = ["#85C1E9", "#82E0AA", "#F8C471", "#D7BDE2", "#F0B27A", "#AED6F1"]

REGION_OPTIONS = ["", "IAP", "EMEA", "NAFTA", "LATAM", "ROW"]


def _pick_color(index: int, project: Project) -> str:
    if project.color:
        return project.color
    return cfg.DEFAULT_COLORS[index % len(cfg.DEFAULT_COLORS)]


def _status_edge_color(project: Project) -> str:
    return cfg.STATUS_COLORS.get(project.computed_status(), cfg.STATUS_COLORS["on-track"])


# ─────────────────────────────────────────────────────────────────────────
# Windows 11 style theme
# ─────────────────────────────────────────────────────────────────────────

def _apply_win11_style(root: tk.Tk):
    style = ttk.Style()
    style.theme_use("clam")

    BG = "#F3F3F3"
    CARD_BG = "#FFFFFF"
    ACCENT = "#0067C0"
    TEXT = "#1A1A1A"
    TEXT_SECONDARY = "#666666"
    BORDER = "#E0E0E0"
    HOVER = "#E8F0FE"
    TAB_ACTIVE = "#FFFFFF"
    TAB_INACTIVE = "#E8E8E8"

    root.configure(bg=BG)

    style.configure(".", background=BG, foreground=TEXT, font=("Segoe UI", 10))
    style.configure("TFrame", background=BG)
    style.configure("TLabel", background=BG, foreground=TEXT, font=("Segoe UI", 10))
    style.configure("TButton", background=CARD_BG, foreground=TEXT, font=("Segoe UI", 10),
                     borderwidth=1, relief="flat", padding=(12, 6))
    style.map("TButton",
              background=[("active", HOVER), ("pressed", ACCENT)],
              foreground=[("pressed", "white")])

    style.configure("Accent.TButton", background=ACCENT, foreground="white",
                     font=("Segoe UI", 10, "bold"), padding=(14, 8))
    style.map("Accent.TButton",
              background=[("active", "#005A9E"), ("pressed", "#004578")])

    style.configure("TNotebook", background=BG, borderwidth=0)
    style.configure("TNotebook.Tab", background=TAB_INACTIVE, foreground=TEXT,
                     font=("Segoe UI", 11), padding=(20, 10))
    style.map("TNotebook.Tab",
              background=[("selected", TAB_ACTIVE)],
              foreground=[("selected", ACCENT)])

    style.configure("Card.TFrame", background=CARD_BG, relief="flat", borderwidth=1)
    style.configure("Card.TLabel", background=CARD_BG, foreground=TEXT)
    style.configure("CardTitle.TLabel", background=CARD_BG, foreground=TEXT,
                     font=("Segoe UI", 12, "bold"))
    style.configure("CardSubtitle.TLabel", background=CARD_BG, foreground=TEXT_SECONDARY,
                     font=("Segoe UI", 9))

    style.configure("Header.TLabel", background=BG, foreground=TEXT,
                     font=("Segoe UI", 20, "bold"))
    style.configure("SubHeader.TLabel", background=BG, foreground=TEXT_SECONDARY,
                     font=("Segoe UI", 11))

    style.configure("Status.TLabel", background=BG, font=("Segoe UI", 9))

    style.configure("Treeview", background=CARD_BG, foreground=TEXT,
                     fieldbackground=CARD_BG, font=("Segoe UI", 10),
                     rowheight=28, borderwidth=0)
    style.configure("Treeview.Heading", background=BG, foreground=TEXT,
                     font=("Segoe UI", 10, "bold"), borderwidth=0)
    style.map("Treeview", background=[("selected", HOVER)],
              foreground=[("selected", TEXT)])

    style.configure("TLabelframe", background=CARD_BG, foreground=TEXT,
                     font=("Segoe UI", 10, "bold"), borderwidth=1, relief="flat")
    style.configure("TLabelframe.Label", background=CARD_BG, foreground=ACCENT,
                     font=("Segoe UI", 10, "bold"))

    style.configure("TCombobox", font=("Segoe UI", 10))
    style.configure("TEntry", font=("Segoe UI", 10))

    return {
        "BG": BG, "CARD_BG": CARD_BG, "ACCENT": ACCENT,
        "TEXT": TEXT, "TEXT_SECONDARY": TEXT_SECONDARY,
        "BORDER": BORDER, "HOVER": HOVER,
    }


# ─────────────────────────────────────────────────────────────────────────
# Login Window
# ─────────────────────────────────────────────────────────────────────────

class LoginWindow:
    def __init__(self, root: tk.Tk, db_path: pathlib.Path | None = None):
        self.root = root
        self.db_path = db_path
        self.user = None
        self._password_visible = False

        self.root.title("Project Dashboard")
        
        # Use consistent height - 650 to ensure button is visible
        width = 450
        height = 650
        self.root.geometry(f"{width}x{height}")
        self.root.resizable(False, False)
        
        # Center window on screen
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f"{width}x{height}+{x}+{y}")

        _apply_win11_style(root)

        # ══════════════════════════════════════════════════════════════════
        # FOOTER (pack FIRST with side="bottom")
        # ══════════════════════════════════════════════════════════════════
        footer = ttk.Frame(root, padding=(10, 8))
        footer.pack(side="bottom", fill="x")
        ttk.Separator(root, orient="horizontal").pack(side="bottom", fill="x")
        ttk.Label(footer, text="Developed by IAP VSSQ AI/ML Team",
                  font=("Segoe UI", 9, "italic"), foreground="#888888",
                  anchor="center").pack(fill="x")
        ttk.Label(footer, text="v2.0 | © 2025",
                  font=("Segoe UI", 8), foreground="#AAAAAA",
                  anchor="center").pack(fill="x")

        # ══════════════════════════════════════════════════════════════════
        # MAIN CONTENT
        # ══════════════════════════════════════════════════════════════════
        outer = ttk.Frame(root, padding=(40, 25, 40, 25))
        outer.pack(fill="both", expand=True)

        # ── Company Logo ─────────────────────────────────────────────────
        logo_path = pathlib.Path(
            r"C:\Users\T0276HS\OneDrive - Stellantis\AI Activities"
            r"\Project Timelines\project-timeline-tool\resources\icons\Stellantis.svg.png"
        )
        if logo_path.exists():
            try:
                from PIL import Image, ImageTk
                img = Image.open(logo_path)
                max_h = 55
                ratio = max_h / img.height
                new_size = (int(img.width * ratio), max_h)
                img = img.resize(new_size, Image.LANCZOS)
                self._logo_img = ImageTk.PhotoImage(img)
                logo_label = ttk.Label(outer, image=self._logo_img, background="#F3F3F3")
                logo_label.pack(pady=(0, 12))
            except ImportError:
                ttk.Label(outer, text="Project", font=("Segoe UI", 20, "bold")).pack(pady=(0, 8))
        else:
            ttk.Label(outer, text="Project", font=("Segoe UI", 20, "bold")).pack(pady=(0, 8))

        # ── Title & Subtitle ─────────────────────────────────────────────
        ttk.Label(outer, text="Project Dashboard", 
                  style="Header.TLabel",
                  font=("Segoe UI", 18, "bold")).pack()
        ttk.Label(outer, text="Sign in to access your projects", 
                  style="SubHeader.TLabel",
                  font=("Segoe UI", 10)).pack(pady=(0, 18))

        # ══════════════════════════════════════════════════════════════════
        # LOGIN CARD
        # ══════════════════════════════════════════════════════════════════
        card = tk.Frame(outer, bg="white", bd=1,
                        highlightbackground="#E0E0E0", highlightthickness=1)
        card.pack(fill="x", pady=(0, 12))
        
        form = tk.Frame(card, bg="white", padx=25, pady=20)
        form.pack(fill="x")

        # ── USERNAME FIELD ───────────────────────────────────────────────
        tk.Label(form, text="Username", font=("Segoe UI", 10, "bold"),
                 bg="white", fg="#333333", anchor="w").pack(fill="x")
        
        user_frame = tk.Frame(form, bg="#F5F5F5", highlightbackground="#DDDDDD", 
                              highlightthickness=1)
        user_frame.pack(fill="x", pady=(5, 12))
        
        tk.Label(user_frame, text="👤", font=("Segoe UI", 12), 
                 bg="#F5F5F5", fg="#666666", width=3).pack(side="left", padx=(5, 0))
        
        self.username_entry = tk.Entry(user_frame, font=("Segoe UI", 11),
                                        bg="#F5F5F5", fg="#333333",
                                        relief="flat", bd=0)
        self.username_entry.pack(side="left", fill="x", expand=True, ipady=8, padx=(5, 10))

        # ── PASSWORD FIELD ───────────────────────────────────────────────
        tk.Label(form, text="Password", font=("Segoe UI", 10, "bold"),
                 bg="white", fg="#333333", anchor="w").pack(fill="x")
        
        pass_frame = tk.Frame(form, bg="#F5F5F5", highlightbackground="#DDDDDD",
                              highlightthickness=1)
        pass_frame.pack(fill="x", pady=(5, 12))
        
        tk.Label(pass_frame, text="🔒", font=("Segoe UI", 12),
                 bg="#F5F5F5", fg="#666666", width=3).pack(side="left", padx=(5, 0))
        
        self.password_entry = tk.Entry(pass_frame, font=("Segoe UI", 11),
                                        bg="#F5F5F5", fg="#333333",
                                        relief="flat", bd=0, show="•")
        self.password_entry.pack(side="left", fill="x", expand=True, ipady=8, padx=(5, 0))
        
        # Password visibility toggle button
        self.eye_btn = tk.Button(pass_frame, text="👁", font=("Segoe UI", 10),
                                  bg="#F5F5F5", fg="#666666", relief="flat",
                                  bd=0, cursor="hand2", width=3,
                                  activebackground="#F5F5F5",
                                  command=self._toggle_password_visibility)
        self.eye_btn.pack(side="right", padx=(0, 5))

        # ── REMEMBER ME & FORGOT PASSWORD ────────────────────────────────
        remember_frame = tk.Frame(form, bg="white")
        remember_frame.pack(fill="x", pady=(5, 15))
        
        self.remember_var = tk.BooleanVar(value=False)
        self.remember_check = tk.Checkbutton(
            remember_frame, text="Remember me",
            variable=self.remember_var,
            font=("Segoe UI", 9), bg="white", fg="#666666",
            activebackground="white", selectcolor="white",
            cursor="hand2"
        )
        self.remember_check.pack(side="left")
        
        # Forgot password link
        forgot_label = tk.Label(remember_frame, text="Forgot password?",
                                font=("Segoe UI", 9, "underline"),
                                bg="white", fg="#0067C0", cursor="hand2")
        forgot_label.pack(side="right")
        forgot_label.bind("<Button-1>", self._on_forgot_password)

        # ── SIGN IN BUTTON ───────────────────────────────────────────────
        # Use simple text without emoji for maximum compatibility
        self.signin_btn = tk.Button(form, text="Sign In",
                                     font=("Segoe UI", 12, "bold"),
                                     bg="#0067C0", fg="white",
                                     relief="flat", cursor="hand2",
                                     activebackground="#005A9E",
                                     activeforeground="white",
                                     command=self._login)
        self.signin_btn.pack(fill="x", ipady=12, pady=(5, 0))
        
        # Hover effects for button
        self.signin_btn.bind("<Enter>", lambda e: self.signin_btn.config(bg="#005A9E"))
        self.signin_btn.bind("<Leave>", lambda e: self.signin_btn.config(bg="#0067C0"))

        # ══════════════════════════════════════════════════════════════════
        # STATUS LABEL
        # ══════════════════════════════════════════════════════════════════
        self.status_label = ttk.Label(outer, text="", foreground="red", 
                                       font=("Segoe UI", 9),
                                       wraplength=350, justify="center")
        self.status_label.pack(pady=(8, 0))

        # ══════════════════════════════════════════════════════════════════
        # HELP TEXT
        # ══════════════════════════════════════════════════════════════════
        help_frame = ttk.Frame(outer)
        help_frame.pack(pady=(10, 0))
        ttk.Label(help_frame, text="Need help? Contact your administrator",
                  font=("Segoe UI", 9), foreground="#999999").pack()

        # ══════════════════════════════════════════════════════════════════
        # BINDINGS
        # ══════════════════════════════════════════════════════════════════
        self.password_entry.bind("<Return>", lambda e: self._login())
        self.username_entry.bind("<Return>", lambda e: self.password_entry.focus())
        self.username_entry.focus()
        
        # Load remembered username if exists
        self._load_remembered_user()

    def _toggle_password_visibility(self):
        """Toggle password visibility."""
        self._password_visible = not self._password_visible
        if self._password_visible:
            self.password_entry.config(show="")
            self.eye_btn.config(text="🙈")
        else:
            self.password_entry.config(show="•")
            self.eye_btn.config(text="👁")

    def _on_forgot_password(self, event=None):
        """Handle forgot password click."""
        messagebox.showinfo(
            "Reset Password",
            "To reset your password, please contact your system administrator.\n\n"
            "Email: admin@company.com\n"
            "Phone: ext. 1234\n\n"
            "Please have your username ready when you contact support."
        )

    def _load_remembered_user(self):
        """Load remembered username from file."""
        try:
            remember_file = pathlib.Path.home() / ".project_dashboard_remember"
            if remember_file.exists():
                username = remember_file.read_text().strip()
                if username:
                    self.username_entry.insert(0, username)
                    self.remember_var.set(True)
                    self.password_entry.focus()
        except Exception:
            pass

    def _save_remembered_user(self, username: str):
        """Save username for remember-me functionality."""
        try:
            remember_file = pathlib.Path.home() / ".project_dashboard_remember"
            if self.remember_var.get() and username:
                remember_file.write_text(username)
            elif remember_file.exists():
                remember_file.unlink()
        except Exception:
            pass

    def _login(self):
        username = self.username_entry.get().strip()
        password = self.password_entry.get().strip()
        
        if not username:
            self.status_label.config(text="Please enter your username.")
            self.username_entry.focus()
            return
        
        if not password:
            self.status_label.config(text="Please enter your password.")
            self.password_entry.focus()
            return

        # Show loading state
        original_text = self.signin_btn.cget("text")
        self.signin_btn.config(text="Signing in...", state="disabled")
        self.root.update()

        user = auth.authenticate(username, password, self.db_path)
        
        if user is None:
            self.signin_btn.config(text=original_text, state="normal")
            self.status_label.config(text="Invalid username or password. Please try again.")
            self.password_entry.delete(0, tk.END)
            self.password_entry.focus()
            
            # Flash the button red briefly for error feedback
            self.signin_btn.config(bg="#DC2626")
            self.root.after(200, lambda: self.signin_btn.config(bg="#0067C0"))
            return

        # Save remember-me preference
        self._save_remembered_user(username)

        self.user = user
        self.root.destroy()



# ─────────────────────────────────────────────────────────────────────────
# Main Application (Tab Layout)
# ─────────────────────────────────────────────────────────────────────────

class MainApp:
    def __init__(self, root: tk.Tk, user: dict, db_path: pathlib.Path | None = None):
        self.root = root
        self.user = user
        self.db_path = db_path
        self.can_edit = user["permissions"]["can_edit"]
        self.can_manage = user["permissions"]["can_manage_users"]

        display_name = user.get("full_name") or user["username"]
        role_display = user["role"].upper()
        self.root.title("Project Dashboard")
        self.root.state("zoomed")  # Start maximized
        self.root.minsize(1000, 650)

        self.colors = _apply_win11_style(root)

        # ── Set window / taskbar icon ────────────────────────────────────
        self._icon_photo = None
        _icon_path = pathlib.Path(
            r"C:\Users\T0276HS\OneDrive - Stellantis\AI Activities"
            r"\Project Timelines\project-timeline-tool\resources\icons\icon.png"
        )
        if _icon_path.exists():
            # Try .ico first for Windows taskbar compatibility
            _ico_path = _icon_path.with_suffix(".ico")
            if _ico_path.exists():
                try:
                    self.root.iconbitmap(_ico_path)
                except Exception:
                    pass
            # Also set PNG icon using PIL for better cross-platform support
            try:
                from PIL import Image, ImageTk
                _icon_img = Image.open(_icon_path)
                self._icon_photo = ImageTk.PhotoImage(_icon_img)
                self.root.iconphoto(True, self._icon_photo)
            except Exception:
                pass

        # ── Stellantis logo (keep reference to prevent garbage collection) ─
        self._logo_img_main = None
        logo_path = pathlib.Path(
            r"C:\Users\T0276HS\OneDrive - Stellantis\AI Activities"
            r"\Project Timelines\project-timeline-tool\resources\icons\Stellantis.svg.png"
        )

        # ── Top bar ─────────────────────────────────────────────────────
        topbar = ttk.Frame(root, padding=(15, 8))
        topbar.pack(fill="x")

        # Left: app icon + title
        left_frame = ttk.Frame(topbar)
        left_frame.pack(side="left")
        ttk.Label(left_frame, text="📊 Project Dashboard",
                  font=("Segoe UI", 14, "bold")).pack(side="left")

        # Right: logo + user pill
        right_frame = ttk.Frame(topbar)
        right_frame.pack(side="right")

        # Stellantis logo in top-right
        if logo_path.exists():
            try:
                from PIL import Image, ImageTk
                img = Image.open(logo_path)
                max_h = 35
                ratio = max_h / img.height
                new_size = (int(img.width * ratio), max_h)
                img = img.resize(new_size, Image.LANCZOS)
                self._logo_img_main = ImageTk.PhotoImage(img)
                ttk.Label(right_frame, image=self._logo_img_main,
                          background="#F3F3F3").pack(side="right", padx=(15, 0))
            except ImportError:
                pass

        # User pill
        user_frame = ttk.Frame(right_frame)
        user_frame.pack(side="right")
        role_colors = {"admin": "#0067C0", "editor": "#16A34A", "viewer": "#9333EA"}
        role_color = role_colors.get(user["role"], "#666666")

        ttk.Label(user_frame, text=f"👤 {display_name}",
                  font=("Segoe UI", 10)).pack(side="left", padx=(0, 8))
        tk.Label(user_frame, text=f" {role_display} ",
                 bg=role_color, fg="white",
                 font=("Segoe UI", 8, "bold"),
                 padx=8, pady=2).pack(side="left")

        ttk.Separator(root, orient="horizontal").pack(fill="x")

        # ── Footer (pack BEFORE notebook so it stays at bottom) ─────────
        footer_sep = ttk.Separator(root, orient="horizontal")
        footer_sep.pack(side="bottom", fill="x")

        footer = ttk.Frame(root, padding=(10, 6))
        footer.pack(side="bottom", fill="x")

        ttk.Label(footer, text="Developed by IAP VSSQ AI/ML Team",
                  font=("Segoe UI", 9, "bold"), foreground="#555555").pack(side="left")

        ttk.Label(footer, text=f"Database: {db_path or db.DEFAULT_DB_PATH}",
                  font=("Segoe UI", 8), foreground="#999999").pack(side="right")

        # ── Status bar (above footer) ───────────────────────────────────
        status_bar = ttk.Frame(root, padding=(10, 3))
        status_bar.pack(side="bottom", fill="x")

        self.status_text = ttk.Label(status_bar, text="Ready",
                                      font=("Segoe UI", 9), foreground="#666666")
        self.status_text.pack(side="left")

        self.coords_label = ttk.Label(status_bar, text="",
                                       font=("Segoe UI", 8), foreground="#999999")
        self.coords_label.pack(side="right")

        # ── Notebook (tabs) ─────────────────────────────────────────────
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill="both", expand=True, padx=5, pady=(5, 0))

        # Tab 1: Dashboard with embedded chart
        self.tab_dashboard = ttk.Frame(self.notebook, padding=3)
        self.notebook.add(self.tab_dashboard, text="  📊 Dashboard  ")
        self._build_dashboard_tab()

        # Tab 2: QCTP
        self.tab_qctp = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(self.tab_qctp, text="  📋 QCTP  ")
        self._build_qctp_tab()

        # Tab 3: Activities  ← NEW TAB
        self.tab_activities = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(self.tab_activities, text="  📝 Activities  ")
        self._build_activities_tab()

        # Tab 4: Projects  (was Tab 3)
        self.tab_projects = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(self.tab_projects, text="  📁 Projects  ")
        self._build_projects_tab()

        # Tab 5: Milestones & Phases  (was Tab 4)
        self.tab_milestones = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(self.tab_milestones, text="  🎯 Milestones & Phases  ")
        self._build_milestones_tab()

        # Tab 6: Resources  (was Tab 5)
        self.tab_resources = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(self.tab_resources, text="  👥 Resources  ")
        self._build_resources_tab()

        # Tab 7: Admin (admin only)  (was Tab 6)
        if self.can_manage:
            self.tab_admin = ttk.Frame(self.notebook, padding=10)
            self.notebook.add(self.tab_admin, text="  ⚙️ Admin  ")
            self._build_admin_tab()

    def _set_status(self, text: str):
        self.status_text.config(text=text)
        self.root.after(5000, lambda: self.status_text.config(text="Ready"))

    # ═════════════════════════════════════════════════════════════════════
    # TAB 1 — DASHBOARD (Embedded Chart)
    # ═════════════════════════════════════════════════════════════════════

    def _build_dashboard_tab(self):
        tab = self.tab_dashboard

        # ══════════════════════════════════════════════════════════════════
        # TOP ROW: Summary Cards + Action Buttons
        # ══════════════════════════════════════════════════════════════════
        top_row = ttk.Frame(tab)
        top_row.pack(fill="x", pady=(0, 5))

        # Summary cards (left side)
        self.cards_frame = ttk.Frame(top_row)
        self.cards_frame.pack(side="left", fill="x", expand=True)

        # Buttons (right side)
        btn_frame = ttk.Frame(top_row)
        btn_frame.pack(side="right")

        ttk.Button(btn_frame, text="🔄 Refresh", command=self._refresh_dashboard).pack(side="left", padx=3)
        ttk.Button(btn_frame, text="🪟 Pop-out Chart", command=self._popout_timeline).pack(side="left", padx=3)
        ttk.Button(btn_frame, text="💾 Save PNG", command=self._save_timeline_png).pack(side="left", padx=3)
        ttk.Button(btn_frame, text="🌐 Save HTML", command=self._save_timeline_html).pack(side="left", padx=3)
        ttk.Button(btn_frame, text="📄 Export PDF", command=self._export_pdf_report).pack(side="left", padx=3)
        ttk.Button(btn_frame, text="📊 Export Excel", command=self._export_excel_report).pack(side="left", padx=3)
        ttk.Button(btn_frame, text="📈 Summary", command=self._show_summary_dashboard).pack(side="left", padx=3)

        # Theme toggle
        self.theme_var = tk.StringVar(value="light")
        ttk.Radiobutton(btn_frame, text="Light", variable=self.theme_var,
                         value="light", command=self._refresh_dashboard).pack(side="left", padx=3)
        ttk.Radiobutton(btn_frame, text="Dark", variable=self.theme_var,
                         value="dark", command=self._refresh_dashboard).pack(side="left", padx=3)

        # ══════════════════════════════════════════════════════════════════
        # FILTER BAR (Enhanced with more controls)
        # ══════════════════════════════════════════════════════════════════
        filter_frame = ttk.LabelFrame(tab, text="  🔍 Filters & Controls  ", padding=8)
        filter_frame.pack(fill="x", pady=(0, 5))
        
        filter_row = ttk.Frame(filter_frame)
        filter_row.pack(fill="x")

        # Status filter
        ttk.Label(filter_row, text="Status:").pack(side="left", padx=(0, 3))
        self.filter_status = ttk.Combobox(filter_row, 
                                          values=["All", "On Track", "At Risk", "Overdue"],
                                          width=10, font=("Segoe UI", 9), state="readonly")
        self.filter_status.set("All")
        self.filter_status.pack(side="left", padx=(0, 10))
        self.filter_status.bind("<<ComboboxSelected>>", lambda e: self._refresh_dashboard())

        # Dev Region filter
        ttk.Label(filter_row, text="Dev Region:").pack(side="left", padx=(0, 3))
        self.filter_dev_region = ttk.Combobox(filter_row, values=["All", "IAP", "EMEA", "NAFTA", "LATAM", "ROW"],
                                               width=10, font=("Segoe UI", 9), state="readonly")
        self.filter_dev_region.set("All")
        self.filter_dev_region.pack(side="left", padx=(0, 10))
        self.filter_dev_region.bind("<<ComboboxSelected>>", lambda e: self._refresh_dashboard())

        # Sales Region filter
        ttk.Label(filter_row, text="Sales Region:").pack(side="left", padx=(0, 3))
        self.filter_sales_region = ttk.Combobox(filter_row, values=["All", "IAP", "EMEA", "NAFTA", "LATAM", "ROW"],
                                                 width=10, font=("Segoe UI", 9), state="readonly")
        self.filter_sales_region.set("All")
        self.filter_sales_region.pack(side="left", padx=(0, 10))
        self.filter_sales_region.bind("<<ComboboxSelected>>", lambda e: self._refresh_dashboard())

        # Date range filter
        ttk.Label(filter_row, text="Quarter:").pack(side="left", padx=(0, 3))
        self.filter_quarter = ttk.Combobox(filter_row, 
                                           values=["All", "Q1 2025", "Q2 2025", "Q3 2025", "Q4 2025",
                                                   "Q1 2026", "Q2 2026", "Q3 2026", "Q4 2026"],
                                           width=10, font=("Segoe UI", 9), state="readonly")
        self.filter_quarter.set("All")
        self.filter_quarter.pack(side="left", padx=(0, 10))
        self.filter_quarter.bind("<<ComboboxSelected>>", lambda e: self._refresh_dashboard())

        # Project name search
        ttk.Label(filter_row, text="Search:").pack(side="left", padx=(0, 3))
        self.filter_search = ttk.Entry(filter_row, width=20, font=("Segoe UI", 9))
        self.filter_search.pack(side="left", padx=(0, 10))
        self.filter_search.bind("<Return>", lambda e: self._refresh_dashboard())
        self.filter_search.bind("<KeyRelease>", lambda e: self._on_search_keyrelease())

        # Buttons
        ttk.Button(filter_row, text="🔍 Apply", command=self._refresh_dashboard).pack(side="left", padx=3)
        ttk.Button(filter_row, text="🔄 Clear", command=self._clear_filters).pack(side="left", padx=3)
        ttk.Button(filter_row, text="📐 Zoom to Fit", command=self._zoom_to_fit).pack(side="left", padx=3)

        # ══════════════════════════════════════════════════════════════════
        # CHART AREA
        # ══════════════════════════════════════════════════════════════════
        self.chart_frame = ttk.Frame(tab)
        self.chart_frame.pack(fill="both", expand=True)

        # Placeholder for the matplotlib canvas
        self.chart_canvas = None
        self.chart_toolbar = None

        self._refresh_dashboard()

    def _on_search_keyrelease(self):
        """Debounced search - refresh after user stops typing."""
        if hasattr(self, '_search_after_id'):
            self.root.after_cancel(self._search_after_id)
        self._search_after_id = self.root.after(500, self._refresh_dashboard)

    def _zoom_to_fit(self):
        """Reset zoom to show all projects."""
        if self.chart_canvas:
            self._refresh_dashboard()
            self._set_status("Zoomed to fit all projects")

    def _clear_filters(self):
        """Reset all filter dropdowns and refresh."""
        self.filter_status.set("All")
        self.filter_dev_region.set("All")
        self.filter_sales_region.set("All")
        self.filter_quarter.set("All")
        self.filter_search.delete(0, tk.END)
        self._refresh_dashboard()

    def _get_quarter_date_range(self, quarter_str: str) -> tuple:
        """Convert quarter string to date range."""
        if quarter_str == "All" or not quarter_str:
            return None, None
        
        # Parse "Q1 2025" format
        parts = quarter_str.split()
        if len(parts) != 2:
            return None, None
        
        q = parts[0]  # "Q1", "Q2", etc.
        year = int(parts[1])
        
        quarter_starts = {
            "Q1": (1, 1),
            "Q2": (4, 1),
            "Q3": (7, 1),
            "Q4": (10, 1),
        }
        quarter_ends = {
            "Q1": (3, 31),
            "Q2": (6, 30),
            "Q3": (9, 30),
            "Q4": (12, 31),
        }
        
        if q not in quarter_starts:
            return None, None
        
        start_month, start_day = quarter_starts[q]
        end_month, end_day = quarter_ends[q]
        
        start_date = datetime.date(year, start_month, start_day)
        end_date = datetime.date(year, end_month, end_day)
        
        return start_date, end_date

    def _refresh_dashboard(self):
        projects, ref_lines = db.load_all(self.db_path)
        today = datetime.date.today()

        # ══════════════════════════════════════════════════════════════════
        # APPLY FILTERS
        # ══════════════════════════════════════════════════════════════════
        
        # Status filter
        status_filter = self.filter_status.get()
        if status_filter and status_filter != "All":
            status_map = {
                "On Track": "on-track",
                "At Risk": "at-risk",
                "Overdue": "overdue",
            }
            target_status = status_map.get(status_filter)
            if target_status:
                projects = [p for p in projects if p.computed_status(today) == target_status]
        
        # Dev Region filter
        dev_filter = self.filter_dev_region.get()
        if dev_filter and dev_filter != "All":
            projects = [p for p in projects if p.dev_region == dev_filter]
        
        # Sales Region filter
        sales_filter = self.filter_sales_region.get()
        if sales_filter and sales_filter != "All":
            projects = [p for p in projects if p.sales_region == sales_filter]
        
        # Quarter filter
        quarter_filter = self.filter_quarter.get()
        if quarter_filter and quarter_filter != "All":
            q_start, q_end = self._get_quarter_date_range(quarter_filter)
            if q_start and q_end:
                # Show projects that overlap with the quarter
                projects = [p for p in projects 
                           if not (p.end_date < q_start or p.start_date > q_end)]
        
        # Search filter
        search_term = self.filter_search.get().strip().lower()
        if search_term:
            projects = [p for p in projects if search_term in p.name.lower()]

        # ══════════════════════════════════════════════════════════════════
        # UPDATE SUMMARY CARDS
        # ══════════════════════════════════════════════════════════════════
        for w in self.cards_frame.winfo_children():
            w.destroy()

        total = len(projects)
        on_track = sum(1 for p in projects if p.computed_status(today) == "on-track")
        at_risk = sum(1 for p in projects if p.computed_status(today) == "at-risk")
        overdue = sum(1 for p in projects if p.computed_status(today) == "overdue")

        cards_data = [
            ("📁", "Total Projects", str(total), "#0067C0"),
            ("🟢", "On Track", str(on_track), "#16A34A"),
            ("🟡", "At Risk", str(at_risk), "#EAB308"),
            ("🔴", "Overdue", str(overdue), "#DC2626"),
        ]

        for icon, label, value, color in cards_data:
            card = tk.Frame(self.cards_frame, bg="white", relief="flat",
                            bd=1, highlightbackground="#E0E0E0", highlightthickness=1)
            card.pack(side="left", fill="x", expand=True, padx=3, ipady=6, ipadx=10)

            tk.Label(card, text=icon, font=("Segoe UI", 16), bg="white").pack(side="left", padx=(8, 5))
            text_frame = tk.Frame(card, bg="white")
            text_frame.pack(side="left")
            tk.Label(text_frame, text=value, font=("Segoe UI", 18, "bold"),
                     bg="white", fg=color).pack(anchor="w")
            tk.Label(text_frame, text=label, font=("Segoe UI", 8),
                     bg="white", fg="#666666").pack(anchor="w")

        # ══════════════════════════════════════════════════════════════════
        # RENDER CHART
        # ══════════════════════════════════════════════════════════════════
        if not projects:
            for w in self.chart_frame.winfo_children():
                w.destroy()
            self.chart_canvas = None
            self.chart_toolbar = None
            ttk.Label(self.chart_frame, 
                      text="No projects match the current filters.\nAdjust filters or add projects in the Projects tab.",
                      font=("Segoe UI", 14), foreground="#999999",
                      anchor="center", justify="center").pack(expand=True)
            self._set_status("Dashboard refreshed — no matching projects")
            return

        self._render_embedded_chart(projects, ref_lines)
        
        # Build filter info string
        filter_info = ""
        if status_filter != "All":
            filter_info += f" | Status: {status_filter}"
        if dev_filter != "All":
            filter_info += f" | Dev: {dev_filter}"
        if sales_filter != "All":
            filter_info += f" | Sales: {sales_filter}"
        if quarter_filter != "All":
            filter_info += f" | {quarter_filter}"
        if search_term:
            filter_info += f" | Search: '{search_term}'"
            
        self._set_status(f"Dashboard refreshed — {total} project(s) loaded{filter_info}")

    def _render_embedded_chart(self, projects: list[Project], ref_lines: list[ReferenceLine]):
        """Render the full Gantt chart embedded inside the Dashboard tab."""

        # Clean up previous chart
        for w in self.chart_frame.winfo_children():
            w.destroy()

        today = datetime.date.today()
        theme_name = self.theme_var.get()
        theme_cfg = cfg.THEMES.get(theme_name, cfg.THEMES["light"])
        n = len(projects)

        # ── Create figure ────────────────────────────────────────────────────
        MAX_VISIBLE = 6
        fig_h = max(4.0, min(n, MAX_VISIBLE) * 0.9 + 1.5)
        fig, ax = plt.subplots(figsize=(16, fig_h))
        fig.patch.set_facecolor(theme_cfg["bg_color"])
        ax.set_facecolor(theme_cfg["axis_bg"])
        ax.title.set_color(theme_cfg["text_color"])
        ax.xaxis.label.set_color(theme_cfg["text_color"])
        ax.tick_params(colors=theme_cfg["text_color"])
        for spine in ax.spines.values():
            spine.set_color(theme_cfg["grid_color"])

        x_min, x_max = date_range_padded(projects)
        y_min_orig = 0
        y_max_orig = n + 1

        y_positions = []
        y_labels = []
        milestone_artists = []
        project_bars = []  # Store project bar info for click detection

        for idx, project in enumerate(projects):
            y = n - idx
            y_positions.append(y)
            y_labels.append(project.name)
            color = _pick_color(idx, project)
            status_color = cfg.STATUS_COLORS.get(project.computed_status(today), cfg.STATUS_COLORS["on-track"])

            # Full bar (background - lighter) - make it pickable
            bar = ax.barh(y, (project.end_date - project.start_date).days,
                    left=project.start_date, height=cfg.BAR_HEIGHT,
                    color=color, alpha=0.25,
                    edgecolor="none", linewidth=0, picker=True)
            
            # Store project bar info for click detection
            project_bars.append({
                "bar": bar[0],
                "project": project,
                "y": y,
                "y_min": y - cfg.BAR_HEIGHT / 2,
                "y_max": y + cfg.BAR_HEIGHT / 2,
            })

            # Progress overlay (filled portion - darker)
            progress = project.progress(today)
            if progress > 0:
                elapsed_days = (project.end_date - project.start_date).days * progress
                ax.barh(y, elapsed_days, left=project.start_date,
                        height=cfg.BAR_HEIGHT, color=color, alpha=0.85,
                        edgecolor="none", linewidth=0, zorder=3)

            # Phase sub-bars
            for p_idx, phase in enumerate(project.phases):
                phase_color = PHASE_PALETTE[p_idx % len(PHASE_PALETTE)]
                phase_days = (phase.end_date - phase.start_date).days
                ax.barh(y, phase_days, left=phase.start_date,
                        height=cfg.PHASE_HEIGHT, color=phase_color, alpha=0.75,
                        edgecolor="#666666", linewidth=0.5, zorder=3)
                if phase_days > 45:
                    mid = phase.start_date + datetime.timedelta(days=phase_days / 2)
                    ax.text(mid, y - cfg.BAR_HEIGHT / 2 - 0.05, phase.name,
                            fontsize=7, ha="center", va="top",
                            color="#333333", fontweight="bold", zorder=4)

            # Date labels
            ax.text(project.start_date, y - cfg.BAR_HEIGHT / 2 - 0.18,
                    project.start_date.strftime("%b %d, %Y"),
                    fontsize=7, ha="left", va="top",
                    color=theme_cfg["date_label_color"])
            ax.text(project.end_date, y - cfg.BAR_HEIGHT / 2 - 0.18,
                    project.end_date.strftime("%b %d, %Y"),
                    fontsize=7, ha="right", va="top",
                    color=theme_cfg["date_label_color"])

            # Milestones
            for ms_idx, ms in enumerate(project.milestones):
                # 4-color milestone logic based on task status + date proximity
                milestone_color = ms.marker_color(today)
                
                sc = ax.scatter(ms.date, y, marker=cfg.MILESTONE_MARKER,
                                s=cfg.MILESTONE_SIZE, color=milestone_color,
                                edgecolors=cfg.MILESTONE_EDGE_COLOR, 
                                linewidths=0.8, zorder=5, picker=5)
                milestone_artists.append((sc, project.name, ms))

                if ms_idx % 2 == 0:
                    label_y = y + cfg.BAR_HEIGHT * 0.6
                    va = "bottom"
                else:
                    label_y = y - cfg.BAR_HEIGHT * 0.6
                    va = "top"

                ax.text(mdates.date2num(ms.date), label_y, ms.name,
                        fontsize=6, ha="center", va=va,
                        color=theme_cfg["text_color"], fontweight="bold",
                        zorder=7)

        # ── Milestone click handler ──────────────────────────────────────────
        def on_milestone_pick(event):
            """Handle milestone click to show tasks"""
            if event.mouseevent.button != 1:
                return
            
            for sc, proj_name, milestone in milestone_artists:
                if event.artist == sc:
                    _MilestoneTaskDialog(
                        self.root,
                        milestone_name=milestone.name,
                        milestone_id=milestone.milestone_id,
                        project_name=proj_name,
                        milestone_date=milestone.date,
                        can_edit=self.can_edit,
                        username=self.user["username"],
                        db_path=self.db_path
                    )
                    break

        # ── Project bar click handler ────────────────────────────────────────
        def on_project_click(event):
            """Handle project bar click to show project details"""
            if event.inaxes != ax:
                return
            if event.button != 1:  # Left click only
                return
            
            # Check if click is on a project bar
            click_x = mdates.num2date(event.xdata).date() if event.xdata else None
            click_y = event.ydata
            
            if click_x is None or click_y is None:
                return
            
            for bar_info in project_bars:
                project = bar_info["project"]
                y_min = bar_info["y_min"]
                y_max = bar_info["y_max"]
                
                # Check if click is within this project's bar
                if (y_min <= click_y <= y_max and 
                    project.start_date <= click_x <= project.end_date):
                    # Check if we clicked on a milestone (don't open project details)
                    is_milestone_click = False
                    for sc, _, milestone in milestone_artists:
                        contains, _ = sc.contains(event)
                        if contains:
                            is_milestone_click = True
                            break
                    
                    if not is_milestone_click:
                        # Open project details dialog
                        _ProjectDetailsDialog(
                            self.root,
                            project=project,
                            db_path=self.db_path,
                            can_edit=self.can_edit,
                            username=self.user["username"]
                        )
                    break

        # Connect events
        fig.canvas.mpl_connect('pick_event', on_milestone_pick)
        fig.canvas.mpl_connect('button_press_event', on_project_click)

        # Today line
        ax.axvline(today, color=cfg.TODAY_LINE_COLOR,
                linewidth=cfg.TODAY_LINE_WIDTH, linestyle=cfg.TODAY_LINE_STYLE,
                label=f"Today ({today.strftime('%b %d, %Y')})", zorder=4)

        # ── Today date label at top of the vertical line ──────────────────
        ax.text(
            mdates.date2num(today), y_max_orig - 0.05,
            today.strftime("%b %d, %Y"),
            fontsize=7, fontweight="bold",
            color=cfg.TODAY_LINE_COLOR,
            ha="center", va="top",
            zorder=8,
            bbox=dict(boxstyle="round,pad=0.2", facecolor="white",
                      edgecolor=cfg.TODAY_LINE_COLOR, alpha=0.85),
        )

        # Reference lines
        for ref in ref_lines:
            ax.axvline(ref.date, color=ref.color, linewidth=1.8,
                       linestyle=ref.style, zorder=4)
            ax.text(ref.date, y_max_orig - 0.15,
                    f"  {ref.name}\n  {ref.date.strftime('%b %d, %Y')}",
                    fontsize=7, color=ref.color, fontweight="bold",
                    ha="left", va="top")

        # Axes
        ax.set_xlim(x_min, x_max)
        # Show a viewport if there are many projects
        if n > MAX_VISIBLE:
            y_view_min = n - MAX_VISIBLE + 0.5
            y_view_max = n + 1
        else:
            y_view_min = y_min_orig
            y_view_max = y_max_orig
        ax.set_ylim(y_view_min, y_view_max)
        ax.set_yticks(y_positions)
        ax.set_yticklabels([""] * len(y_positions))  # Clear default labels

        # Draw project name + status as custom y-axis labels
        for idx, project in enumerate(projects):
            y = y_positions[idx]
            status_text = project.computed_status(today).replace("-", " ").upper()
            s_color = cfg.STATUS_COLORS.get(project.computed_status(today), cfg.STATUS_COLORS["on-track"])

            # Project name (main label)
            ax.text(-0.01, y + 0.12, project.name,
                    transform=ax.get_yaxis_transform(),
                    fontsize=9, ha="right", va="center",
                    color=theme_cfg["text_color"], fontweight="bold",
                    clip_on=False)
            # Status below the name
            ax.text(-0.01, y - 0.18, f"● {status_text}",
                    transform=ax.get_yaxis_transform(),
                    fontsize=7, ha="right", va="center",
                    color=s_color, fontweight="bold",
                    clip_on=False)
        ax.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
        ax.xaxis.set_minor_locator(mdates.MonthLocator())
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
        fig.autofmt_xdate(rotation=45, ha="right")
        ax.grid(axis="x", linestyle=":", linewidth=0.5, alpha=0.6,
                color=theme_cfg["grid_color"])
        ax.set_axisbelow(True)
        ax.set_title("Project Timelines", fontsize=14, fontweight="bold",
                      pad=15, color=theme_cfg["text_color"], loc="left")

        # Legend
        legend_handles = []
        legend_handles.append(
            plt.Line2D([0], [0], color=cfg.TODAY_LINE_COLOR,
                       linewidth=cfg.TODAY_LINE_WIDTH, linestyle=cfg.TODAY_LINE_STYLE,
                       label=f"Today ({today.strftime('%b %d, %Y')})")
        )
        for ref in ref_lines:
            legend_handles.append(
                plt.Line2D([0], [0], color=ref.color, linewidth=1.8,
                           linestyle=ref.style, label=ref.name)
            )
        for status, scolor in cfg.STATUS_COLORS.items():
            legend_handles.append(
                plt.Line2D([0], [0], marker="s", color="w", markerfacecolor=scolor,
                           markersize=8, label=status.replace("-", " ").title())
            )
        # Milestone color legend
        milestone_legend = [
            ("#27AE60", "Completed"),
            ("#F39C12", "At Risk (≤5 days)"),
            ("#E74C3C", "Overdue"),
            ("#5DADE2", "Upcoming"),
        ]
        for ms_color, ms_label in milestone_legend:
            legend_handles.append(
                plt.Line2D([0], [0], marker=cfg.MILESTONE_MARKER, color="w",
                           markerfacecolor=ms_color, markeredgecolor=cfg.MILESTONE_EDGE_COLOR,
                           markersize=10, label=ms_label)
            )
        ax.legend(handles=legend_handles, loc="upper right", fontsize=8,
                  framealpha=0.9, facecolor=theme_cfg["bg_color"],
                  labelcolor=theme_cfg["text_color"])

        fig.tight_layout()

        # Store figure reference for PDF export
        self._current_figure = fig

        # ── Embed in Tkinter ─────────────────────────────────────────────
        canvas = FigureCanvasTkAgg(fig, master=self.chart_frame)
        canvas_widget = canvas.get_tk_widget()
        canvas_widget.pack(fill="both", expand=True)

        toolbar = NavigationToolbar2Tk(canvas, self.chart_frame)
        toolbar.update()
        toolbar.pack(fill="x")

        self.chart_canvas = canvas
        self.chart_toolbar = toolbar

        # ── Hover tooltip ────────────────────────────────────────────
        tooltip = ax.annotate(
            "", xy=(0, 0), xytext=(15, 15),
            textcoords="offset points",
            fontsize=9, fontweight="bold",
            color=theme_cfg["tooltip_text"],
            bbox=dict(boxstyle="round,pad=0.5", facecolor=theme_cfg["tooltip_bg"],
                    edgecolor="grey", alpha=0.95),
            arrowprops=dict(arrowstyle="->", color=theme_cfg["tooltip_bg"], lw=1.5),
            zorder=10, visible=False,
        )

        def on_hover(event):
            if event.inaxes != ax:
                if tooltip.get_visible():
                    tooltip.set_visible(False)
                    canvas.draw_idle()
                return
            found = False
            for sc, proj_name, milestone in milestone_artists:
                contains, _ = sc.contains(event)
                if contains:
                    # Build tooltip text with task status summary
                    tooltip_text = f"{proj_name}\n{milestone.name}: {milestone.date.strftime('%b %d, %Y')}"
                    
                    # Add task summary if tasks exist
                    if milestone.task_statuses:
                        total = sum(milestone.task_statuses.values())
                        if total > 0:
                            tooltip_text += f"\n\n📊 Task Summary ({total} total):"
                            
                            completed = milestone.task_statuses.get("Completed", 0)
                            wip = milestone.task_statuses.get("WIP", 0)
                            yet_to_start = milestone.task_statuses.get("Yet to Start", 0)
                            not_applicable = milestone.task_statuses.get("Not Applicable", 0)
                            
                            if completed > 0:
                                tooltip_text += f"\n  ✅ Completed: {completed}"
                            if wip > 0:
                                tooltip_text += f"\n  🔄 WIP: {wip}"
                            if yet_to_start > 0:
                                tooltip_text += f"\n  ⏳ Yet to Start: {yet_to_start}"
                            if not_applicable > 0:
                                tooltip_text += f"\n  ➖ Not Applicable: {not_applicable}"
                            
                            # Add completion percentage
                            completion_pct = int((completed / total) * 100) if total > 0 else 0
                            tooltip_text += f"\n\n📈 Progress: {completion_pct}%"
                    
                    tooltip.xy = (mdates.date2num(milestone.date), sc.get_offsets()[0][1])
                    tooltip.set_text(tooltip_text)
                    tooltip.set_visible(True)
                    found = True
                    break
            if not found and tooltip.get_visible():
                tooltip.set_visible(False)
            canvas.draw_idle()

        # ── Y-axis-only scroll (X-axis stays fixed) ─────────────────────
        Y_SCROLL_STEP = 0.4  # how many y-units per scroll tick
        x_min_num = mdates.date2num(x_min)
        x_max_num = mdates.date2num(x_max)

        def on_scroll(event):
            if event.inaxes != ax:
                return
            cur_y_min, cur_y_max = ax.get_ylim()
            view_height = cur_y_max - cur_y_min

            if event.button == "up":
                # Scroll up → show higher-numbered projects (move viewport up)
                new_y_max = min(cur_y_max + Y_SCROLL_STEP, y_max_orig)
                new_y_min = new_y_max - view_height
                if new_y_min < y_min_orig:
                    new_y_min = y_min_orig
                    new_y_max = new_y_min + view_height
            elif event.button == "down":
                # Scroll down → show lower-numbered projects (move viewport down)
                new_y_min = max(cur_y_min - Y_SCROLL_STEP, y_min_orig)
                new_y_max = new_y_min + view_height
                if new_y_max > y_max_orig:
                    new_y_max = y_max_orig
                    new_y_min = new_y_max - view_height
            else:
                return

            ax.set_ylim(new_y_min, new_y_max)
            canvas.draw_idle()

        # ── Middle-click reset ───────────────────────────────────────────
        def on_click(event):
            if event.inaxes != ax:
                return
            if event.button == 2:
                ax.set_xlim(x_min_num, x_max_num)
                ax.set_ylim(y_min_orig, y_max_orig)
                canvas.draw_idle()

        canvas.mpl_connect("motion_notify_event", on_hover)
        canvas.mpl_connect("scroll_event", on_scroll)
        canvas.mpl_connect("button_press_event", on_click)
    
    def _popout_timeline(self):
        """Open the chart in a separate Matplotlib window."""
        projects, ref_lines = db.load_all(self.db_path)
        if not projects:
            messagebox.showinfo("Empty", "No projects in the database.")
            return
        render_timeline(
            projects=projects, today=datetime.date.today(),
            title="Project Timelines", reference_lines=ref_lines,
            theme=self.theme_var.get(), show=True,
        )

    def _save_timeline_png(self):
        projects, ref_lines = db.load_all(self.db_path)
        if not projects:
            messagebox.showinfo("Empty", "No projects.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".png",
            filetypes=[("PNG", "*.png"), ("SVG", "*.svg"), ("PDF", "*.pdf")])
        if path:
            render_timeline(projects=projects, today=datetime.date.today(),
                            title="Project Timelines", reference_lines=ref_lines,
                            theme=self.theme_var.get(),
                            output_path=path, show=False)
            self._set_status(f"Saved to {path}")
            messagebox.showinfo("Saved", f"Chart saved to:\n{path}")

    def _save_timeline_html(self):
        projects, ref_lines = db.load_all(self.db_path)
        if not projects:
            messagebox.showinfo("Empty", "No projects.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".html", filetypes=[("HTML", "*.html")])
        if path:
            render_timeline(projects=projects, today=datetime.date.today(),
                            title="Project Timelines", reference_lines=ref_lines,
                            theme=self.theme_var.get(),
                            html_path=path, show=False)
            self._set_status(f"HTML saved to {path}")
            messagebox.showinfo("Saved", f"Interactive chart saved to:\n{path}")

    def _export_pdf_report(self):
        """Export a professional PDF report."""
        if not REPORTLAB_AVAILABLE:
            messagebox.showerror(
                "Missing Dependency",
                "reportlab is required for PDF export.\n\nInstall with:\npip install reportlab"
            )
            return
        
        # Ask for save location - use initialfile, not initialname
        filepath = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            title="Save PDF Report",
            initialfile=f"project_report_{datetime.date.today().strftime('%Y%m%d')}.pdf"  # FIXED
        )
        
        if not filepath:
            return
        
        try:
            # Save current Gantt chart as temp image
            gantt_path = None
            if hasattr(self, '_current_figure') and self._current_figure:
                gantt_path = pathlib.Path(filepath).parent / "_temp_gantt.png"
                self._current_figure.savefig(str(gantt_path), dpi=150, bbox_inches='tight')
            
            # Load projects
            projects, _ = db.load_all(self.db_path)
            
            # Generate PDF
            generate_pdf_report(
                projects=projects,
                output_path=filepath,
                title="Project Status Report",
                include_gantt=True,
                include_milestones=True,
                include_kpis=True,
                gantt_image_path=gantt_path,
            )
            
            # Clean up temp file
            if gantt_path and gantt_path.exists():
                gantt_path.unlink()
            
            self._set_status(f"PDF report saved to {filepath}")
            messagebox.showinfo("Export Complete", f"PDF report saved successfully!\n\n{filepath}")
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to generate PDF report:\n\n{e}")

    def _export_excel_report(self):
        """Export an Excel report with multiple sheets."""
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror(
                "Missing Dependency",
                "openpyxl is required for Excel export.\n\nInstall with:\npip install openpyxl"
            )
            return
        
        # Use initialfile, not initialname
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save Excel Report",
            initialfile=f"project_report_{datetime.date.today().strftime('%Y%m%d')}.xlsx"  # FIXED
        )
        
        if not filepath:
            return
        
        try:
            projects, _ = db.load_all(self.db_path)
            
            generate_excel_report(
                projects=projects,
                output_path=filepath,
                title="Project Status Report",
            )
            
            self._set_status(f"Excel report saved to {filepath}")
            messagebox.showinfo("Export Complete", f"Excel report saved successfully!\n\n{filepath}")
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to generate Excel report:\n\n{e}")


    # ═════════════════════════════════════════════════════════════════════
    # TAB 2 — QCTP (Quality, Cost, Time, Performance)
    # ═════════════════════════════════════════════════════════════════════

    def _build_qctp_tab(self):
        tab = self.tab_qctp

        # Load QCTP template from Excel on startup
        try:
            from timeline_tool.qctp_template import get_qctp_template
            self._qctp_template = get_qctp_template()
        except Exception as e:
            print(f"Warning: Could not load QCTP template: {e}")
            self._qctp_template = {}

        # Initialize current week number from SYSTEM DATE
        self._current_week = datetime.date.today().isocalendar()[1]
        self._current_year = datetime.date.today().year

        # Cache for project list - MUST be initialized before any method calls
        self._qctp_projects_cache = []

        # ── Header ───────────────────────────────────────────────────────
        header = ttk.Frame(tab)
        header.pack(fill="x", pady=(0, 10))
        ttk.Label(header, text="QCTP Overview", style="Header.TLabel").pack(side="left")
        
        # Legend
        legend_frame = ttk.Frame(header)
        legend_frame.pack(side="right")
        ttk.Label(legend_frame, text="Status: ", font=("Segoe UI", 9)).pack(side="left")
        tk.Label(legend_frame, text="● Green", fg="#16A34A", bg="#F3F3F3", 
                 font=("Segoe UI", 9, "bold")).pack(side="left", padx=5)
        tk.Label(legend_frame, text="● Orange", fg="#F97316", bg="#F3F3F3",
                 font=("Segoe UI", 9, "bold")).pack(side="left", padx=5)
        tk.Label(legend_frame, text="● Red", fg="#DC2626", bg="#F3F3F3",
                 font=("Segoe UI", 9, "bold")).pack(side="left", padx=5)

        # ── Project selector ─────────────────────────────────────────────
        selector_frame = ttk.LabelFrame(tab, text="  Select Project  ", padding=8)
        selector_frame.pack(fill="x", pady=(0, 10))

        sel_inner = ttk.Frame(selector_frame)
        sel_inner.pack(fill="x")

        ttk.Label(sel_inner, text="Project:", style="Card.TLabel").pack(side="left", padx=(0, 8))
        self.qctp_project_combo = ttk.Combobox(sel_inner, width=40, state="readonly",
                                                 font=("Segoe UI", 10))
        self.qctp_project_combo.pack(side="left", padx=(0, 10))
        self.qctp_project_combo.bind("<<ComboboxSelected>>", self._on_qctp_project_change)
        ttk.Button(sel_inner, text="🔄 Refresh", command=self._refresh_qctp_project_list).pack(side="left")

        if self.can_edit:
            # Save All button - styled blue
            save_btn = tk.Button(sel_inner, text="💾 Save All", font=("Segoe UI", 10, "bold"),
                                  bg="#0067C0", fg="white", relief="flat", cursor="hand2",
                                  activebackground="#005A9E", activeforeground="white",
                                  command=self._save_qctp, padx=15, pady=5)
            save_btn.pack(side="right", padx=5)
            save_btn.bind("<Enter>", lambda e: save_btn.config(bg="#005A9E"))
            save_btn.bind("<Leave>", lambda e: save_btn.config(bg="#0067C0"))

        # ── Phase Sub-Tabs ───────────────────────────────────────────────
        self.qctp_notebook = ttk.Notebook(tab)
        self.qctp_notebook.pack(fill="both", expand=True, pady=(10, 0))

        # Define the 3 phases
        self.qctp_phases = [
            ("pre_program", "Pre-Program", "Before CM"),
            ("detailed_design", "Detailed Design Phase", "CM to Sync 5/SHRM"),
            ("industrialization", "Industrialization Phase", "Sync5 to SOP"),
        ]

        # Define QCTP categories
        self.qctp_categories = [
            ("quality", "Quality", "🎯"),
            ("cost", "Cost", "💰"),
            ("time", "Time", "⏱️"),
            ("performance", "Performance", "📈"),
        ]

        # Store widgets for each phase/category/line item
        self.qctp_line_widgets = {}
        
        # CRITICAL: Initialize notes widgets BEFORE _refresh_qctp_project_list is called
        self.qctp_notes_widgets = {}
        
        # Store week label references for updating
        self._week_labels = []

        for phase_key, phase_name, phase_desc in self.qctp_phases:
            phase_tab = ttk.Frame(self.qctp_notebook, padding=5)
            self.qctp_notebook.add(phase_tab, text=f"  {phase_name}  ")
            
            # Phase description header with week navigation
            desc_frame = ttk.Frame(phase_tab)
            desc_frame.pack(fill="x", pady=(0, 8))
            
            # Left side: phase name
            left_desc = ttk.Frame(desc_frame)
            left_desc.pack(side="left")
            ttk.Label(left_desc, text=f"📋 {phase_name}", 
                      font=("Segoe UI", 12, "bold")).pack(side="left")
            ttk.Label(left_desc, text=f"  ({phase_desc})", 
                      font=("Segoe UI", 10, "italic"), foreground="#666666").pack(side="left")

            # Right side: Week navigation
            week_frame = ttk.Frame(desc_frame)
            week_frame.pack(side="right")
            
            # Previous Week button
            prev_btn = tk.Button(week_frame, text="◀ Previous Week", font=("Segoe UI", 9),
                                  bg="#F3F3F3", fg="#333333", relief="flat", cursor="hand2",
                                  command=lambda: self._change_week(-1), padx=10, pady=3)
            prev_btn.pack(side="left", padx=2)
            
            # Current Week indicator (blue button style) - gets current week from system
            week_label = tk.Label(week_frame, text=f"Week {self._current_week}",
                                   font=("Segoe UI", 10, "bold"),
                                   bg="#0067C0", fg="white", padx=15, pady=5)
            week_label.pack(side="left", padx=2)
            self._week_labels.append(week_label)
            
            # Next Week button
            next_btn = tk.Button(week_frame, text="Next Week ▶", font=("Segoe UI", 9),
                                  bg="#F3F3F3", fg="#333333", relief="flat", cursor="hand2",
                                  command=lambda: self._change_week(1), padx=10, pady=3)
            next_btn.pack(side="left", padx=2)

            # Main content area - horizontal split
            main_content = ttk.Frame(phase_tab)
            main_content.pack(fill="both", expand=True)

            # Left side: Scrollable QCTP grid
            left_frame = ttk.Frame(main_content)
            left_frame.pack(side="left", fill="both", expand=True)
            
            # Create scrollable frame for this phase
            phase_canvas = tk.Canvas(left_frame, bg="#F3F3F3", highlightthickness=0)
            phase_scrollbar = ttk.Scrollbar(left_frame, orient="vertical", command=phase_canvas.yview)
            phase_scrollable = ttk.Frame(phase_canvas)
            
            phase_scrollable.bind(
                "<Configure>",
                lambda e, c=phase_canvas: c.configure(scrollregion=c.bbox("all"))
            )
            
            phase_canvas.create_window((0, 0), window=phase_scrollable, anchor="nw")
            phase_canvas.configure(yscrollcommand=phase_scrollbar.set)
            
            phase_canvas.pack(side="left", fill="both", expand=True)
            phase_scrollbar.pack(side="right", fill="y")

            # 2×2 Grid for QCTP categories
            grid_frame = ttk.Frame(phase_scrollable)
            grid_frame.pack(fill="both", expand=True, padx=5, pady=5)
            
            grid_frame.grid_columnconfigure(0, weight=1, uniform="col")
            grid_frame.grid_columnconfigure(1, weight=1, uniform="col")

            self.qctp_line_widgets[phase_key] = {}

            # Create each category box
            positions = [(0, 0), (0, 1), (1, 0), (1, 1)]
            for (cat_key, cat_name, cat_icon), (row, col) in zip(self.qctp_categories, positions):
                category_frame = self._create_qctp_category_box(
                    grid_frame, phase_key, cat_key, cat_name, cat_icon
                )
                category_frame.grid(row=row, column=col, padx=5, pady=5, sticky="nsew")
                grid_frame.grid_rowconfigure(row, weight=1)

            # Right side: Notes panels (Highlights, Red Points, VSSQ Escalation)
            right_frame = ttk.Frame(main_content, width=600)
            right_frame.pack(side="right", fill="both", padx=(10, 0))
            right_frame.pack_propagate(False)  # Maintain fixed width
            
            self.qctp_notes_widgets[phase_key] = {}
            
            # Highlights panel
            highlights_frame = tk.LabelFrame(right_frame, text="HIGHLIGHTS (Focus on main difficulties all phases)",
                                              font=("Segoe UI", 9, "bold"), bg="white", fg="#1A1A1A")
            highlights_frame.pack(fill="both", expand=True, pady=(0, 5))
            
            highlights_text = tk.Text(highlights_frame, font=("Segoe UI", 9), wrap="word",
                                    height=8, bg="white", relief="flat", padx=5, pady=5)
            highlights_text.pack(fill="both", expand=True, padx=2, pady=2)
            if not self.can_edit:
                highlights_text.config(state="disabled")
            self.qctp_notes_widgets[phase_key]["highlights"] = highlights_text
            
            # Red Points panel
            red_points_frame = tk.LabelFrame(right_frame, text="RED points explanation",
                                              font=("Segoe UI", 9, "bold"), bg="white", fg="#DC2626")
            red_points_frame.pack(fill="both", expand=True, pady=5)
            
            red_points_text = tk.Text(red_points_frame, font=("Segoe UI", 9), wrap="word",
                                    height=7, bg="white", relief="flat", padx=5, pady=5)
            red_points_text.pack(fill="both", expand=True, padx=2, pady=2)
            if not self.can_edit:
                red_points_text.config(state="disabled")
            self.qctp_notes_widgets[phase_key]["red_points"] = red_points_text
            
            # VSSQ Escalation panel
            escalation_frame = tk.LabelFrame(right_frame, text="VSSQ Escalation",
                                              font=("Segoe UI", 9, "bold"), bg="white", fg="#1A1A1A")
            escalation_frame.pack(fill="both", expand=True, pady=(5, 0))
            
            escalation_text = tk.Text(escalation_frame, font=("Segoe UI", 9), wrap="word",
                                    height=7, bg="white", relief="flat", padx=5, pady=5)
            escalation_text.pack(fill="both", expand=True, padx=2, pady=2)
            if not self.can_edit:
                escalation_text.config(state="disabled")
            self.qctp_notes_widgets[phase_key]["escalation"] = escalation_text

        self._refresh_qctp_project_list()

    def _change_week(self, delta: int):
        """Change the current week by delta (-1 for previous, +1 for next)."""
        self._current_week += delta
        
        # Handle year transitions
        if self._current_week < 1:
            self._current_year -= 1
            self._current_week = 52  # Approximate
        elif self._current_week > 52:
            self._current_year += 1
            self._current_week = 1
        
        # Update all week labels
        if hasattr(self, '_week_labels'):
            for label in self._week_labels:
                label.config(text=f"Week {self._current_week}")
        
        # Reload data for the new week
        self._on_qctp_project_change()
        self._set_status(f"Switched to Week {self._current_week}, {self._current_year}")

    def _get_qctp_selected_project_id(self) -> int | None:
        """Get the currently selected project ID from the QCTP combo box."""
        idx = self.qctp_project_combo.current()
        if idx < 0 or not self._qctp_projects_cache:
            return None
        return self._qctp_projects_cache[idx]["id"]

    def _refresh_qctp_project_list(self):
        """Refresh the project dropdown list."""
        with db._connect(self.db_path) as conn:
            rows = conn.execute("SELECT id, name FROM projects ORDER BY start_date").fetchall()
        self._qctp_projects_cache = list(rows)
        self.qctp_project_combo["values"] = [f"{r['id']} — {r['name']}" for r in rows]
        if rows:
            self.qctp_project_combo.current(0)
            self._on_qctp_project_change()

    def _create_qctp_category_box(self, parent, phase_key: str, cat_key: str, 
                                   cat_name: str, cat_icon: str) -> tk.Frame:
        """Create a QCTP category box with DYNAMIC line items based on Excel data."""
        
        # Get descriptions from template - filter out empty ones
        all_descriptions = self._qctp_template.get(phase_key, {}).get(cat_key, [])
        
        # Filter to only non-empty descriptions
        descriptions_with_index = [
            (idx + 1, desc.strip()) 
            for idx, desc in enumerate(all_descriptions) 
            if desc and desc.strip()
        ]
        
        # Main card frame
        card = tk.Frame(parent, bg="white", bd=1,
                        highlightbackground="#CCCCCC", highlightthickness=1)
        
        # Header - NO COUNT NUMBER (just icon + category name)
        header = tk.Frame(card, bg="#E8E8E8")
        header.pack(fill="x")
        
        header_text = f"{cat_icon} {cat_name}"  # No count number
        
        tk.Label(header, text=header_text, font=("Segoe UI", 11, "bold"),
                 bg="#E8E8E8", fg="#1A1A1A", anchor="w", padx=10, pady=8).pack(fill="x")
        
        # Separator
        tk.Frame(card, bg="#CCCCCC", height=1).pack(fill="x")
        
        # Content area
        content = tk.Frame(card, bg="white", padx=8, pady=8)
        content.pack(fill="both", expand=True)
        
        # Initialize line widgets storage
        self.qctp_line_widgets[phase_key][cat_key] = []
        
        if not descriptions_with_index:
            # No data - show placeholder
            tk.Label(content, text="No items defined", font=("Segoe UI", 9, "italic"),
                     bg="white", fg="#999999", pady=20).pack()
            return card
        
        # Column headers
        header_frame = tk.Frame(content, bg="white")
        header_frame.pack(fill="x", pady=(0, 5))
        
        tk.Label(header_frame, text="#", font=("Segoe UI", 8, "bold"), 
                 bg="white", fg="#666666", width=2).pack(side="left")
        tk.Label(header_frame, text="Description", font=("Segoe UI", 8, "bold"),
                 bg="white", fg="#666666", width=38, anchor="w").pack(side="left", padx=(5, 0))
        tk.Label(header_frame, text="Status", font=("Segoe UI", 8, "bold"),
                 bg="white", fg="#666666", width=8).pack(side="left", padx=(5, 0))
        tk.Label(header_frame, text="Remarks", font=("Segoe UI", 8, "bold"),
                 bg="white", fg="#666666", width=18, anchor="w").pack(side="left", padx=(5, 0))
        tk.Label(header_frame, text="Attach", font=("Segoe UI", 8, "bold"),
                 bg="white", fg="#666666", width=6).pack(side="left", padx=(5, 0))
        
        # Separator
        tk.Frame(content, bg="#E0E0E0", height=1).pack(fill="x", pady=3)
        
        # Create only rows that have data
        for display_num, (original_line_num, description) in enumerate(descriptions_with_index, 1):
            line_widgets = self._create_qctp_line_item(
                content, phase_key, cat_key, display_num, original_line_num, description
            )
            self.qctp_line_widgets[phase_key][cat_key].append(line_widgets)
        
        return card

    def _create_qctp_line_item(self, parent, phase_key: str, cat_key: str, 
                                display_num: int, original_line_num: int, 
                                description: str) -> dict:
        """Create a single QCTP line item row with READ-ONLY description from template.
        Status dropdown starts BLANK (empty).
        """
        
        row_frame = tk.Frame(parent, bg="white")
        row_frame.pack(fill="x", pady=3)
        
        # Line number (display number)
        tk.Label(row_frame, text=str(display_num), font=("Segoe UI", 9, "bold"),
                 bg="white", fg="#0067C0", width=2).pack(side="left")
        
        # Description - READ-ONLY Label with word wrap for multi-line
        lines = description.split('\n')
        if len(lines) > 1 or len(description) > 55:
            # Multi-line description - use Text widget (read-only)
            desc_frame = tk.Frame(row_frame, bg="#F0F0F0", relief="groove", bd=1)
            desc_frame.pack(side="left", padx=(5, 0), fill="x")
            
            num_lines = len(lines)
            max_line_len = max(len(line) for line in lines) if lines else 0
            wrapped_lines = (max_line_len // 45) + 1
            height = min(3, max(1, num_lines + wrapped_lines - 1))
            
            desc_text = tk.Text(desc_frame, font=("Segoe UI", 8), width=42, height=height,
                                bg="#F0F0F0", fg="#333333", relief="flat", wrap="word",
                                padx=3, pady=2)
            desc_text.insert("1.0", description)
            desc_text.config(state="disabled")
            desc_text.pack(fill="both", expand=True)
            desc_widget = desc_text
        else:
            # Single line - use Label
            display_desc = description[:55] + "..." if len(description) > 55 else description
            desc_label = tk.Label(row_frame, text=display_desc, font=("Segoe UI", 8),
                                  bg="#F0F0F0", fg="#333333", anchor="w", width=42,
                                  relief="groove", padx=5, pady=3)
            desc_label.pack(side="left", padx=(5, 0))
            desc_widget = desc_label
            
            # Tooltip for truncated descriptions
            if len(description) > 55:
                desc_label.bind("<Enter>", lambda e, d=description: self._show_tooltip(e, d))
                desc_label.bind("<Leave>", lambda e: self._hide_tooltip())
        
        # Status dropdown - STARTS BLANK (empty string)
        status_var = tk.StringVar(value="")  # Empty = blank on startup
        status_combo = ttk.Combobox(row_frame, textvariable=status_var,
                                     values=["", "Green", "Orange", "Red"],  # Blank as first option
                                     width=6, state="readonly" if self.can_edit else "disabled",
                                     font=("Segoe UI", 8))
        status_combo.pack(side="left", padx=(5, 0))
        status_combo.set("")  # Explicitly set to blank
        
        # Status color indicator - STARTS GRAY (no selection)
        status_indicator = tk.Label(row_frame, text="●", font=("Segoe UI", 10),
                                     bg="white", fg="#CCCCCC")  # Gray = unset
        status_indicator.pack(side="left", padx=(2, 0))
        
        # Bind status change to update color indicator
        def on_status_change(*args):
            colors = {
                "": "#CCCCCC",        # Gray for blank
                "Green": "#16A34A", 
                "Orange": "#F97316", 
                "Red": "#DC2626"
            }
            color = colors.get(status_var.get(), "#CCCCCC")
            status_indicator.config(fg=color)
        
        status_var.trace_add("write", on_status_change)
        
        # Remarks entry (editable)
        remarks_entry = tk.Entry(row_frame, font=("Segoe UI", 8), width=18,
                                  bg="#FAFAFA", relief="solid", bd=1)
        remarks_entry.pack(side="left", padx=(5, 0), ipady=2)
        
        if not self.can_edit:
            remarks_entry.config(state="disabled")
        
        # Attachment button and indicator
        attach_frame = tk.Frame(row_frame, bg="white")
        attach_frame.pack(side="left", padx=(5, 0))
        
        attachment_path = tk.StringVar(value="")
        
        # Create attach_label first
        attach_label = tk.Label(attach_frame, text="", font=("Segoe UI", 8),
                                 bg="white", fg="#666666")
        
        def browse_attachment():
            filepath = filedialog.askopenfilename(
                title=f"Select attachment for {cat_key.title()} Item {display_num}",
                filetypes=[
                    ("All files", "*.*"),
                    ("PDF files", "*.pdf"),
                    ("Word files", "*.docx"),
                    ("Excel files", "*.xlsx"),
                    ("Images", "*.png *.jpg *.jpeg")
                ]
            )
            if filepath:
                attachment_path.set(filepath)
                attach_label.config(text="📎", fg="#16A34A")
                if self.can_edit:
                    attach_btn.config(text="✓")
        
        def view_attachment():
            path = attachment_path.get()
            if path and pathlib.Path(path).exists():
                import os
                import platform
                if platform.system() == 'Windows':
                    os.startfile(path)
                elif platform.system() == 'Darwin':
                    os.system(f'open "{path}"')
                else:
                    os.system(f'xdg-open "{path}"')
            elif path:
                messagebox.showwarning("Not Found", "Attachment file not found.")
            else:
                messagebox.showinfo("No Attachment", "No attachment for this item.")
        
        def clear_attachment():
            attachment_path.set("")
            attach_label.config(text="", fg="#666666")
            if self.can_edit:
                attach_btn.config(text="📎")
        
        if self.can_edit:
            attach_btn = tk.Button(attach_frame, text="📎", font=("Segoe UI", 8),
                                    bg="white", fg="#666666", relief="flat",
                                    bd=0, cursor="hand2", width=2, pady=0,
                                    command=browse_attachment)
            attach_btn.pack(side="left")
            
            # Right-click menu for attachment
            attach_menu = tk.Menu(attach_btn, tearoff=0)
            attach_menu.add_command(label="Browse...", command=browse_attachment)
            attach_menu.add_command(label="View", command=view_attachment)
            attach_menu.add_separator()
            attach_menu.add_command(label="Clear", command=clear_attachment)
            
            def show_attach_menu(event):
                attach_menu.tk_popup(event.x_root, event.y_root)
            
            attach_btn.bind("<Button-3>", show_attach_menu)
        else:
            attach_btn = tk.Button(attach_frame, text="👁", font=("Segoe UI", 8),
                                    bg="white", fg="#666666", relief="flat",
                                    cursor="hand2", width=2, pady=0,
                                    command=view_attachment)
            attach_btn.pack(side="left")
        
        attach_label.pack(side="left")
        
        return {
            "display_num": display_num,
            "line_num": original_line_num,
            "description": description,
            "desc_widget": desc_widget,
            "status_var": status_var,
            "status_combo": status_combo,
            "status_indicator": status_indicator,
            "remarks_entry": remarks_entry,
            "attachment_path": attachment_path,
            "attach_btn": attach_btn,
            "attach_label": attach_label,
        }

    def _show_tooltip(self, event, text):
        """Show tooltip with full description."""
        if hasattr(self, '_tooltip') and self._tooltip:
            self._tooltip.destroy()
        
        x = event.widget.winfo_rootx() + 10
        y = event.widget.winfo_rooty() + event.widget.winfo_height() + 5
        
        self._tooltip = tk.Toplevel(self.root)
        self._tooltip.wm_overrideredirect(True)
        self._tooltip.wm_geometry(f"+{x}+{y}")
        
        label = tk.Label(self._tooltip, text=text, font=("Segoe UI", 9),
                         bg="#FFFDE7", fg="#333333", relief="solid", bd=1,
                         padx=8, pady=4, wraplength=350, justify="left")
        label.pack()

    def _hide_tooltip(self):
        """Hide the tooltip."""
        if hasattr(self, '_tooltip') and self._tooltip:
            self._tooltip.destroy()
            self._tooltip = None

    def _on_qctp_project_change(self, event=None):
        """Load QCTP data when project selection changes."""
        pid = self._get_qctp_selected_project_id()
        if pid is None:
            return
        
        for phase_key, _, _ in self.qctp_phases:
            for cat_key, _, _ in self.qctp_categories:
                try:
                    line_items = db.get_qctp_line_items(pid, phase_key, cat_key, self.db_path)
                except Exception:
                    line_items = []
                
                items_by_line = {item["line_number"]: item for item in line_items}
                
                for widget_data in self.qctp_line_widgets[phase_key][cat_key]:
                    line_num = widget_data["line_num"]
                    item = items_by_line.get(line_num, {})
                    
                    # Update status - ALWAYS default to "" (blank) if not explicitly set
                    saved_status = item.get("status", "")
                    if not item or saved_status == "":
                        status = ""
                    else:
                        status = saved_status
                    
                    widget_data["status_var"].set(status)
                    widget_data["status_combo"].set(status)  # Explicitly set combo too
                    
                    # Manually update the indicator color
                    colors = {
                        "": "#CCCCCC",
                        "Green": "#16A34A", 
                        "Orange": "#F97316", 
                        "Red": "#DC2626"
                    }
                    widget_data["status_indicator"].config(fg=colors.get(status, "#CCCCCC"))
                    
                    # Update remarks
                    remarks_entry = widget_data["remarks_entry"]
                    was_disabled = (remarks_entry.cget("state") == "disabled")
                    if was_disabled:
                        remarks_entry.config(state="normal")
                    remarks_entry.delete(0, tk.END)
                    remarks_entry.insert(0, item.get("remarks", ""))
                    if was_disabled:
                        remarks_entry.config(state="disabled")
                    
                    # Update attachment
                    attachment = item.get("attachment_path", "")
                    widget_data["attachment_path"].set(attachment)
                    if attachment:
                        widget_data["attach_label"].config(text="📎", fg="#16A34A")
                        if self.can_edit:
                            widget_data["attach_btn"].config(text="✓")
                    else:
                        widget_data["attach_label"].config(text="", fg="#666666")
                        if self.can_edit:
                            widget_data["attach_btn"].config(text="📎")
            
            # Load notes (highlights, red_points, escalation) for this phase
            if phase_key in self.qctp_notes_widgets:
                try:
                    notes = db.get_qctp_notes(pid, phase_key, self._current_week, self._current_year, self.db_path)
                except Exception:
                    notes = {}
                
                for note_key in ["highlights", "red_points", "escalation"]:
                    text_widget = self.qctp_notes_widgets[phase_key].get(note_key)
                    if text_widget:
                        was_disabled = (text_widget.cget("state") == "disabled")
                        if was_disabled:
                            text_widget.config(state="normal")
                        text_widget.delete("1.0", tk.END)
                        text_widget.insert("1.0", notes.get(note_key, ""))
                        if was_disabled:
                            text_widget.config(state="disabled")

    def _save_qctp(self):
        """Save all QCTP line items and notes for the selected project."""
        pid = self._get_qctp_selected_project_id()
        if pid is None:
            messagebox.showwarning("Select", "Select a project first.")
            return
        
        try:
            saved_count = 0
            
            for phase_key, _, _ in self.qctp_phases:
                for cat_key, _, _ in self.qctp_categories:
                    for widget_data in self.qctp_line_widgets[phase_key][cat_key]:
                        line_num = widget_data["line_num"]
                        description = widget_data["description"]
                        status = widget_data["status_var"].get()
                        remarks = widget_data["remarks_entry"].get().strip()
                        attachment = widget_data["attachment_path"].get()
                        
                        db.save_qctp_line_item(
                            project_id=pid,
                            phase=phase_key,
                            category=cat_key,
                            line_number=line_num,
                            description=description,
                            status=status,
                            remarks=remarks,
                            attachment_path=attachment,
                            username=self.user["username"],
                            db_path=self.db_path
                        )
                        saved_count += 1
                
                # Save notes for this phase
                if phase_key in self.qctp_notes_widgets:
                    notes_data = {}
                    for note_key in ["highlights", "red_points", "escalation"]:
                        text_widget = self.qctp_notes_widgets[phase_key].get(note_key)
                        if text_widget:
                            notes_data[note_key] = text_widget.get("1.0", tk.END).strip()
                    
                    db.save_qctp_notes(
                        project_id=pid,
                        phase=phase_key,
                        week_number=self._current_week,
                        year=self._current_year,
                        highlights=notes_data.get("highlights", ""),
                        red_points=notes_data.get("red_points", ""),
                        escalation=notes_data.get("escalation", ""),
                        username=self.user["username"],
                        db_path=self.db_path
                    )
            
            # Log the action
            with db._connect(self.db_path) as conn:
                db.log_action(conn, self.user["username"], "SAVE_QCTP",
                              f"Saved {saved_count} QCTP line items for project ID {pid}, Week {self._current_week}")
            
            self._set_status(f"QCTP saved successfully ({saved_count} items, Week {self._current_week})")
            messagebox.showinfo("Saved", f"QCTP data saved successfully!\n\n{saved_count} line items saved for Week {self._current_week}.")
            
        except Exception as e:
            messagebox.showerror("Save Error", f"Failed to save QCTP data:\n\n{str(e)}")


    # ═════════════════════════════════════════════════════════════════════
    # TAB — ACTIVITIES
    # ═════════════════════════════════════════════════════════════════════

    def _build_activities_tab(self):
        """Build the Activities tab for weekly activity tracking."""
        tab = self.tab_activities

        # ── Header ───────────────────────────────────────────────────────
        header = ttk.Frame(tab)
        header.pack(fill="x", pady=(0, 10))
        ttk.Label(header, text="Weekly Activities", style="Header.TLabel").pack(side="left")

        # ── Project + Week Selector ──────────────────────────────────────
        selector_frame = ttk.LabelFrame(tab, text="  Select Project & Week  ", padding=8)
        selector_frame.pack(fill="x", pady=(0, 10))

        sel_inner = ttk.Frame(selector_frame)
        sel_inner.pack(fill="x")

        # Project dropdown
        ttk.Label(sel_inner, text="Project:", style="Card.TLabel").pack(side="left", padx=(0, 8))
        self.act_project_combo = ttk.Combobox(sel_inner, width=35, state="readonly",
                                               font=("Segoe UI", 10))
        self.act_project_combo.pack(side="left", padx=(0, 10))
        self.act_project_combo.bind("<<ComboboxSelected>>", self._on_activity_project_change)

        ttk.Button(sel_inner, text="🔄 Refresh", command=self._refresh_activity_project_list).pack(side="left")

        # Week navigation
        week_frame = ttk.Frame(sel_inner)
        week_frame.pack(side="left", padx=(20, 0))

        self._act_current_week = datetime.date.today().isocalendar()[1]
        self._act_current_year = datetime.date.today().year

        prev_btn = tk.Button(week_frame, text="◀", font=("Segoe UI", 10, "bold"),
                              bg="#E0E0E0", relief="flat", cursor="hand2",
                              command=lambda: self._change_activity_week(-1), padx=10, pady=3)
        prev_btn.pack(side="left", padx=2)

        self.act_week_label = ttk.Label(week_frame,
                                         text=f"Week {self._act_current_week}, {self._act_current_year}",
                                         font=("Segoe UI", 11, "bold"))
        self.act_week_label.pack(side="left", padx=10)

        next_btn = tk.Button(week_frame, text="▶", font=("Segoe UI", 10, "bold"),
                              bg="#E0E0E0", relief="flat", cursor="hand2",
                              command=lambda: self._change_activity_week(1), padx=10, pady=3)
        next_btn.pack(side="left", padx=2)

        # Add / Edit / Delete buttons (only for editors)
        if self.can_edit:
            btn_frame = ttk.Frame(sel_inner)
            btn_frame.pack(side="right")
            ttk.Button(btn_frame, text="➕ Add Activity", style="Accent.TButton",
                       command=self._add_activity).pack(side="left", padx=3)
            ttk.Button(btn_frame, text="✏️ Edit", command=self._edit_activity).pack(side="left", padx=3)
            ttk.Button(btn_frame, text="🗑️ Delete", command=self._delete_activity).pack(side="left", padx=3)

        # ── Activities Table ─────────────────────────────────────────────
        table_frame = ttk.LabelFrame(tab, text="  Activities  ", padding=10)
        table_frame.pack(fill="both", expand=True)

        cols = ("id", "project", "activity", "start_date", "end_date", "time_taken",
                "members", "hard_points", "status", "attachment", "updated_by")
        self.act_tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=18)

        self.act_tree.heading("id", text="ID")
        self.act_tree.heading("project", text="Project")
        self.act_tree.heading("activity", text="Activity")
        self.act_tree.heading("start_date", text="Start Date")
        self.act_tree.heading("end_date", text="End Date")
        self.act_tree.heading("time_taken", text="Time Taken")
        self.act_tree.heading("members", text="Members Involved")
        self.act_tree.heading("hard_points", text="Hard Points")
        self.act_tree.heading("status", text="Status")
        self.act_tree.heading("attachment", text="Attachment")
        self.act_tree.heading("updated_by", text="Updated By")

        self.act_tree.column("id", width=40, anchor="center")
        self.act_tree.column("project", width=160)
        self.act_tree.column("activity", width=200)
        self.act_tree.column("start_date", width=90, anchor="center")
        self.act_tree.column("end_date", width=90, anchor="center")
        self.act_tree.column("time_taken", width=80, anchor="center")
        self.act_tree.column("members", width=180)
        self.act_tree.column("hard_points", width=160)
        self.act_tree.column("status", width=80, anchor="center")
        self.act_tree.column("attachment", width=120)
        self.act_tree.column("updated_by", width=90, anchor="center")

        scrollbar_y = ttk.Scrollbar(table_frame, orient="vertical", command=self.act_tree.yview)
        scrollbar_x = ttk.Scrollbar(table_frame, orient="horizontal", command=self.act_tree.xview)
        self.act_tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

        self.act_tree.pack(side="left", fill="both", expand=True)
        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(side="bottom", fill="x")

        # Status color tags
        self.act_tree.tag_configure("completed", background="#DCFCE7")  # light green
        self.act_tree.tag_configure("wip", background="#FEF9C3")  # light yellow

        # Cache for project list
        self._act_projects_cache = []

        self._refresh_activity_project_list()

    def _change_activity_week(self, delta: int):
        """Navigate weeks for the activities tab."""
        self._act_current_week += delta
        if self._act_current_week > 52:
            self._act_current_week = 1
            self._act_current_year += 1
        elif self._act_current_week < 1:
            self._act_current_week = 52
            self._act_current_year -= 1
        self.act_week_label.config(
            text=f"Week {self._act_current_week}, {self._act_current_year}")
        self._refresh_activities_table()

    def _refresh_activity_project_list(self):
        """Refresh the project dropdown in the Activities tab."""
        with db._connect(self.db_path) as conn:
            rows = conn.execute("SELECT id, name FROM projects ORDER BY start_date").fetchall()
        self._act_projects_cache = list(rows)
        # Add All Projects option
        values = ["All Projects"] + [f"{r['name']} ({r['id']})" for r in rows]

        self._act_projects_cache = rows  # keep original list

        self.act_project_combo["values"] = values
        self.act_project_combo.current(0)   # Default to All Projects
        self._on_activity_project_change()
        self._on_activity_project_change()

        def _get_act_selected_project_id(self):
            """Returns None when 'All Projects' is chosen."""
            idx = self.act_project_combo.current()
            if idx == 0:
                return None   # All projects mode
            return self._act_projects_cache[idx - 1]["id"]


    def _on_activity_project_change(self, event=None):
        """Reload activities when project selection changes."""
        self._refresh_activities_table()

    def _get_act_selected_project_id(self):
        """Get the selected project ID. Returns None when 'All Projects' is chosen."""
        idx = self.act_project_combo.current()
        if idx == 0:
            return None  # All Projects mode
        return self._act_projects_cache[idx - 1]["id"]

    def _refresh_activities_table(self):
        """Reload the activities treeview for the selected project + week."""
        self.act_tree.delete(*self.act_tree.get_children())

        project_id = self._get_act_selected_project_id()

        # If "All Projects" selected → fetch all
        if project_id is None:
            activities = []
            for proj in self._act_projects_cache:
                activities += db.get_activities(
                    proj["id"],
                    self._act_current_week,
                    self._act_current_year,
                    db_path=self.db_path
                )
        else:
            activities = db.get_activities(
                project_id,
                self._act_current_week,
                self._act_current_year,
                db_path=self.db_path
            )

        # Populate tree
        for act in activities:
            tag = "completed" if act["status"] == "Completed" else "wip"
            attachment_display = (
                pathlib.Path(act["attachment_path"]).name
                if act["attachment_path"] else ""
            )

            self.act_tree.insert(
                "",
                "end",
                values=(
                    act["id"],
                    act["project_name"],
                    act["activity_name"],
                    act["start_date"],
                    act["end_date"],
                    act["time_taken"],
                    act["members"],
                    act["hard_points"],
                    act["status"],
                    attachment_display,
                    act.get("updated_by", ""),
                ),
                tags=(tag,),
            )


    def _get_selected_activity_id(self) -> int | None:
        """Get the ID of the selected activity row."""
        sel = self.act_tree.selection()
        if not sel:
            return None
        return int(self.act_tree.item(sel[0], "values")[0])

    def _add_activity(self):
        """Open dialog to add a new activity."""
        project_id = self._get_act_selected_project_id()
        if project_id is None:
            from tkinter import messagebox
            messagebox.showwarning("No Project", "Please select a project first.")
            return

        dialog = _ActivityDialog(
            self.root, "Add Activity", self.db_path,
            week=self._act_current_week, year=self._act_current_year
        )
        self.root.wait_window(dialog.top)

        if dialog.result:
            db.add_activity(
                project_id=project_id,
                week_number=self._act_current_week,
                year=self._act_current_year,
                username=self.user["username"],
                db_path=self.db_path,
                **dialog.result
            )
            self._refresh_activities_table()
            self._set_status("Activity added successfully")

    def _edit_activity(self):
        """Open dialog to edit the selected activity."""
        activity_id = self._get_selected_activity_id()
        if activity_id is None:
            from tkinter import messagebox
            messagebox.showwarning("No Selection", "Please select an activity to edit.")
            return

        # Fetch current data
        with db._connect(self.db_path) as conn:
            row = conn.execute("SELECT * FROM activities WHERE id = ?", (activity_id,)).fetchone()
        if not row:
            return

        initial = {
            "activity_name": row["activity_name"],
            "start_date": row["start_date"],
            "end_date": row["end_date"],
            "time_taken": row["time_taken"],
            "members": row["members"],
            "hard_points": row["hard_points"],
            "status": row["status"],
            "attachment_path": row["attachment_path"],
        }

        dialog = _ActivityDialog(
            self.root, "Edit Activity", self.db_path,
            week=self._act_current_week, year=self._act_current_year,
            initial=initial
        )
        self.root.wait_window(dialog.top)

        if dialog.result:
            db.update_activity(
                activity_id=activity_id,
                username=self.user["username"],
                db_path=self.db_path,
                **dialog.result
            )
            self._refresh_activities_table()
            self._set_status("Activity updated successfully")

    def _delete_activity(self):
        """Delete the selected activity after confirmation."""
        activity_id = self._get_selected_activity_id()
        if activity_id is None:
            from tkinter import messagebox
            messagebox.showwarning("No Selection", "Please select an activity to delete.")
            return

        from tkinter import messagebox
        if messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this activity?"):
            db.delete_activity(activity_id, username=self.user["username"], db_path=self.db_path)
            self._refresh_activities_table()
            self._set_status("Activity deleted")


    # ═════════════════════════════════════════════════════════════════════
    # TAB 4 — PROJECT MANAGER
    # ═════════════════════════════════════════════════════════════════════

    def _build_projects_tab(self):
        tab = self.tab_projects

        header = ttk.Frame(tab)
        header.pack(fill="x", pady=(0, 10))
        ttk.Label(header, text="Project Manager", style="Header.TLabel").pack(side="left")

        if self.can_edit:
            btn_frame = ttk.Frame(header)
            btn_frame.pack(side="right")
            ttk.Button(btn_frame, text="➕ Add Project", style="Accent.TButton",
                       command=self._add_project).pack(side="left", padx=3)
            ttk.Button(btn_frame, text="✏️ Edit", command=self._edit_project).pack(side="left", padx=3)
            ttk.Button(btn_frame, text="🗑️ Delete", command=self._delete_project).pack(side="left", padx=3)
            ttk.Button(btn_frame, text="📥 Import JSON", command=self._import_json).pack(side="left", padx=3)

        ttk.Button(header, text="🔄 Refresh", command=self._refresh_projects_tab).pack(side="right", padx=3)

        table_frame = ttk.LabelFrame(tab, text="  All Projects  ", padding=10)
        table_frame.pack(fill="both", expand=True)

        cols = ("id", "name", "start_date", "end_date", "status", "dev_region", "sales_region", "color", "created_by", "updated_at")
        self.proj_tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=18)
        self.proj_tree.heading("id", text="ID")
        self.proj_tree.heading("name", text="Project Name")
        self.proj_tree.heading("start_date", text="Start Date")
        self.proj_tree.heading("end_date", text="End Date")
        self.proj_tree.heading("status", text="Status")
        self.proj_tree.heading("dev_region", text="Dev Region")
        self.proj_tree.heading("sales_region", text="Sales Region")
        self.proj_tree.heading("color", text="Color")
        self.proj_tree.heading("created_by", text="Created By")
        self.proj_tree.heading("updated_at", text="Last Updated")
        self.proj_tree.column("id", width=40)
        self.proj_tree.column("name", width=160)
        self.proj_tree.column("start_date", width=90)
        self.proj_tree.column("end_date", width=90)
        self.proj_tree.column("status", width=80)
        self.proj_tree.column("dev_region", width=90)
        self.proj_tree.column("sales_region", width=90)
        self.proj_tree.column("color", width=70)
        self.proj_tree.column("created_by", width=90)
        self.proj_tree.column("updated_at", width=130)

        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.proj_tree.yview)
        self.proj_tree.configure(yscrollcommand=scrollbar.set)
        self.proj_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        self._refresh_projects_tab()

    def _refresh_projects_tab(self):
        self.proj_tree.delete(*self.proj_tree.get_children())
        today = datetime.date.today()
        with db._connect(self.db_path) as conn:
            rows = conn.execute("SELECT * FROM projects ORDER BY start_date").fetchall()
            for r in rows:
                dev_reg = r["dev_region"] if "dev_region" in r.keys() else ""
                sales_reg = r["sales_region"] if "sales_region" in r.keys() else ""

                # Compute status from milestones for this project
                from timeline_tool.models import Milestone as MilestoneModel
                ms_rows = conn.execute(
                    "SELECT id, name, date FROM milestones WHERE project_id = ? ORDER BY date",
                    (r["id"],)
                ).fetchall()
                temp_milestones = []
                for ms_row in ms_rows:
                    task_statuses = {}
                    task_rows = conn.execute(
                        "SELECT status, COUNT(*) as cnt FROM milestone_tasks WHERE milestone_id = ? GROUP BY status",
                        (ms_row["id"],)
                    ).fetchall()
                    for tr in task_rows:
                        task_statuses[tr["status"]] = tr["cnt"]
                    temp_milestones.append(MilestoneModel(
                        name=ms_row["name"],
                        date=datetime.date.fromisoformat(ms_row["date"]),
                        milestone_id=ms_row["id"],
                        task_statuses=task_statuses,
                    ))
                # Build a temporary Project to compute status
                temp_proj = Project(
                    name=r["name"],
                    start_date=datetime.date.fromisoformat(r["start_date"]),
                    end_date=datetime.date.fromisoformat(r["end_date"]),
                    milestones=temp_milestones,
                )
                computed = temp_proj.computed_status(today).replace("-", " ").title()

                self.proj_tree.insert("", "end", values=(
                    r["id"], r["name"], r["start_date"], r["end_date"],
                    computed, dev_reg, sales_reg,
                    r["color"] or "auto", r["created_by"] or "", r["updated_at"],
                ))
        self._set_status(f"Projects tab refreshed — {len(rows)} project(s)")

    def _get_selected_proj_id(self) -> int | None:
        sel = self.proj_tree.selection()
        if not sel:
            messagebox.showwarning("Select", "Please select a project first.")
            return None
        return self.proj_tree.item(sel[0])["values"][0]

    def _add_project(self):
        dialog = _ProjectDialog(self.root, "Add Project")
        self.root.wait_window(dialog.top)
        if dialog.result:
            try:
                project_id = db.add_project(**dialog.result, username=self.user["username"], db_path=self.db_path)            
                self._refresh_projects_tab()
                self._refresh_dashboard()
                self._set_status(f"Project '{dialog.result['name']}' added")
            except Exception as e:
                messagebox.showerror("Error", str(e))

    def _edit_project(self):
        pid = self._get_selected_proj_id()
        if pid is None:
            return
        with db._connect(self.db_path) as conn:
            proj = conn.execute("SELECT * FROM projects WHERE id = ?", (pid,)).fetchone()
        initial = {
            "name": proj["name"], "start_date": proj["start_date"],
            "end_date": proj["end_date"], "color": proj["color"] or "",
            "status": proj["status"],
            "dev_region": proj["dev_region"] if "dev_region" in proj.keys() else "",
            "sales_region": proj["sales_region"] if "sales_region" in proj.keys() else "",
        }
        dialog = _ProjectDialog(self.root, "Edit Project", initial)
        self.root.wait_window(dialog.top)
        if dialog.result:
            try:
                db.update_project(pid, **dialog.result, username=self.user["username"], db_path=self.db_path)
                self._refresh_projects_tab()
                self._refresh_dashboard()
                self._set_status("Project updated")
            except Exception as e:
                messagebox.showerror("Error", str(e))

    def _delete_project(self):
        pid = self._get_selected_proj_id()
        if pid is None:
            return
        if messagebox.askyesno("Confirm", "Delete this project and all its milestones/phases?"):
            db.delete_project(pid, username=self.user["username"], db_path=self.db_path)
            self._refresh_projects_tab()
            self._refresh_dashboard()
            self._set_status("Project deleted")

    def _import_json(self):
        path = filedialog.askopenfilename(
            title="Select JSON file",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")])
        if path:
            try:
                db.import_from_json(path, username=self.user["username"], db_path=self.db_path)
                self._refresh_projects_tab()
                self._refresh_dashboard()
                self._set_status("JSON data imported successfully")
                messagebox.showinfo("Success", "Data imported successfully!")
            except Exception as e:
                messagebox.showerror("Error", str(e))

    # ═════════════════════════════════════════════════════════════════════
    # TAB 5 — MILESTONES & PHASES
    # ═════════════════════════════════════════════════════════════════════

    def _build_milestones_tab(self):
        tab = self.tab_milestones

        header = ttk.Frame(tab)
        header.pack(fill="x", pady=(0, 10))
        ttk.Label(header, text="Milestones & Phases", style="Header.TLabel").pack(side="left")

        selector_frame = ttk.LabelFrame(tab, text="  Select Project  ", padding=8)
        selector_frame.pack(fill="x", pady=(0, 10))

        sel_inner = ttk.Frame(selector_frame)
        sel_inner.pack(fill="x")

        ttk.Label(sel_inner, text="Project:", style="Card.TLabel").pack(side="left", padx=(0, 8))
        self.ms_project_combo = ttk.Combobox(sel_inner, width=40, state="readonly", font=("Segoe UI", 10))
        self.ms_project_combo.pack(side="left", padx=(0, 10))
        self.ms_project_combo.bind("<<ComboboxSelected>>", self._on_ms_project_change)
        ttk.Button(sel_inner, text="🔄 Refresh", command=self._refresh_ms_project_list).pack(side="left")

        paned = ttk.PanedWindow(tab, orient="horizontal")
        paned.pack(fill="both", expand=True)

        # Milestones panel
        ms_frame = ttk.LabelFrame(paned, text="  🎯 Milestones  ", padding=8)
        paned.add(ms_frame, weight=1)

        cols_ms = ("id", "name", "date", "updated_by")
        self.ms_tree = ttk.Treeview(ms_frame, columns=cols_ms, show="headings", height=12)
        self.ms_tree.heading("id", text="ID")
        self.ms_tree.heading("name", text="Milestone")
        self.ms_tree.heading("date", text="Date")
        self.ms_tree.heading("updated_by", text="Updated By")
        self.ms_tree.column("id", width=40)
        self.ms_tree.column("name", width=100)
        self.ms_tree.column("date", width=100)
        self.ms_tree.column("updated_by", width=100)
        self.ms_tree.pack(fill="both", expand=True)

        if self.can_edit:
            ms_btn = ttk.Frame(ms_frame)
            ms_btn.pack(fill="x", pady=(5, 0))
            ttk.Button(ms_btn, text="➕ Add", style="Accent.TButton",
                       command=self._add_milestone).pack(side="left", padx=2)
            ttk.Button(ms_btn, text="✏️ Edit", command=self._edit_milestone).pack(side="left", padx=2)
            ttk.Button(ms_btn, text="🗑️ Delete", command=self._delete_milestone).pack(side="left", padx=2)

        # Phases panel
        ph_frame = ttk.LabelFrame(paned, text="  📐 Phases  ", padding=8)
        paned.add(ph_frame, weight=1)

        cols_ph = ("id", "name", "start_date", "end_date")
        self.ph_tree = ttk.Treeview(ph_frame, columns=cols_ph, show="headings", height=12)
        self.ph_tree.heading("id", text="ID")
        self.ph_tree.heading("name", text="Phase")
        self.ph_tree.heading("start_date", text="Start")
        self.ph_tree.heading("end_date", text="End")
        self.ph_tree.column("id", width=40)
        self.ph_tree.column("name", width=120)
        self.ph_tree.column("start_date", width=100)
        self.ph_tree.column("end_date", width=100)
        self.ph_tree.pack(fill="both", expand=True)

        if self.can_edit:
            ph_btn = ttk.Frame(ph_frame)
            ph_btn.pack(fill="x", pady=(5, 0))
            ttk.Button(ph_btn, text="➕ Add", style="Accent.TButton",
                       command=self._add_phase).pack(side="left", padx=2)
            ttk.Button(ph_btn, text="✏️ Edit", command=self._edit_phase).pack(side="left", padx=2)
            ttk.Button(ph_btn, text="🗑️ Delete", command=self._delete_phase).pack(side="left", padx=2)

        self._refresh_ms_project_list()

    def _refresh_ms_project_list(self):
        with db._connect(self.db_path) as conn:
            rows = conn.execute("SELECT id, name FROM projects ORDER BY start_date").fetchall()
        self._ms_projects_cache = list(rows)
        self.ms_project_combo["values"] = [f"{r['id']} — {r['name']}" for r in rows]
        if rows:
            self.ms_project_combo.current(0)
            self._on_ms_project_change()

    def _get_ms_selected_project_id(self) -> int | None:
        idx = self.ms_project_combo.current()
        if idx < 0 or not self._ms_projects_cache:
            return None
        return self._ms_projects_cache[idx]["id"]

    def _on_ms_project_change(self, event=None):
        pid = self._get_ms_selected_project_id()
        if pid is None:
            return
        with db._connect(self.db_path) as conn:
            ms_rows = conn.execute(
                "SELECT id, name, date, created_by FROM milestones WHERE project_id = ? ORDER BY date",
                (pid,)).fetchall()
            ph_rows = conn.execute(
                "SELECT id, name, start_date, end_date FROM phases WHERE project_id = ? ORDER BY start_date",
                (pid,)).fetchall()

        self.ms_tree.delete(*self.ms_tree.get_children())
        for m in ms_rows:
            self.ms_tree.insert("", "end", values=(m["id"], m["name"], m["date"], m["created_by"] or ""))

        self.ph_tree.delete(*self.ph_tree.get_children())
        for p in ph_rows:
            self.ph_tree.insert("", "end", values=(p["id"], p["name"], p["start_date"], p["end_date"]))

    def _add_milestone(self):
        pid = self._get_ms_selected_project_id()
        if pid is None:
            messagebox.showwarning("Select", "Select a project first.")
            return
        dialog = _MilestoneDialog(self.root, "Add Milestone")
        self.root.wait_window(dialog.top)
        if dialog.result:
            try:
                db.add_milestone(pid, **dialog.result, username=self.user["username"], db_path=self.db_path)
                self._on_ms_project_change()
                self._set_status("Milestone added")
            except Exception as e:
                messagebox.showerror("Error", str(e))

    def _edit_milestone(self):
        sel = self.ms_tree.selection()
        if not sel:
            messagebox.showwarning("Select", "Select a milestone first.")
            return
        values = self.ms_tree.item(sel[0])["values"]
        ms_id = values[0]
        initial = {"name": str(values[1]), "date": str(values[2])}
        dialog = _MilestoneDialog(self.root, "Edit Milestone", initial)
        self.root.wait_window(dialog.top)
        if dialog.result:
            try:
                db.update_milestone(ms_id, **dialog.result, username=self.user["username"], db_path=self.db_path)
                self._on_ms_project_change()
                self._set_status("Milestone updated")
            except Exception as e:
                messagebox.showerror("Error", str(e))

    def _delete_milestone(self):
        sel = self.ms_tree.selection()
        if not sel:
            messagebox.showwarning("Select", "Select a milestone first.")
            return
        ms_id = self.ms_tree.item(sel[0])["values"][0]
        if messagebox.askyesno("Confirm", "Delete this milestone?"):
            db.delete_milestone(ms_id, username=self.user["username"], db_path=self.db_path)
            self._on_ms_project_change()
            self._set_status("Milestone deleted")

    def _add_phase(self):
        pid = self._get_ms_selected_project_id()
        if pid is None:
            messagebox.showwarning("Select", "Select a project first.")
            return
        dialog = _PhaseDialog(self.root, "Add Phase")
        self.root.wait_window(dialog.top)
        if dialog.result:
            try:
                db.add_phase(pid, **dialog.result, username=self.user["username"], db_path=self.db_path)
                self._on_ms_project_change()
                self._set_status("Phase added")
            except Exception as e:
                messagebox.showerror("Error", str(e))

    def _edit_phase(self):
        sel = self.ph_tree.selection()
        if not sel:
            messagebox.showwarning("Select", "Select a phase first.")
            return
        values = self.ph_tree.item(sel[0])["values"]
        ph_id = values[0]
        initial = {"name": str(values[1]), "start_date": str(values[2]), "end_date": str(values[3])}
        dialog = _PhaseDialog(self.root, "Edit Phase", initial)
        self.root.wait_window(dialog.top)
        if dialog.result:
            try:
                db.update_phase(ph_id, **dialog.result, username=self.user["username"], db_path=self.db_path)
                self._on_ms_project_change()
                self._set_status("Phase updated")
            except Exception as e:
                messagebox.showerror("Error", str(e))    
    
    def _delete_phase(self):
        sel = self.ph_tree.selection()
        if not sel:
            messagebox.showwarning("Select", "Select a phase first.")
            return
        ph_id = self.ph_tree.item(sel[0])["values"][0]
        if messagebox.askyesno("Confirm", "Delete this phase?"):
            db.delete_phase(ph_id, username=self.user["username"], db_path=self.db_path)
            self._on_ms_project_change()
            self._set_status("Phase deleted")

    # ═════════════════════════════════════════════════════════════════════
    # TAB 6 — RESOURCES
    # ═════════════════════════════════════════════════════════════════════

    def _build_resources_tab(self):
        """Build the Resources & Team Management tab."""
        tab = self.tab_resources
        
        header = ttk.Frame(tab)
        header.pack(fill="x", pady=(0, 10))
        ttk.Label(header, text="Resource & Team Management", style="Header.TLabel").pack(side="left")
        
        # Create sub-notebook for Resources
        res_nb = ttk.Notebook(tab)
        res_nb.pack(fill="both", expand=True)
        
        # Subtab 1: Team Members
        team_tab = ttk.Frame(res_nb, padding=10)
        res_nb.add(team_tab, text="  👤 Team Members  ")
        self._build_team_subtab(team_tab)
        
        # Subtab 2: Project Assignments
        assign_tab = ttk.Frame(res_nb, padding=10)
        res_nb.add(assign_tab, text="  📋 Assignments  ")
        self._build_assignments_subtab(assign_tab)
        
        # Subtab 3: Workload / Utilization
        workload_tab = ttk.Frame(res_nb, padding=10)
        res_nb.add(workload_tab, text="  📊 Workload  ")
        self._build_workload_subtab(workload_tab)

    def _build_team_subtab(self, tab):
        """Build the team members management subtab."""
        # Buttons
        btn_frame = ttk.Frame(tab)
        btn_frame.pack(fill="x", pady=(0, 10))
        
        if self.can_edit:
            ttk.Button(btn_frame, text="➕ Add Team Member", style="Accent.TButton",
                    command=self._add_team_member).pack(side="left", padx=2)
            ttk.Button(btn_frame, text="✏️ Edit", command=self._edit_team_member).pack(side="left", padx=2)
            ttk.Button(btn_frame, text="🗑️ Delete", command=self._delete_team_member).pack(side="left", padx=2)
        
        # Add Import Excel button for admins
        if self.can_manage:
            ttk.Separator(btn_frame, orient="vertical").pack(side="left", fill="y", padx=10)
            ttk.Button(btn_frame, text="📥 Import from Excel", 
                      command=self._import_team_from_excel).pack(side="left", padx=2)
        
        ttk.Button(btn_frame, text="🔄 Refresh", command=self._refresh_team_list).pack(side="right", padx=2)
        
        # Team member list
        cols = ("id", "name", "role", "department", "email", "allocation")
        self.team_tree = ttk.Treeview(tab, columns=cols, show="headings", height=15)
        self.team_tree.heading("id", text="ID")
        self.team_tree.heading("name", text="Name")
        self.team_tree.heading("role", text="Role")
        self.team_tree.heading("department", text="Department")
        self.team_tree.heading("email", text="Email")
        self.team_tree.heading("allocation", text="Capacity %")
        self.team_tree.column("id", width=40)
        self.team_tree.column("name", width=150)
        self.team_tree.column("role", width=120)
        self.team_tree.column("department", width=120)
        self.team_tree.column("email", width=180)
        self.team_tree.column("allocation", width=80)
        self.team_tree.pack(fill="both", expand=True)
        
        self._refresh_team_list()

    def _import_team_from_excel(self):
        """Import team members from the Excel file."""
        # Default path
        default_path = r"C:\Users\T0276HS\Stellantis\AI ML - VEHE STRUCTURE - Documents\Shared\project_timelines\User_Database.xlsx"
        
        # Ask user to select file (with default path if it exists)
        initial_dir = ""
        initial_file = ""
        if pathlib.Path(default_path).exists():
            initial_dir = str(pathlib.Path(default_path).parent)
            initial_file = pathlib.Path(default_path).name
        
        filepath = filedialog.askopenfilename(
            title="Select User Database Excel File",
            initialdir=initial_dir,
            initialfile=initial_file,
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if not filepath:
            return
        
        try:
            # Check if openpyxl is available
            try:
                import openpyxl
            except ImportError:
                messagebox.showerror(
                    "Missing Dependency",
                    "openpyxl is required for Excel import.\n\nInstall with:\npip install openpyxl"
                )
                return
            
            # Load the Excel file
            wb = openpyxl.load_workbook(filepath, read_only=True)
            ws = wb.active
            
            # Get header row to find column indices
            headers = {}
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            for idx, cell in enumerate(header_row):
                if cell:
                    headers[cell.strip().upper()] = idx
            
            # Map expected columns
            name_col = headers.get("NAME")
            designation_col = headers.get("DESIGNATION")
            dept_col = headers.get("META FUNCTION")
            user_id_col = headers.get("USER ID")
            role_col = headers.get("ROLE")
            email_col = headers.get("E-MAIL ID")
            
            if name_col is None:
                messagebox.showerror("Invalid Format", "Could not find 'NAME' column in the Excel file.")
                return
            
            # Parse email from "NAME <email>" format
            def parse_email(email_str):
                if not email_str:
                    return ""
                # Handle format: "NAME <email@domain.com>"
                if "<" in email_str and ">" in email_str:
                    start = email_str.find("<") + 1
                    end = email_str.find(">")
                    return email_str[start:end].strip()
                return email_str.strip()
            
            # Import data
            imported = 0
            skipped = 0
            errors = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Skip empty rows
                    if not row or not row[name_col]:
                        continue
                    
                    name = str(row[name_col]).strip() if row[name_col] else ""
                    if not name:
                        continue
                    
                    # Get other fields
                    designation = str(row[designation_col]).strip() if designation_col is not None and row[designation_col] else ""
                    department = str(row[dept_col]).strip() if dept_col is not None and row[dept_col] else ""
                    role = str(row[role_col]).strip() if role_col is not None and row[role_col] else ""
                    email_raw = str(row[email_col]).strip() if email_col is not None and row[email_col] else ""
                    email = parse_email(email_raw)
                    
                    # Map role from Excel to our role options
                    role_mapping = {
                        "MANAGER": "PM",
                        "LEAD": "Lead",
                        "ADMIN": "PM",
                        "DEVELOPER": "Developer",
                        "ENGINEER": "Developer",
                        "DESIGNER": "Designer",
                        "QA": "QA",
                        "ANALYST": "Analyst",
                    }
                    mapped_role = role_mapping.get(role.upper(), role if role else "Other")
                    
                    # Check if resource already exists (by name or email)
                    existing = self._check_resource_exists(name, email)
                    if existing:
                        skipped += 1
                        continue
                    
                    # Add resource
                    add_resource(
                        name=name,
                        role=mapped_role,
                        department=department,
                        email=email,
                        allocation_pct=100.0,
                        username=self.user["username"],
                        db_path=self.db_path
                    )
                    imported += 1
                    
                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)}")
            
            wb.close()
            
            # Refresh the list
            self._refresh_team_list()
            
            # Show result
            result_msg = f"Import completed!\n\n"
            result_msg += f"✅ Imported: {imported} team members\n"
            result_msg += f"⏭️ Skipped (already exist): {skipped}\n"
            
            if errors:
                result_msg += f"\n⚠️ Errors ({len(errors)}):\n"
                for err in errors[:5]:  # Show first 5 errors
                    result_msg += f"  • {err}\n"
                if len(errors) > 5:
                    result_msg += f"  ... and {len(errors) - 5} more"
            
            messagebox.showinfo("Import Complete", result_msg)
            self._set_status(f"Imported {imported} team members from Excel")
            
        except Exception as e:
            messagebox.showerror("Import Error", f"Failed to import from Excel:\n\n{str(e)}")

    def _check_resource_exists(self, name: str, email: str) -> bool:
        """Check if a resource with the given name or email already exists."""
        try:
            resources = get_all_resources(self.db_path)
            for res in resources:
                # Check by name (case-insensitive)
                if res.name.lower() == name.lower():
                    return True
                # Check by email (if provided)
                if email and res.email and res.email.lower() == email.lower():
                    return True
            return False
        except Exception:
            return False

    def _refresh_team_list(self):
        """Refresh the team members list."""
        self.team_tree.delete(*self.team_tree.get_children())
        
        try:
            resources = get_all_resources(self.db_path)
            for res in resources:
                self.team_tree.insert("", "end", values=(
                    res.id, res.name, res.role, res.department, res.email, f"{res.allocation_pct:.0f}%"
                ))
        except Exception as e:
            print(f"Error loading resources: {e}")

    def _add_team_member(self):
        """Add a new team member."""
        dialog = _ResourceDialog(self.root, "Add Team Member")
        self.root.wait_window(dialog.top)
        
        if dialog.result:
            try:
                add_resource(
                    **dialog.result,
                    username=self.user["username"],
                    db_path=self.db_path
                )
                self._refresh_team_list()
                self._set_status(f"Team member '{dialog.result['name']}' added")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to add team member:\n{e}")

    def _edit_team_member(self):
        """Edit selected team member."""
        sel = self.team_tree.selection()
        if not sel:
            messagebox.showwarning("Select", "Please select a team member to edit.")
            return
        
        values = self.team_tree.item(sel[0])["values"]
        resource_id = values[0]
        
        initial = {
            "name": values[1],
            "role": values[2],
            "department": values[3],
            "email": values[4],
            "allocation_pct": float(str(values[5]).replace("%", "")),
        }
        
        dialog = _ResourceDialog(self.root, "Edit Team Member", initial)
        self.root.wait_window(dialog.top)
        
        if dialog.result:
            try:
                update_resource(resource_id, **dialog.result, 
                            username=self.user["username"], db_path=self.db_path)
                self._refresh_team_list()
                self._set_status(f"Team member updated")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to update:\n{e}")

    def _delete_team_member(self):
        """Delete selected team member."""
        sel = self.team_tree.selection()
        if not sel:
            messagebox.showwarning("Select", "Please select a team member to delete.")
            return
        
        values = self.team_tree.item(sel[0])["values"]
        if messagebox.askyesno("Confirm", f"Delete team member '{values[1]}'?"):
            try:
                delete_resource(values[0], username=self.user["username"], db_path=self.db_path)
                self._refresh_team_list()
                self._set_status(f"Team member deleted")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete:\n{e}")

    def _build_assignments_subtab(self, tab):
        """Build the project assignments subtab."""
        # Project selector
        selector = ttk.Frame(tab)
        selector.pack(fill="x", pady=(0, 10))
        
        ttk.Label(selector, text="Project:").pack(side="left", padx=(0, 8))
        self.assign_project_combo = ttk.Combobox(selector, width=40, state="readonly")
        self.assign_project_combo.pack(side="left", padx=(0, 10))
        self.assign_project_combo.bind("<<ComboboxSelected>>", self._on_assign_project_change)
        
        if self.can_edit:
            ttk.Button(selector, text="➕ Assign Member", command=self._assign_member_to_project).pack(side="left", padx=5)
            ttk.Button(selector, text="🗑️ Remove", command=self._remove_member_from_project).pack(side="left", padx=5)
        ttk.Button(selector, text="🔄 Refresh", command=self._refresh_assignments).pack(side="right", padx=5)
        
        # Assignment list
        cols = ("resource_id", "name", "role", "project_role", "allocation", "notes")
        self.assign_tree = ttk.Treeview(tab, columns=cols, show="headings", height=12)
        self.assign_tree.heading("resource_id", text="ID")
        self.assign_tree.heading("name", text="Name")
        self.assign_tree.heading("role", text="Team Role")
        self.assign_tree.heading("project_role", text="Project Role")
        self.assign_tree.heading("allocation", text="Allocation %")
        self.assign_tree.heading("notes", text="Notes")
        self.assign_tree.column("resource_id", width=40)
        self.assign_tree.column("name", width=150)
        self.assign_tree.column("role", width=100)
        self.assign_tree.column("project_role", width=120)
        self.assign_tree.column("allocation", width=80)
        self.assign_tree.column("notes", width=200)
        self.assign_tree.pack(fill="both", expand=True)
        
        self._assign_projects_cache = []
        self._refresh_assignment_projects()

    def _refresh_assignment_projects(self):
        """Refresh project dropdown for assignments."""
        with db._connect(self.db_path) as conn:
            rows = conn.execute("SELECT id, name FROM projects ORDER BY name").fetchall()
        self._assign_projects_cache = list(rows)
        self.assign_project_combo["values"] = [f"{r['id']} — {r['name']}" for r in rows]
        if rows:
            self.assign_project_combo.current(0)
            self._on_assign_project_change()

    def _on_assign_project_change(self, event=None):
        """Load assignments for the selected project."""
        idx = self.assign_project_combo.current()
        if idx < 0 or not self._assign_projects_cache:
            return
        
        project_id = self._assign_projects_cache[idx]["id"]
        
        self.assign_tree.delete(*self.assign_tree.get_children())
        
        try:
            assignments = get_project_assignments(project_id, self.db_path)
            for a in assignments:
                self.assign_tree.insert("", "end", values=(
                    a["resource_id"], a["name"], a["role"], 
                    a["role_in_project"], f"{a['allocation_pct']:.0f}%", a["notes"]
                ))
        except Exception as e:
            print(f"Error loading assignments: {e}")

    def _refresh_assignments(self):
        """Refresh the assignments view."""
        self._refresh_assignment_projects()

    def _assign_member_to_project(self):
        """Assign a team member to the selected project."""
        idx = self.assign_project_combo.current()
        if idx < 0:
            messagebox.showwarning("Select", "Please select a project first.")
            return
        
        project_id = self._assign_projects_cache[idx]["id"]
        project_name = self._assign_projects_cache[idx]["name"]
        
        dialog = _AssignmentDialog(self.root, f"Assign to: {project_name}", self.db_path)
        self.root.wait_window(dialog.top)
        
        if dialog.result:
            try:
                assign_resource_to_project(
                    project_id=project_id,
                    resource_id=dialog.result["resource_id"],
                    role_in_project=dialog.result.get("role_in_project", ""),
                    allocation_pct=dialog.result.get("allocation_pct", 100),
                    notes=dialog.result.get("notes", ""),
                    username=self.user["username"],
                    db_path=self.db_path
                )
                self._on_assign_project_change()
                self._set_status("Team member assigned to project")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to assign:\n{e}")

    def _remove_member_from_project(self):
        """Remove selected assignment."""
        sel = self.assign_tree.selection()
        if not sel:
            messagebox.showwarning("Select", "Please select an assignment to remove.")
            return
        
        idx = self.assign_project_combo.current()
        project_id = self._assign_projects_cache[idx]["id"]
        resource_id = self.assign_tree.item(sel[0])["values"][0]
        
        if messagebox.askyesno("Confirm", "Remove this assignment?"):
            try:
                remove_assignment(project_id, resource_id, 
                                username=self.user["username"], db_path=self.db_path)
                self._on_assign_project_change()
                self._set_status("Assignment removed")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to remove:\n{e}")

    def _build_workload_subtab(self, tab):
        """Build the workload/utilization subtab."""
        btn_frame = ttk.Frame(tab)
        btn_frame.pack(fill="x", pady=(0, 10))
        ttk.Button(btn_frame, text="🔄 Refresh Workload", command=self._refresh_workload).pack(side="left", padx=5)
        
        # Utilization summary
        cols = ("name", "role", "capacity", "allocated", "available", "projects", "status")
        self.workload_tree = ttk.Treeview(tab, columns=cols, show="headings", height=15)
        self.workload_tree.heading("name", text="Name")
        self.workload_tree.heading("role", text="Role")
        self.workload_tree.heading("capacity", text="Capacity %")
        self.workload_tree.heading("allocated", text="Allocated %")
        self.workload_tree.heading("available", text="Available %")
        self.workload_tree.heading("projects", text="Projects")
        self.workload_tree.heading("status", text="Status")
        self.workload_tree.column("name", width=150)
        self.workload_tree.column("role", width=100)
        self.workload_tree.column("capacity", width=80)
        self.workload_tree.column("allocated", width=80)
        self.workload_tree.column("available", width=80)
        self.workload_tree.column("projects", width=70)
        self.workload_tree.column("status", width=100)
        self.workload_tree.pack(fill="both", expand=True)
        
        self._refresh_workload()

    def _refresh_workload(self):
        """Refresh the workload utilization view."""
        self.workload_tree.delete(*self.workload_tree.get_children())
        
        try:
            summary = get_team_utilization_summary(self.db_path)
            
            for util in summary:
                if "error" in util:
                    continue
                
                # Colored circle only: Yellow=Under, Green=Optimal, Red=Over
                status_display = {
                    "under": "🟡 Under",      # Yellow circle - underutilized
                    "optimal": "🟢 Optimal",  # Green circle - optimal
                    "over": "🔴 Over",        # Red circle - overutilized
                }.get(util["status"], util["status"])
                
                self.workload_tree.insert("", "end", values=(
                    util["resource_name"],
                    "",  # Role would need to be fetched
                    f"{util['max_capacity']:.0f}%",
                    f"{util['total_allocation']:.0f}%",
                    f"{util['available_capacity']:.0f}%",
                    util["project_count"],
                    status_display,
                ))
            
        except Exception as e:
            print(f"Error loading workload: {e}")
            
    # ═════════════════════════════════════════════════════════════════════
    # TAB 7 — ADMIN
    # ═════════════════════════════════════════════════════════════════════

    def _build_admin_tab(self):
        tab = self.tab_admin

        header = ttk.Frame(tab)
        header.pack(fill="x", pady=(0, 10))
        ttk.Label(header, text="Administration", style="Header.TLabel").pack(side="left")

        admin_nb = ttk.Notebook(tab)
        admin_nb.pack(fill="both", expand=True)

        users_tab = ttk.Frame(admin_nb, padding=10)
        admin_nb.add(users_tab, text="  👤 Users  ")
        self._build_users_subtab(users_tab)

        audit_tab = ttk.Frame(admin_nb, padding=10)
        admin_nb.add(audit_tab, text="  📋 Audit Log  ")
        self._build_audit_subtab(audit_tab)

        ref_tab = ttk.Frame(admin_nb, padding=10)
        admin_nb.add(ref_tab, text="  📏 Reference Lines  ")
        self._build_reflines_subtab(ref_tab)

        backup_tab = ttk.Frame(admin_nb, padding=10)
        admin_nb.add(backup_tab, text="  💾 Backup & Restore  ")
        self._build_backup_subtab(backup_tab)

    def _build_users_subtab(self, tab):
        btn_frame = ttk.Frame(tab)
        btn_frame.pack(fill="x", pady=(0, 8))
        ttk.Button(btn_frame, text="➕ Add User", style="Accent.TButton",
                   command=self._add_user).pack(side="left", padx=2)
        ttk.Button(btn_frame, text="🔄 Change Role", command=self._change_user_role).pack(side="left", padx=2)
        ttk.Button(btn_frame, text="🗑️ Delete User", command=self._delete_user).pack(side="left", padx=2)
        ttk.Button(btn_frame, text="🔄 Refresh", command=self._refresh_users).pack(side="right", padx=2)

        cols = ("username", "role", "full_name", "created_at")
        self.users_tree = ttk.Treeview(tab, columns=cols, show="headings", height=14)
        self.users_tree.heading("username", text="Username")
        self.users_tree.heading("role", text="Role")
        self.users_tree.heading("full_name", text="Full Name")
        self.users_tree.heading("created_at", text="Created")
        self.users_tree.column("username", width=150)
        self.users_tree.column("role", width=100)
        self.users_tree.column("full_name", width=200)
        self.users_tree.column("created_at", width=180)
        self.users_tree.pack(fill="both", expand=True)

        self._refresh_users()

    def _refresh_users(self):
        self.users_tree.delete(*self.users_tree.get_children())
        for u in auth.list_users(self.db_path):
            self.users_tree.insert("", "end", values=(
                u["username"], u["role"], u["full_name"] or "", u["created_at"]))

    def _add_user(self):
        dialog = _UserDialog(self.root, "Add User")
        self.root.wait_window(dialog.top)
        if dialog.result:
            try:
                auth.create_user(**dialog.result, db_path=self.db_path)
                self._refresh_users()
                self._set_status(f"User '{dialog.result['username']}' created")
            except Exception as e:
                messagebox.showerror("Error", str(e))

    def _change_user_role(self):
        sel = self.users_tree.selection()
        if not sel:
            messagebox.showwarning("Select", "Select a user first.")
            return
        username = self.users_tree.item(sel[0])["values"][0]
        new_role = simpledialog.askstring("Change Role",
                                          f"New role for '{username}':\n(admin / editor / viewer)")
        if new_role:
            try:
                auth.update_user_role(username, new_role.strip(), self.user["username"], self.db_path)
                self._refresh_users()
                self._set_status(f"Role for '{username}' changed to '{new_role.strip()}'")
            except Exception as e:
                messagebox.showerror("Error", str(e))

    def _delete_user(self):
        sel = self.users_tree.selection()
        if not sel:
            messagebox.showwarning("Select", "Select a user first.")
            return
        username = self.users_tree.item(sel[0])["values"][0]
        if username == self.user["username"]:
            messagebox.showwarning("Error", "Cannot delete yourself.")
            return
        if messagebox.askyesno("Confirm", f"Delete user '{username}'?"):
            auth.delete_user(username, self.user["username"], self.db_path)
            self._refresh_users()
            self._set_status(f"User '{username}' deleted")

    def _build_audit_subtab(self, tab):
        """Build the enhanced audit log subtab with filtering and search."""
        
        # ── Filter bar ──────────────────────────────────────────────────────
        filter_frame = ttk.LabelFrame(tab, text="  🔍 Filters  ", padding=8)
        filter_frame.pack(fill="x", pady=(0, 10))
        
        filter_row1 = ttk.Frame(filter_frame)
        filter_row1.pack(fill="x", pady=(0, 5))
        
        # Username filter
        ttk.Label(filter_row1, text="User:").pack(side="left", padx=(0, 5))
        self.audit_user_combo = ttk.Combobox(filter_row1, width=15, state="readonly")
        self.audit_user_combo.pack(side="left", padx=(0, 15))
        
        # Action filter
        ttk.Label(filter_row1, text="Action:").pack(side="left", padx=(0, 5))
        self.audit_action_combo = ttk.Combobox(filter_row1, width=20, state="readonly")
        self.audit_action_combo.pack(side="left", padx=(0, 15))
        
        # Date filters
        ttk.Label(filter_row1, text="From:").pack(side="left", padx=(0, 5))
        self.audit_from_entry = ttk.Entry(filter_row1, width=12)
        self.audit_from_entry.pack(side="left", padx=(0, 10))
        self.audit_from_entry.insert(0, (datetime.date.today() - datetime.timedelta(days=30)).strftime("%Y-%m-%d"))
        
        ttk.Label(filter_row1, text="To:").pack(side="left", padx=(0, 5))
        self.audit_to_entry = ttk.Entry(filter_row1, width=12)
        self.audit_to_entry.pack(side="left", padx=(0, 15))
        self.audit_to_entry.insert(0, datetime.date.today().strftime("%Y-%m-%d"))
        
        filter_row2 = ttk.Frame(filter_frame)
        filter_row2.pack(fill="x")
        
        # Search
        ttk.Label(filter_row2, text="Search:").pack(side="left", padx=(0, 5))
        self.audit_search_entry = ttk.Entry(filter_row2, width=30)
        self.audit_search_entry.pack(side="left", padx=(0, 15))
        self.audit_search_entry.bind("<Return>", lambda e: self._refresh_audit())
        
        # Buttons
        ttk.Button(filter_row2, text="🔍 Apply Filters", 
                   command=self._refresh_audit).pack(side="left", padx=5)
        ttk.Button(filter_row2, text="🔄 Clear Filters", 
                   command=self._clear_audit_filters).pack(side="left", padx=5)
        ttk.Button(filter_row2, text="📊 Summary", 
                   command=self._show_audit_summary).pack(side="left", padx=5)
        ttk.Button(filter_row2, text="📤 Export CSV", 
                   command=self._export_audit_csv).pack(side="left", padx=5)
        
        # ── Audit log table ─────────────────────────────────────────────────
        table_frame = ttk.Frame(tab)
        table_frame.pack(fill="both", expand=True)
        
        cols = ("icon", "timestamp", "username", "action", "detail")
        self.audit_tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=18)
        self.audit_tree.heading("icon", text="")
        self.audit_tree.heading("timestamp", text="Timestamp")
        self.audit_tree.heading("username", text="User")
        self.audit_tree.heading("action", text="Action")
        self.audit_tree.heading("detail", text="Details")
        self.audit_tree.column("icon", width=30, anchor="center")
        self.audit_tree.column("timestamp", width=150)
        self.audit_tree.column("username", width=100)
        self.audit_tree.column("action", width=150)
        self.audit_tree.column("detail", width=400)
        
        # Scrollbars
        scrollbar_y = ttk.Scrollbar(table_frame, orient="vertical", command=self.audit_tree.yview)
        scrollbar_x = ttk.Scrollbar(table_frame, orient="horizontal", command=self.audit_tree.xview)
        self.audit_tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        self.audit_tree.pack(side="left", fill="both", expand=True)
        scrollbar_y.pack(side="right", fill="y")
        
        # Double-click to view details
        self.audit_tree.bind("<Double-1>", self._show_audit_detail)
        
        # ── Status bar ──────────────────────────────────────────────────────
        self.audit_status = ttk.Label(tab, text="", font=("Segoe UI", 9), foreground="#666666")
        self.audit_status.pack(anchor="w", pady=(5, 0))
        
        # Initialize filter dropdowns
        self._init_audit_filters()
        
        # Load initial data
        self._refresh_audit()

    def _init_audit_filters(self):
        """Initialize the filter dropdowns with available values."""
        try:
            # Get unique users
            users = get_unique_users(self.db_path)
            self.audit_user_combo["values"] = ["All"] + users
            self.audit_user_combo.set("All")
            
            # Get unique actions
            actions = get_unique_actions(self.db_path)
            self.audit_action_combo["values"] = ["All"] + actions
            self.audit_action_combo.set("All")
        except Exception as e:
            print(f"Error initializing audit filters: {e}")

    def _refresh_audit(self):
        """Refresh the audit log with current filters."""
        self.audit_tree.delete(*self.audit_tree.get_children())
        
        try:
            # Get filter values
            username = self.audit_user_combo.get()
            if username == "All":
                username = None
            
            action = self.audit_action_combo.get()
            if action == "All":
                action = None
            
            # Parse dates
            start_date = None
            end_date = None
            try:
                from_str = self.audit_from_entry.get().strip()
                if from_str:
                    start_date = datetime.datetime.strptime(from_str, "%Y-%m-%d").date()
            except ValueError:
                pass
            
            try:
                to_str = self.audit_to_entry.get().strip()
                if to_str:
                    end_date = datetime.datetime.strptime(to_str, "%Y-%m-%d").date()
            except ValueError:
                pass
            
            search_term = self.audit_search_entry.get().strip() or None
            
            # Fetch filtered data
            entries = get_audit_log(
                db_path=self.db_path,
                limit=1000,
                username=username,
                action=action,
                start_date=start_date,
                end_date=end_date,
                search_term=search_term,
            )
            
            # Populate tree
            for entry in entries:
                icon = get_action_icon(entry["action"])
                self.audit_tree.insert("", "end", values=(
                    icon,
                    entry["timestamp"],
                    entry["username"],
                    entry["action"],
                    entry["detail"] or "",
                ), tags=(entry["action"],))
            
            # Update status
            self.audit_status.config(text=f"Showing {len(entries)} entries")
            
            # Color-code rows by action type
            self.audit_tree.tag_configure("DELETE_PROJECT", background="#FFEBEE")
            self.audit_tree.tag_configure("DELETE_MILESTONE", background="#FFEBEE")
            self.audit_tree.tag_configure("DELETE_USER", background="#FFEBEE")
            self.audit_tree.tag_configure("ADD_PROJECT", background="#E8F5E9")
            self.audit_tree.tag_configure("CREATE_USER", background="#E8F5E9")
            
        except Exception as e:
            self.audit_status.config(text=f"Error: {e}")

    def _clear_audit_filters(self):
        """Clear all audit filters and refresh."""
        self.audit_user_combo.set("All")
        self.audit_action_combo.set("All")
        self.audit_from_entry.delete(0, tk.END)
        self.audit_from_entry.insert(0, (datetime.date.today() - datetime.timedelta(days=30)).strftime("%Y-%m-%d"))
        self.audit_to_entry.delete(0, tk.END)
        self.audit_to_entry.insert(0, datetime.date.today().strftime("%Y-%m-%d"))
        self.audit_search_entry.delete(0, tk.END)
        self._refresh_audit()

    def _show_audit_detail(self, event):
        """Show detailed view of selected audit entry."""
        sel = self.audit_tree.selection()
        if not sel:
            return
        
        values = self.audit_tree.item(sel[0])["values"]
        
        # Create detail popup
        popup = tk.Toplevel(self.root)
        popup.title("Audit Entry Details")
        popup.geometry("500x300")
        popup.transient(self.root)
        
        frame = ttk.Frame(popup, padding=20)
        frame.pack(fill="both", expand=True)
        
        # Title
        ttk.Label(frame, text=f"{values[0]} Audit Entry", 
                  font=("Segoe UI", 14, "bold")).pack(anchor="w", pady=(0, 15))
        
        # Details
        details = [
            ("Timestamp:", values[1]),
            ("User:", values[2]),
            ("Action:", values[3]),
            ("Description:", get_action_description(values[3])),
        ]
        
        for label, value in details:
            row = ttk.Frame(frame)
            row.pack(fill="x", pady=3)
            ttk.Label(row, text=label, font=("Segoe UI", 10, "bold"), width=12).pack(side="left")
            ttk.Label(row, text=value, font=("Segoe UI", 10)).pack(side="left")
        
        # Detail text (can be long)
        ttk.Label(frame, text="Details:", font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(15, 5))
        
        detail_text = tk.Text(frame, height=6, wrap="word", font=("Segoe UI", 10))
        detail_text.insert("1.0", values[4] or "(No details)")
        detail_text.config(state="disabled")
        detail_text.pack(fill="both", expand=True)
        
        ttk.Button(frame, text="Close", command=popup.destroy).pack(pady=(15, 0))

    def _show_audit_summary(self):
        """Show audit activity summary popup."""
        try:
            summary = get_activity_summary(self.db_path, days=30)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to get summary: {e}")
            return
        
        popup = tk.Toplevel(self.root)
        popup.title("Activity Summary (Last 30 Days)")
        popup.geometry("600x500")
        popup.transient(self.root)
        
        frame = ttk.Frame(popup, padding=20)
        frame.pack(fill="both", expand=True)
        
        # Header
        ttk.Label(frame, text="📊 Activity Summary", 
                  font=("Segoe UI", 16, "bold")).pack(anchor="w", pady=(0, 5))
        ttk.Label(frame, text=f"Last {summary['period_days']} days | Total: {summary['total_actions']} actions",
                  font=("Segoe UI", 10), foreground="#666666").pack(anchor="w", pady=(0, 15))
        
        ttk.Separator(frame).pack(fill="x", pady=10)
        
        # Create notebook for tabs
        nb = ttk.Notebook(frame)
        nb.pack(fill="both", expand=True)
        
        # Tab 1: By Action
        action_tab = ttk.Frame(nb, padding=10)
        nb.add(action_tab, text="  By Action Type  ")
        
        action_tree = ttk.Treeview(action_tab, columns=("action", "count"), show="headings", height=10)
        action_tree.heading("action", text="Action")
        action_tree.heading("count", text="Count")
        action_tree.column("action", width=250)
        action_tree.column("count", width=80)
        action_tree.pack(fill="both", expand=True)
        
        for item in summary["by_action"][:20]:
            action_tree.insert("", "end", values=(item["action"], item["count"]))
        
        # Tab 2: By User
        user_tab = ttk.Frame(nb, padding=10)
        nb.add(user_tab, text="  By User  ")
        
        user_tree = ttk.Treeview(user_tab, columns=("user", "count"), show="headings", height=10)
        user_tree.heading("user", text="User")
        user_tree.heading("count", text="Actions")
        user_tree.column("user", width=200)
        user_tree.column("count", width=80)
        user_tree.pack(fill="both", expand=True)
        
        for item in summary["by_user"][:20]:
            user_tree.insert("", "end", values=(item["username"], item["count"]))
        
        # Tab 3: By Day
        day_tab = ttk.Frame(nb, padding=10)
        nb.add(day_tab, text="  By Day  ")
        
        day_tree = ttk.Treeview(day_tab, columns=("day", "count"), show="headings", height=10)
        day_tree.heading("day", text="Date")
        day_tree.heading("count", text="Actions")
        day_tree.column("day", width=150)
        day_tree.column("count", width=80)
        day_tree.pack(fill="both", expand=True)
        
        for item in summary["by_day"][:30]:
            day_tree.insert("", "end", values=(item["day"], item["count"]))
        
        ttk.Button(frame, text="Close", command=popup.destroy).pack(pady=(10, 0))

    def _export_audit_csv(self):
        """Export filtered audit log to CSV."""
        # Use initialfile, not initialname
        filepath = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
            title="Export Audit Log",
            initialfile=f"audit_log_{datetime.date.today().strftime('%Y%m%d')}.csv"
        )
        
        if not filepath:
            return
        
        try:
            # Get current filter values
            username = self.audit_user_combo.get()
            if username == "All":
                username = None
            
            action = self.audit_action_combo.get()
            if action == "All":
                action = None
            
            start_date = None
            end_date = None
            try:
                from_str = self.audit_from_entry.get().strip()
                if from_str:
                    start_date = datetime.datetime.strptime(from_str, "%Y-%m-%d").date()
            except ValueError:
                pass
            
            try:
                to_str = self.audit_to_entry.get().strip()
                if to_str:
                    end_date = datetime.datetime.strptime(to_str, "%Y-%m-%d").date()
            except ValueError:
                pass
            
            search_term = self.audit_search_entry.get().strip() or None
            
            count = export_audit_log_csv(
                output_path=pathlib.Path(filepath),
                db_path=self.db_path,
                username=username,
                action=action,
                start_date=start_date,
                end_date=end_date,
                search_term=search_term,
            )
            
            self._set_status(f"Exported {count} entries to CSV")
            messagebox.showinfo("Export Complete", f"Exported {count} audit entries to:\n\n{filepath}")
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export:\n{e}")

    def _build_reflines_subtab(self, tab):
        btn_frame = ttk.Frame(tab)
        btn_frame.pack(fill="x", pady=(0, 8))
        ttk.Button(btn_frame, text="➕ Add Reference Line", style="Accent.TButton",
                   command=self._add_ref_line).pack(side="left", padx=2)
        ttk.Button(btn_frame, text="🗑️ Delete", command=self._delete_ref_line).pack(side="left", padx=2)
        ttk.Button(btn_frame, text="🔄 Refresh", command=self._refresh_ref_lines).pack(side="right", padx=2)

        cols = ("id", "name", "date", "color", "style")
        self.ref_tree = ttk.Treeview(tab, columns=cols, show="headings", height=10)
        self.ref_tree.heading("id", text="ID")
        self.ref_tree.heading("name", text="Name")
        self.ref_tree.heading("date", text="Date")
        self.ref_tree.heading("color", text="Color")
        self.ref_tree.heading("style", text="Line Style")
        self.ref_tree.column("id", width=40)
        self.ref_tree.column("name", width=200)
        self.ref_tree.column("date", width=120)
        self.ref_tree.column("color", width=100)
        self.ref_tree.column("style", width=80)
        self.ref_tree.pack(fill="both", expand=True)

        self._refresh_ref_lines()

    def _refresh_ref_lines(self):
        self.ref_tree.delete(*self.ref_tree.get_children())
        with db._connect(self.db_path) as conn:
            rows = conn.execute("SELECT * FROM reference_lines ORDER BY date").fetchall()
        for r in rows:
            self.ref_tree.insert("", "end", values=(r["id"], r["name"], r["date"], r["color"], r["style"]))

    def _add_ref_line(self):
        dialog = _RefLineDialog(self.root, "Add Reference Line")
        self.root.wait_window(dialog.top)
        if dialog.result:
            try:
                db.add_reference_line(**dialog.result, username=self.user["username"], db_path=self.db_path)
                self._refresh_ref_lines()
                self._set_status("Reference line added")
            except Exception as e:
                messagebox.showerror("Error", str(e))

    def _delete_ref_line(self):
        sel = self.ref_tree.selection()
        if not sel:
            messagebox.showwarning("Select", "Select a reference line first.")
            return
        ref_id = self.ref_tree.item(sel[0])["values"][0]
        if messagebox.askyesno("Confirm", "Delete this reference line?"):
            db.delete_reference_line(ref_id, username=self.user["username"], db_path=self.db_path)
            self._refresh_ref_lines()

    def _build_backup_subtab(self, tab):
        """Build the backup & restore subtab."""
        
        # Action buttons
        btn_frame = ttk.Frame(tab)
        btn_frame.pack(fill="x", pady=(0, 15))
        
        ttk.Button(btn_frame, text="📦 Create Backup", command=self._create_manual_backup).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="📂 Restore Selected", command=self._restore_selected_backup).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="🗑️ Delete Selected", command=self._delete_selected_backup).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="🔄 Refresh", command=self._refresh_backups).pack(side="left", padx=5)
        
        ttk.Separator(btn_frame, orient="vertical").pack(side="left", fill="y", padx=15)
        
        ttk.Button(btn_frame, text="📤 Export to JSON", command=self._export_json).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="📥 Import from JSON", command=self._import_json_backup).pack(side="left", padx=5)
        
        # Backup list
        ttk.Label(tab, text="Available Backups:", font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(10, 5))
        
        cols = ("name", "created", "size")
        self.backup_tree = ttk.Treeview(tab, columns=cols, show="headings", height=12)
        self.backup_tree.heading("name", text="Backup Name")
        self.backup_tree.heading("created", text="Created")
        self.backup_tree.heading("size", text="Size (KB)")
        self.backup_tree.column("name", width=300)
        self.backup_tree.column("created", width=180)
        self.backup_tree.column("size", width=100)
        self.backup_tree.pack(fill="both", expand=True)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(tab, orient="vertical", command=self.backup_tree.yview)
        self.backup_tree.configure(yscrollcommand=scrollbar.set)
        
        self._refresh_backups()

    def _refresh_backups(self):
        """Refresh the backup list."""
        self.backup_tree.delete(*self.backup_tree.get_children())
        
        try:
            backups = list_backups(self.db_path or db.DEFAULT_DB_PATH)
            for backup in backups:
                self.backup_tree.insert("", "end", values=(
                    backup["name"],
                    backup["created"].strftime("%Y-%m-%d %H:%M:%S"),
                    f"{backup['size_kb']:.1f}",
                ), tags=(str(backup["path"]),))
        except Exception as e:
            print(f"Error loading backups: {e}")

    def _create_manual_backup(self):
        """Create a manual backup."""
        try:
            backup_path = create_backup(self.db_path or db.DEFAULT_DB_PATH, backup_name="manual")
            self._set_status(f"Backup created: {backup_path.name}")
            messagebox.showinfo("Backup Created", f"Backup saved successfully!\n\n{backup_path}")
            self._refresh_backups()
        except Exception as e:
            messagebox.showerror("Backup Error", f"Failed to create backup:\n\n{e}")

    def _restore_selected_backup(self):
        """Restore from the selected backup."""
        selected = self.backup_tree.selection()
        if not selected:
            messagebox.showwarning("No Selection", "Please select a backup to restore.")
            return
        
        item = self.backup_tree.item(selected[0])
        backup_name = item["values"][0]
        
        if not messagebox.askyesno(
            "Confirm Restore",
            f"Are you sure you want to restore from:\n\n{backup_name}\n\n"
            "This will replace the current database. A pre-restore backup will be created."
        ):
            return
        
        try:
            backup_dir = get_backup_dir(self.db_path or db.DEFAULT_DB_PATH)
            backup_path = backup_dir / f"{backup_name}.db"
            
            restore_backup(backup_path, self.db_path or db.DEFAULT_DB_PATH)
            
            self._set_status("Database restored successfully")
            messagebox.showinfo("Restore Complete", "Database restored successfully!\n\nPlease restart the application.")
            self._refresh_backups()
            
        except Exception as e:
            messagebox.showerror("Restore Error", f"Failed to restore backup:\n\n{e}")

    def _delete_selected_backup(self):
        """Delete the selected backup."""
        selected = self.backup_tree.selection()
        if not selected:
            messagebox.showwarning("No Selection", "Please select a backup to delete.")
            return
        
        item = self.backup_tree.item(selected[0])
        backup_name = item["values"][0]
        
        if not messagebox.askyesno("Confirm Delete", f"Delete backup:\n\n{backup_name}"):
            return
        
        try:
            backup_dir = get_backup_dir(self.db_path or db.DEFAULT_DB_PATH)
            backup_path = backup_dir / f"{backup_name}.db"
            
            delete_backup(backup_path)
            self._set_status(f"Backup deleted: {backup_name}")
            self._refresh_backups()
            
        except Exception as e:
            messagebox.showerror("Delete Error", f"Failed to delete backup:\n\n{e}")

    def _export_json(self):
        """Export database to JSON file."""
        # Use initialfile, not initialname
        filepath = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json")],
            title="Export to JSON",
            initialfile=f"project_data_{datetime.date.today().strftime('%Y%m%d')}.json"  # FIXED
        )
        
        if not filepath:
            return
        
        try:
            export_to_json(
                self.db_path or db.DEFAULT_DB_PATH,
                pathlib.Path(filepath)
            )
            self._set_status(f"Exported to {filepath}")
            messagebox.showinfo("Export Complete", f"Data exported successfully!\n\n{filepath}")
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export:\n\n{e}")

    def _import_json_backup(self):
        """Import data from JSON file (from backup tab)."""
        filepath = filedialog.askopenfilename(
            filetypes=[("JSON files", "*.json")],
            title="Import from JSON"
        )
        
        if not filepath:
            return
        
        merge = messagebox.askyesno(
            "Import Mode",
            "How do you want to import?\n\n"
            "YES = Merge with existing data\n"
            "NO = Replace all data (backup will be created)"
        )
        
        try:
            # Create backup before import
            if not merge:
                create_backup(self.db_path or db.DEFAULT_DB_PATH, backup_name="pre_import")
            
            stats = import_from_json(
                pathlib.Path(filepath),
                self.db_path or db.DEFAULT_DB_PATH,
                merge=merge
            )
            
            self._set_status("Import complete")
            messagebox.showinfo(
                "Import Complete",
                f"Data imported successfully!\n\n"
                f"Projects: {stats['projects']}\n"
                f"Milestones: {stats['milestones']}\n"
                f"Phases: {stats['phases']}\n"
                f"Tasks: {stats['tasks']}"
            )
            
            # Refresh views
            self._refresh_backups()
            self._refresh_dashboard()
            
        except Exception as e:
            messagebox.showerror("Import Error", f"Failed to import:\n\n{e}")

    def _show_summary_dashboard(self):
        """Show a comprehensive summary dashboard popup with charts."""
        projects, ref_lines = db.load_all(self.db_path)
        today = datetime.date.today()
        
        if not projects:
            messagebox.showinfo("No Data", "No projects to summarize.")
            return
        
        # Create popup window
        popup = tk.Toplevel(self.root)
        popup.title("📊 Portfolio Summary Dashboard")
        popup.geometry("1000x700")
        popup.transient(self.root)
        
        # Main frame with scrolling
        main_frame = ttk.Frame(popup, padding=20)
        main_frame.pack(fill="both", expand=True)
        
        # ══════════════════════════════════════════════════════════════════
        # HEADER
        # ══════════════════════════════════════════════════════════════════
        header = ttk.Frame(main_frame)
        header.pack(fill="x", pady=(0, 15))
        
        ttk.Label(header, text="📊 Portfolio Summary Dashboard", 
                  font=("Segoe UI", 18, "bold")).pack(side="left")
        ttk.Label(header, text=f"Generated: {today.strftime('%B %d, %Y')}",
                  font=("Segoe UI", 10), foreground="#666666").pack(side="right")
        
        ttk.Separator(main_frame).pack(fill="x", pady=10)
        
        # ══════════════════════════════════════════════════════════════════
        # TOP ROW: Summary Stats Cards
        # ══════════════════════════════════════════════════════════════════
        stats_frame = ttk.Frame(main_frame)
        stats_frame.pack(fill="x", pady=(0, 15))
        
        # Calculate statistics
        total = len(projects)
        on_track = sum(1 for p in projects if p.computed_status(today) == "on-track")
        at_risk = sum(1 for p in projects if p.computed_status(today) == "at-risk")
        overdue = sum(1 for p in projects if p.computed_status(today) == "overdue")
        
        # Active projects (end date >= today)
        active = sum(1 for p in projects if p.end_date >= today)
        completed = sum(1 for p in projects if p.end_date < today)
        
        # Calculate health score (0-100)
        if total > 0:
            health_score = int((on_track / total) * 100)
        else:
            health_score = 0
        
        # Health score color
        if health_score >= 80:
            health_color = "#16A34A"  # Green
            health_label = "Excellent"
        elif health_score >= 60:
            health_color = "#EAB308"  # Yellow
            health_label = "Good"
        elif health_score >= 40:
            health_color = "#F97316"  # Orange
            health_label = "Fair"
        else:
            health_color = "#DC2626"  # Red
            health_label = "Needs Attention"
        
        # Create stat cards
        stats_data = [
            ("📁", "Total Projects", str(total), "#0067C0"),
            ("✅", "Active", str(active), "#16A34A"),
            ("🏁", "Completed", str(completed), "#6B7280"),
            ("🟢", "On Track", str(on_track), "#16A34A"),
            ("🟡", "At Risk", str(at_risk), "#EAB308"),
            ("🔴", "Overdue", str(overdue), "#DC2626"),
            ("💚", f"Health: {health_label}", f"{health_score}%", health_color),
        ]
        
        for icon, label, value, color in stats_data:
            card = tk.Frame(stats_frame, bg="white", relief="flat",
                            bd=1, highlightbackground="#E0E0E0", highlightthickness=1)
            card.pack(side="left", fill="x", expand=True, padx=3, ipady=8, ipadx=12)
            
            tk.Label(card, text=icon, font=("Segoe UI", 18), bg="white").pack(side="left", padx=(8, 5))
            text_frame = tk.Frame(card, bg="white")
            text_frame.pack(side="left")
            tk.Label(text_frame, text=value, font=("Segoe UI", 20, "bold"),
                     bg="white", fg=color).pack(anchor="w")
            tk.Label(text_frame, text=label, font=("Segoe UI", 9),
                     bg="white", fg="#666666").pack(anchor="w")
        
        # ══════════════════════════════════════════════════════════════════
        # MIDDLE ROW: Charts
        # ══════════════════════════════════════════════════════════════════
        charts_frame = ttk.Frame(main_frame)
        charts_frame.pack(fill="both", expand=True, pady=(0, 15))
        
        # Left: Status Pie Chart
        pie_frame = ttk.LabelFrame(charts_frame, text="  📊 Project Status Distribution  ", padding=10)
        pie_frame.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        fig_pie, ax_pie = plt.subplots(figsize=(4, 3))
        fig_pie.patch.set_facecolor('#FFFFFF')
        
        status_counts = [on_track, at_risk, overdue]
        status_labels = ['On Track', 'At Risk', 'Overdue']
        status_colors_pie = ['#16A34A', '#EAB308', '#DC2626']
        
        # Only show non-zero values
        non_zero = [(c, l, col) for c, l, col in zip(status_counts, status_labels, status_colors_pie) if c > 0]
        if non_zero:
            counts, labels, colors = zip(*non_zero)
            wedges, texts, autotexts = ax_pie.pie(counts, labels=labels, colors=colors,
                                                   autopct='%1.0f%%', startangle=90,
                                                   textprops={'fontsize': 9})
            for autotext in autotexts:
                autotext.set_color('white')
                autotext.set_fontweight('bold')
        else:
            ax_pie.text(0.5, 0.5, "No data", ha='center', va='center', fontsize=12)
        
        ax_pie.set_title("Status Breakdown", fontsize=11, fontweight='bold')
        
        canvas_pie = FigureCanvasTkAgg(fig_pie, master=pie_frame)
        canvas_pie.draw()
        canvas_pie.get_tk_widget().pack(fill="both", expand=True)
        
        # Right: Upcoming Milestones
        ms_frame = ttk.LabelFrame(charts_frame, text="  📅 Upcoming Milestones (Next 30 Days)  ", padding=10)
        ms_frame.pack(side="right", fill="both", expand=True, padx=(10, 0))
        
        # Collect upcoming milestones
        upcoming_milestones = []
        cutoff = today + datetime.timedelta(days=30)
        for p in projects:
            for ms in p.milestones:
                if today <= ms.date <= cutoff:
                    days_until = (ms.date - today).days
                    upcoming_milestones.append({
                        "project": p.name,
                        "milestone": ms.name,
                        "date": ms.date,
                        "days_until": days_until,
                        "is_complete": ms.is_complete(),
                    })
        
        # Sort by date
        upcoming_milestones.sort(key=lambda x: x["date"])
        
        if upcoming_milestones:
            # Create treeview
            cols = ("project", "milestone", "date", "days", "status")
            ms_tree = ttk.Treeview(ms_frame, columns=cols, show="headings", height=8)
            ms_tree.heading("project", text="Project")
            ms_tree.heading("milestone", text="Milestone")
            ms_tree.heading("date", text="Date")
            ms_tree.heading("days", text="Days")
            ms_tree.heading("status", text="Status")
            ms_tree.column("project", width=120)
            ms_tree.column("milestone", width=80)
            ms_tree.column("date", width=90)
            ms_tree.column("days", width=50)
            ms_tree.column("status", width=80)
            
            for ms in upcoming_milestones[:15]:  # Show top 15
                status = "✅ Complete" if ms["is_complete"] else "⏳ Pending"
                days_text = "Today" if ms["days_until"] == 0 else f"{ms['days_until']}d"
                
                tags = ()
                if ms["days_until"] <= 3 and not ms["is_complete"]:
                    tags = ("urgent",)
                elif ms["is_complete"]:
                    tags = ("complete",)
                
                ms_tree.insert("", "end", values=(
                    ms["project"],
                    ms["milestone"],
                    ms["date"].strftime("%Y-%m-%d"),
                    days_text,
                    status,
                ), tags=tags)
            
            ms_tree.tag_configure("urgent", background="#FFEBEE")
            ms_tree.tag_configure("complete", background="#E8F5E9")
            
            ms_tree.pack(fill="both", expand=True)
        else:
            ttk.Label(ms_frame, text="No milestones in the next 30 days",
                     font=("Segoe UI", 11), foreground="#999999").pack(expand=True)
        
        # ══════════════════════════════════════════════════════════════════
        # BOTTOM: Project Timeline Overview
        # ══════════════════════════════════════════════════════════════════
        timeline_frame = ttk.LabelFrame(main_frame, text="  📈 Project Timeline Overview  ", padding=10)
        timeline_frame.pack(fill="x", pady=(0, 10))
        
        # Mini timeline chart
        fig_timeline, ax_timeline = plt.subplots(figsize=(10, 2))
        fig_timeline.patch.set_facecolor('#FFFFFF')
        
        # Sort projects by start date
        sorted_projects = sorted(projects, key=lambda p: p.start_date)[:10]  # Top 10
        
        for idx, p in enumerate(sorted_projects):
            color = cfg.STATUS_COLORS.get(p.computed_status(today), "#0067C0")
            ax_timeline.barh(idx, (p.end_date - p.start_date).days,
                            left=p.start_date, height=0.6, color=color, alpha=0.7)
        
        ax_timeline.axvline(today, color='red', linewidth=1.5, linestyle='--', label='Today')
        ax_timeline.set_yticks(range(len(sorted_projects)))
        ax_timeline.set_yticklabels([p.name[:20] for p in sorted_projects], fontsize=8)
        ax_timeline.xaxis.set_major_formatter(mdates.DateFormatter('%b %Y'))
        ax_timeline.set_title("Project Timelines (Top 10 by Start Date)", fontsize=10, fontweight='bold')
        fig_timeline.tight_layout()
        
        canvas_timeline = FigureCanvasTkAgg(fig_timeline, master=timeline_frame)
        canvas_timeline.draw()
        canvas_timeline.get_tk_widget().pack(fill="x")
        
        # ══════════════════════════════════════════════════════════════════
        # CLOSE BUTTON
        # ══════════════════════════════════════════════════════════════════
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill="x")
        
        ttk.Button(btn_frame, text="Close", command=popup.destroy).pack(side="right")
        ttk.Button(btn_frame, text="📄 Export Report", 
                   command=lambda: self._export_pdf_report()).pack(side="right", padx=5)
        
        # Cleanup matplotlib figures when popup closes
        def on_close():
            plt.close(fig_pie)
            plt.close(fig_timeline)
            popup.destroy()
        
        popup.protocol("WM_DELETE_WINDOW", on_close)

# ─────────────────────────────────────────────────────────────────────────
# Input Dialogs
# ─────────────────────────────────────────────────────────────────────────

class _ProjectDialog:
    def __init__(self, parent, title, initial=None):
        self.result = None
        self.top = tk.Toplevel(parent)
        self.top.title(title)
        self.top.geometry("380x500")
        self.top.grab_set()

        f = ttk.Frame(self.top, padding=20)
        f.pack(fill="both", expand=True)

        fields = [
            ("Name:", "name"), ("Start Date (YYYY-MM-DD):", "start_date"),
            ("End Date (YYYY-MM-DD):", "end_date"),
        ]
        self.entries = {}
        for label, key in fields:
            ttk.Label(f, text=label).pack(anchor="w")
            entry = ttk.Entry(f, width=42, font=("Segoe UI", 10))
            entry.pack(pady=(0, 6), ipady=2)
            self.entries[key] = entry

        # Status is now auto-calculated — show info label instead
        ttk.Label(f, text="Status:").pack(anchor="w")
        ttk.Label(f, text="⚡ Auto-calculated from milestone task completion",
                  font=("Segoe UI", 9, "italic"), foreground="#888888").pack(anchor="w", pady=(0, 8))

        ttk.Label(f, text="Development Region:").pack(anchor="w")
        self.dev_region_combo = ttk.Combobox(f, values=REGION_OPTIONS,
                                              width=39, font=("Segoe UI", 10),
                                              state="readonly")
        self.dev_region_combo.set("")
        self.dev_region_combo.pack(pady=(0, 8))

        ttk.Label(f, text="Sales Region:").pack(anchor="w")
        self.sales_region_combo = ttk.Combobox(f, values=REGION_OPTIONS,
                                                width=39, font=("Segoe UI", 10),
                                                state="readonly")
        self.sales_region_combo.set("")
        self.sales_region_combo.pack(pady=(0, 12))

        ttk.Button(f, text="💾 Save", style="Accent.TButton", command=self._save).pack(fill="x", ipady=2)

        if initial:
            for key, entry in self.entries.items():
                entry.insert(0, initial.get(key, ""))
            self.dev_region_combo.set(initial.get("dev_region", ""))
            self.sales_region_combo.set(initial.get("sales_region", ""))

    def _save(self):
        name = self.entries["name"].get().strip()
        start = self.entries["start_date"].get().strip()
        end = self.entries["end_date"].get().strip()
        if not name or not start or not end:
            messagebox.showwarning("Required", "Name, Start Date, and End Date are required.")
            return
        self.result = {
            "name": name, "start_date": start, "end_date": end,
            "color": cfg.DEFAULT_PROJECT_COLOR,
            "status": "on-track",  # Default; actual status is auto-calculated at runtime
            "dev_region": self.dev_region_combo.get(),
            "sales_region": self.sales_region_combo.get(),
        }
        self.top.destroy()


class _MilestoneDialog:
    MILESTONE_NAMES = [
        "IM", "CM", "PM", "SFM", "SHRM", "Post SHRM",
        "Mule Build", "X0", "MPRM", "X1", "X2", "LRM", "X3", "SOP",
        "Other",
    ]

    def __init__(self, parent, title, initial=None):
        self.result = None
        self.top = tk.Toplevel(parent)
        self.top.title(title)
        self.top.resizable(False, False)
        self.top.grab_set()
        self.top.update_idletasks()
        self.top.geometry(f"350x{self.top.winfo_reqheight()}")

        f = ttk.Frame(self.top, padding=20)
        f.pack(fill="both", expand=True)

        ttk.Label(f, text="Milestone Name:").pack(anchor="w")
        self.name_combo = ttk.Combobox(f, values=self.MILESTONE_NAMES,
                                        width=36, font=("Segoe UI", 10),
                                        state="readonly")
        self.name_combo.pack(pady=(0, 8), ipady=2)
        self.name_combo.bind("<<ComboboxSelected>>", self._on_name_change)

        self.other_label = ttk.Label(f, text="Custom Milestone Name:")
        self.other_entry = ttk.Entry(f, width=38, font=("Segoe UI", 10))
        # Widgets are created but not packed yet; shown on demand

        self._date_label = ttk.Label(f, text="Date (YYYY-MM-DD):")
        self._date_label.pack(anchor="w")
        self.date_entry = ttk.Entry(f, width=38, font=("Segoe UI", 10))
        self.date_entry.pack(pady=(0, 15), ipady=2)

        ttk.Button(f, text="💾 Save", style="Accent.TButton", command=self._save).pack(fill="x", ipady=2)

        if initial:
            name = initial.get("name", "")
            if name in self.MILESTONE_NAMES:
                self.name_combo.set(name)
                if name == "Other":
                    self._show_other_entry()
            else:
                # Custom name not in predefined list — select "Other" and pre-fill
                self.name_combo.set("Other")
                self._show_other_entry()
                self.other_entry.insert(0, name)
            self.date_entry.insert(0, initial.get("date", ""))

    def _on_name_change(self, event=None):
        if self.name_combo.get() == "Other":
            self._show_other_entry()
        else:
            self._hide_other_entry()

    def _show_other_entry(self):
        self.other_label.pack(anchor="w", before=self._date_label)
        self.other_entry.pack(pady=(0, 8), ipady=2, before=self._date_label)
        self.top.update_idletasks()
        self.top.geometry(f"350x{self.top.winfo_reqheight()}")

    def _hide_other_entry(self):
        self.other_entry.pack_forget()
        self.other_label.pack_forget()
        self.other_entry.delete(0, tk.END)
        self.top.update_idletasks()
        self.top.geometry(f"350x{self.top.winfo_reqheight()}")

    def _save(self):
        combo_val = self.name_combo.get().strip()
        if combo_val == "Other":
            name = self.other_entry.get().strip()
        else:
            name = combo_val
        date = self.date_entry.get().strip()
        if not name or not date:
            messagebox.showwarning("Required", "Both fields are required.")
            return
        self.result = {"name": name, "date": date}
        self.top.destroy()


class _PhaseDialog:
    def __init__(self, parent, title, initial=None):
        self.result = None
        self.top = tk.Toplevel(parent)
        self.top.title(title)
        self.top.geometry("350x250")
        self.top.grab_set()

        f = ttk.Frame(self.top, padding=20)
        f.pack(fill="both", expand=True)

        for label, attr in [("Phase Name:", "name_entry"),
                             ("Start Date (YYYY-MM-DD):", "start_entry"),
                             ("End Date (YYYY-MM-DD):", "end_entry")]:
            ttk.Label(f, text=label).pack(anchor="w")
            entry = ttk.Entry(f, width=38, font=("Segoe UI", 10))
            entry.pack(pady=(0, 8), ipady=2)
            setattr(self, attr, entry)

        ttk.Button(f, text="💾 Save", style="Accent.TButton", command=self._save).pack(fill="x", ipady=2)

        if initial:
            self.name_entry.insert(0, initial.get("name", ""))
            self.start_entry.insert(0, initial.get("start_date", ""))
            self.end_entry.insert(0, initial.get("end_date", ""))

    def _save(self):
        name = self.name_entry.get().strip()
        start = self.start_entry.get().strip()
        end = self.end_entry.get().strip()
        if not name or not start or not end:
            messagebox.showwarning("Required", "All fields are required.")
            return
        self.result = {"name": name, "start_date": start, "end_date": end}
        self.top.destroy()


class _UserDialog:
    def __init__(self, parent, title):
        self.result = None
        self.top = tk.Toplevel(parent)
        self.top.title(title)
        self.top.geometry("350x300")
        self.top.grab_set()

        f = ttk.Frame(self.top, padding=20)
        f.pack(fill="both", expand=True)

        for label, attr, show in [("Username:", "user_entry", None),
                                    ("Full Name:", "name_entry", None),
                                    ("Password:", "pass_entry", "•")]:
            ttk.Label(f, text=label).pack(anchor="w")
            entry = ttk.Entry(f, width=38, font=("Segoe UI", 10), show=show or "")
            entry.pack(pady=(0, 8), ipady=2)
            setattr(self, attr, entry)

        ttk.Label(f, text="Role:").pack(anchor="w")
        self.role_combo = ttk.Combobox(f, values=["admin", "editor", "viewer"],
                                        width=35, font=("Segoe UI", 10))
        self.role_combo.set("viewer")
        self.role_combo.pack(pady=(0, 12))

        ttk.Button(f, text="👤 Create User", style="Accent.TButton", command=self._save).pack(fill="x", ipady=2)

    def _save(self):
        username = self.user_entry.get().strip()
        password = self.pass_entry.get().strip()
        if not username or not password:
            messagebox.showwarning("Required", "Username and password are required.")
            return
        self.result = {
            "username": username, "password": password,
            "role": self.role_combo.get(),
            "full_name": self.name_entry.get().strip(),
        }
        self.top.destroy()


class _RefLineDialog:
    def __init__(self, parent, title):
        self.result = None
        self.top = tk.Toplevel(parent)
        self.top.title(title)
        self.top.geometry("350x280")
        self.top.grab_set()

        f = ttk.Frame(self.top, padding=20)
        f.pack(fill="both", expand=True)

        for label, attr in [("Name:", "name_entry"),
                             ("Date (YYYY-MM-DD):", "date_entry"),
                             ("Color (#RRGGBB):", "color_entry")]:
            ttk.Label(f, text=label).pack(anchor="w")
            entry = ttk.Entry(f, width=38, font=("Segoe UI", 10))
            entry.pack(pady=(0, 8), ipady=2)
            setattr(self, attr, entry)

        self.color_entry.insert(0, "#2196F3")

        ttk.Label(f, text="Line Style:").pack(anchor="w")
        self.style_combo = ttk.Combobox(f, values=["--", "-.", ":", "-"],
                                         width=35, font=("Segoe UI", 10))
        self.style_combo.set("-.")
        self.style_combo.pack(pady=(0, 12))

        ttk.Button(f, text="💾 Save", style="Accent.TButton", command=self._save).pack(fill="x", ipady=2)

    def _save(self):
        name = self.name_entry.get().strip()
        date = self.date_entry.get().strip()
        if not name or not date:
            messagebox.showwarning("Required", "Name and Date are required.")
            return
        self.result = {
            "name": name, "date": date,
            "color": self.color_entry.get().strip() or "#2196F3",
            "style": self.style_combo.get(),
        }
        self.top.destroy()


class _ResourceDialog:
    """Dialog for adding/editing team members."""
    
    ROLE_OPTIONS = ["Developer", "Designer", "PM", "QA", "Analyst", "Architect", "Lead", "Other"]
    
    def __init__(self, parent, title, initial=None):
        self.result = None
        self.top = tk.Toplevel(parent)
        self.top.title(title)
        self.top.geometry("400x350")
        self.top.grab_set()
        
        f = ttk.Frame(self.top, padding=20)
        f.pack(fill="both", expand=True)
        
        # Name
        ttk.Label(f, text="Name:").pack(anchor="w")
        self.name_entry = ttk.Entry(f, width=45)
        self.name_entry.pack(pady=(0, 8), ipady=2)
        
        # Role
        ttk.Label(f, text="Role:").pack(anchor="w")
        self.role_combo = ttk.Combobox(f, values=self.ROLE_OPTIONS, width=42)
        self.role_combo.pack(pady=(0, 8))
        
        # Department
        ttk.Label(f, text="Department:").pack(anchor="w")
        self.dept_entry = ttk.Entry(f, width=45)
        self.dept_entry.pack(pady=(0, 8), ipady=2)
        
        # Email
        ttk.Label(f, text="Email:").pack(anchor="w")
        self.email_entry = ttk.Entry(f, width=45)
        self.email_entry.pack(pady=(0, 8), ipady=2)
        
        # Allocation
        ttk.Label(f, text="Capacity % (0-100):").pack(anchor="w")
        self.alloc_entry = ttk.Entry(f, width=15)
        self.alloc_entry.insert(0, "100")
        self.alloc_entry.pack(anchor="w", pady=(0, 15), ipady=2)
        
        ttk.Button(f, text="💾 Save", style="Accent.TButton", command=self._save).pack(fill="x", ipady=2)
        
        if initial:
            self.name_entry.insert(0, initial.get("name", ""))
            self.role_combo.set(initial.get("role", ""))
            self.dept_entry.insert(0, initial.get("department", ""))
            self.email_entry.insert(0, initial.get("email", ""))
            self.alloc_entry.delete(0, tk.END)
            self.alloc_entry.insert(0, str(initial.get("allocation_pct", 100)))
    
    def _save(self):
        name = self.name_entry.get().strip()
        if not name:
            messagebox.showwarning("Required", "Name is required.")
            return
        
        try:
            alloc = float(self.alloc_entry.get().strip() or "100")
        except ValueError:
            alloc = 100.0
        
        self.result = {
            "name": name,
            "role": self.role_combo.get().strip(),
            "department": self.dept_entry.get().strip(),
            "email": self.email_entry.get().strip(),
            "allocation_pct": alloc,
        }
        self.top.destroy()


class _AssignmentDialog:
    """Dialog for assigning a team member to a project."""
    
    def __init__(self, parent, title, db_path):
        self.result = None
        self.db_path = db_path
        self.top = tk.Toplevel(parent)
        self.top.title(title)
        self.top.geometry("400x300")
        self.top.grab_set()
        
        f = ttk.Frame(self.top, padding=20)
        f.pack(fill="both", expand=True)
        
        # Team member selector
        ttk.Label(f, text="Team Member:").pack(anchor="w")
        self.member_combo = ttk.Combobox(f, width=42, state="readonly")
        self.member_combo.pack(pady=(0, 10))
        
        # Load team members
        self._resources = get_all_resources(db_path)
        self.member_combo["values"] = [f"{r.id} — {r.name} ({r.role})" for r in self._resources]
        if self._resources:
            self.member_combo.current(0)
        
        # Role in project
        ttk.Label(f, text="Role in Project:").pack(anchor="w")
        self.role_entry = ttk.Entry(f, width=45)
        self.role_entry.pack(pady=(0, 10), ipady=2)
        
        # Allocation
        ttk.Label(f, text="Allocation % for this project:").pack(anchor="w")
        self.alloc_entry = ttk.Entry(f, width=15)
        self.alloc_entry.insert(0, "100")
        self.alloc_entry.pack(anchor="w", pady=(0, 10), ipady=2)
        
        # Notes
        ttk.Label(f, text="Notes:").pack(anchor="w")
        self.notes_entry = ttk.Entry(f, width=45)
        self.notes_entry.pack(pady=(0, 15), ipady=2)
        
        ttk.Button(f, text="💾 Assign", style="Accent.TButton", command=self._save).pack(fill="x", ipady=2)
    
    def _save(self):
        idx = self.member_combo.current()
        if idx < 0:
            messagebox.showwarning("Select", "Please select a team member.")
            return
        
        try:
            alloc = float(self.alloc_entry.get().strip() or "100")
        except ValueError:
            alloc = 100.0
        
        self.result = {
            "resource_id": self._resources[idx].id,
            "role_in_project": self.role_entry.get().strip(),
            "allocation_pct": alloc,
            "notes": self.notes_entry.get().strip(),
        }
        self.top.destroy()


class _MilestoneTaskDialog:
    """Dialog to view and manage milestone tasks with status and attachments."""
    
    def __init__(self, parent: tk.Tk, milestone_name: str, milestone_id: int, 
                 project_name: str, milestone_date: datetime.date, 
                 can_edit: bool, username: str, db_path: pathlib.Path | None = None):
        self.milestone_id = milestone_id
        self.can_edit = can_edit
        self.username = username
        self.db_path = db_path
        
        self.top = tk.Toplevel(parent)
        self.top.title(f"Tasks: {milestone_name}")
        self.top.geometry("1100x650")
        self.top.grab_set()
        
        # Header
        header = ttk.Frame(self.top, padding=20)
        header.pack(fill="x")
        
        ttk.Label(header, text=f"📌 {milestone_name}", 
                  font=("Segoe UI", 16, "bold")).pack(anchor="w")
        ttk.Label(header, text=f"📁 Project: {project_name}", 
                  font=("Segoe UI", 10)).pack(anchor="w")
        ttk.Label(header, text=f"📅 Date: {milestone_date.strftime('%b %d, %Y')}", 
                  font=("Segoe UI", 10)).pack(anchor="w")
        
        ttk.Separator(self.top, orient="horizontal").pack(fill="x", pady=10)
        
        # Tasks frame
        tasks_frame = ttk.Frame(self.top, padding=(20, 0, 20, 10))
        tasks_frame.pack(fill="both", expand=True)
        
        # Load tasks
        self.tasks = db.get_milestone_tasks_with_status(milestone_id, db_path)
        
        if not self.tasks:
            ttk.Label(tasks_frame, text="No tasks defined for this milestone.",
                     font=("Segoe UI", 12), foreground="#999999").pack(pady=50)
        else:
            # Create scrollable canvas
            self.canvas = tk.Canvas(tasks_frame, bg="#F3F3F3", highlightthickness=0)
            scrollbar = ttk.Scrollbar(tasks_frame, orient="vertical", command=self.canvas.yview)
            self.scrollable_frame = ttk.Frame(self.canvas)
            
            # Create window in canvas
            self.canvas_window = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
            
            # Configure scrolling
            def configure_scroll_region(event=None):
                # Update scroll region to encompass the scrollable frame
                self.canvas.configure(scrollregion=self.canvas.bbox("all"))
                # Make the canvas window (scrollable_frame) match canvas width
                canvas_width = self.canvas.winfo_width()
                if canvas_width > 1:  # Only update if canvas has been rendered
                    self.canvas.itemconfig(self.canvas_window, width=canvas_width)
            
            self.scrollable_frame.bind("<Configure>", configure_scroll_region)
            self.canvas.bind("<Configure>", configure_scroll_region)
            
            # Configure scrollbar
            self.canvas.configure(yscrollcommand=scrollbar.set)
            
            # Pack canvas and scrollbar
            self.canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            
            # Enable mouse wheel scrolling
            self._bind_mousewheel()
            
            # Store status comboboxes and task data
            self.task_widgets = []
            
            for idx, task in enumerate(self.tasks):
                self._create_task_row(self.scrollable_frame, task, idx)
            
            # Update scroll region after adding all tasks
            self.top.update_idletasks()
            configure_scroll_region()
        
        # Buttons
        btn_frame = ttk.Frame(self.top, padding=20)
        btn_frame.pack(fill="x")
        
        if self.can_edit and self.tasks:
            ttk.Button(btn_frame, text="💾 Save Changes", style="Accent.TButton",
                      command=self._save_changes).pack(side="right", padx=5)
        
        ttk.Button(btn_frame, text="Close", command=self.top.destroy).pack(side="right", padx=5)
    
    def _bind_mousewheel(self):
        """Bind mouse wheel scrolling to canvas."""
        def _on_mousewheel(event):
            # Windows and MacOS
            if event.delta:
                self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            # Linux
            elif event.num == 4:
                self.canvas.yview_scroll(-1, "units")
            elif event.num == 5:
                self.canvas.yview_scroll(1, "units")
        
        # Bind to canvas
        self.canvas.bind_all("<MouseWheel>", _on_mousewheel)  # Windows/MacOS
        self.canvas.bind_all("<Button-4>", _on_mousewheel)    # Linux scroll up
        self.canvas.bind_all("<Button-5>", _on_mousewheel)    # Linux scroll down
        
        # Also bind to scrollable frame for better UX
        self.scrollable_frame.bind_all("<MouseWheel>", _on_mousewheel)
        self.scrollable_frame.bind_all("<Button-4>", _on_mousewheel)
        self.scrollable_frame.bind_all("<Button-5>", _on_mousewheel)
        
        # Unbind when window closes to prevent issues
        def _unbind_mousewheel():
            try:
                self.canvas.unbind_all("<MouseWheel>")
                self.canvas.unbind_all("<Button-4>")
                self.canvas.unbind_all("<Button-5>")
                self.scrollable_frame.unbind_all("<MouseWheel>")
                self.scrollable_frame.unbind_all("<Button-4>")
                self.scrollable_frame.unbind_all("<Button-5>")
            except:
                pass
        
        self.top.protocol("WM_DELETE_WINDOW", lambda: (_unbind_mousewheel(), self.top.destroy()))
    
    def _create_task_row(self, parent: ttk.Frame, task: dict, idx: int):
        """Create a row for each task with status dropdown and attachment button."""
        
        # Define status colors
        status_colors = {
            "Completed": "#27AE60",      # Green
            "WIP": "#F39C12",            # Yellow/Orange
            "Yet to Start": "#E67E22",   # Orange
            "Not Applicable": "#95A5A6"  # Grey
        }
        
        # Get color for current status
        status_color = status_colors.get(task["status"], "#95A5A6")
        
        # Outer container for colored border effect
        border_frame = tk.Frame(parent, bg=status_color, bd=0)
        border_frame.pack(fill="x", padx=5, pady=5)
        
        # Task card - inner frame with left padding for colored border
        card = tk.Frame(border_frame, bg="white", bd=1,
                    highlightbackground="#E0E0E0", highlightthickness=1)
        card.pack(fill="both", expand=True, padx=(5, 0))  # Left padding creates colored border
        
        # Add padding inside card
        inner_frame = tk.Frame(card, bg="white")
        inner_frame.pack(fill="both", expand=True, padx=12, pady=10)
        
        # Configure grid columns to distribute space properly
        inner_frame.grid_columnconfigure(1, weight=1)  # Task name column expands
        
        # Status indicator dot
        status_indicator = tk.Label(inner_frame, text="●", 
                                font=("Segoe UI", 14), bg="white", fg=status_color)
        status_indicator.grid(row=0, column=0, sticky="w", padx=(0, 5), pady=(0, 8))
        
        # Task number and name
        task_header = tk.Frame(inner_frame, bg="white")
        task_header.grid(row=0, column=1, columnspan=4, sticky="ew", pady=(0, 8))
        
        tk.Label(task_header, text=f"{idx + 1}.", 
                font=("Segoe UI", 10, "bold"), bg="white", fg="#0067C0").pack(side="left", padx=(0, 5))
        
        tk.Label(task_header, text=task["task_name"], 
                font=("Segoe UI", 10), bg="white", anchor="w").pack(side="left", fill="x", expand=True)
        
        # Row 1: Status label and dropdown
        tk.Label(inner_frame, text="Status:", font=("Segoe UI", 9), 
                bg="white", fg="#666666").grid(
                    row=1, column=0, columnspan=1, sticky="w", padx=(0, 5))
        
        status_var = tk.StringVar(value=task["status"])
        
        # Custom callback to update color when status changes
        def on_status_change(*args):
            new_color = status_colors.get(status_var.get(), "#95A5A6")
            border_frame.configure(bg=new_color)
            status_indicator.configure(fg=new_color)
        
        status_var.trace_add("write", on_status_change)
        
        status_combo = ttk.Combobox(
            inner_frame,
            textvariable=status_var,
            values=["Yet to Start", "WIP", "Completed", "Not Applicable"],
            state="readonly" if self.can_edit else "disabled",
            width=15,
            font=("Segoe UI", 9)
        )
        status_combo.grid(row=1, column=1, sticky="w", padx=(0, 15))
        
        # Row 1: Attachment info and buttons
        col_offset = 2
        
        if task["attachment_path"]:
            filename = pathlib.Path(task["attachment_path"]).name
            # Truncate long filenames
            if len(filename) > 35:
                display_name = filename[:32] + "..."
            else:
                display_name = filename
            
            attach_label = tk.Label(inner_frame, text=f"📎 {display_name}", 
                    font=("Segoe UI", 9), bg="white", fg="#0067C0",
                    cursor="hand2")
            attach_label.grid(row=1, column=col_offset, sticky="w", padx=(0, 10))
            attach_label.bind("<Button-1>", lambda e, t=task: self._view_attachment(t))
            col_offset += 1
            
            # View button
            ttk.Button(
                inner_frame,
                text="👁️ View",
                command=lambda t=task: self._view_attachment(t),
                width=8
            ).grid(row=1, column=col_offset, sticky="w", padx=2)
            col_offset += 1
            
            # Remove button
            if self.can_edit:
                ttk.Button(
                    inner_frame,
                    text="🗑️ Remove",
                    command=lambda t=task: self._remove_attachment(t),
                    width=10
                ).grid(row=1, column=col_offset, sticky="w", padx=2)
        else:
            # Attach button
            if self.can_edit:
                ttk.Button(
                    inner_frame,
                    text="📎 Attach File",
                    command=lambda t=task: self._add_attachment(t),
                    width=13
                ).grid(row=1, column=col_offset, sticky="w", padx=2)
        
        # Store references
        self.task_widgets.append({
            "task_id": task["id"],
            "status_var": status_var,
            "attachment_path": task["attachment_path"]
        })

    def _add_attachment(self, task: dict):
        """Add an attachment to a task."""
        file_path = filedialog.askopenfilename(
            title=f"Select file for task: {task['task_name']}",
            filetypes=[
                ("All files", "*.*"),
                ("PDF files", "*.pdf"),
                ("Word files", "*.docx"),
                ("Excel files", "*.xlsx"),
                ("Images", "*.png *.jpg *.jpeg")
            ]
        )
        
        if file_path:
            # Copy file to attachments folder
            import shutil
            attachments_dir = db.DEFAULT_DB_DIR / "attachments"
            attachments_dir.mkdir(exist_ok=True)
            
            filename = pathlib.Path(file_path).name
            dest_path = attachments_dir / f"task{task['id']}_{filename}"
            
            try:
                shutil.copy2(file_path, dest_path)
                
                # Update database
                db.update_task_attachment(task["id"], str(dest_path), self.username, self.db_path)
                
                messagebox.showinfo("Success", f"Attachment added:\n{filename}")
                self.top.destroy()  # Close and will be reopened to refresh
            except Exception as e:
                messagebox.showerror("Error", f"Failed to attach file:\n{str(e)}")
    
    def _view_attachment(self, task: dict):
        """View an existing attachment."""
        if task["attachment_path"] and pathlib.Path(task["attachment_path"]).exists():
            import os
            import platform
            
            try:
                if platform.system() == 'Windows':
                    os.startfile(task["attachment_path"])
                elif platform.system() == 'Darwin':  # macOS
                    os.system(f'open "{task["attachment_path"]}"')
                else:  # Linux
                    os.system(f'xdg-open "{task["attachment_path"]}"')
            except Exception as e:
                messagebox.showerror("Error", f"Could not open file:\n{str(e)}")
        else:
            messagebox.showwarning("Not Found", "Attachment file not found.")
    
    def _remove_attachment(self, task: dict):
        """Remove an attachment from a task."""
        filename = pathlib.Path(task["attachment_path"]).name if task["attachment_path"] else "this file"
        
        if messagebox.askyesno("Confirm Removal", 
                               f"Remove attachment:\n{filename}\n\nThis will permanently delete the file."):
            try:
                db.remove_task_attachment(task["id"], self.username, self.db_path)
                messagebox.showinfo("Removed", "Attachment removed successfully.")
                self.top.destroy()  # Close and will be reopened to refresh
            except Exception as e:
                messagebox.showerror("Error", f"Failed to remove attachment:\n{str(e)}")
    
    def _save_changes(self):
        """Save all status changes."""
        for widget in self.task_widgets:
            new_status = widget["status_var"].get()
            db.update_task_status(widget["task_id"], new_status, self.username, self.db_path)
        
        messagebox.showinfo("Saved", "Task statuses updated successfully!")
        self.top.destroy()


class _ProjectDetailsDialog:
    """Comprehensive project details dialog showing timeline, activities, QCTP, and team."""
    
    def __init__(self, parent: tk.Tk, project: Project, db_path: pathlib.Path | None,
                 can_edit: bool, username: str):
        self.project = project
        self.db_path = db_path
        self.can_edit = can_edit
        self.username = username
        self.today = datetime.date.today()
        
        # Create dialog window
        self.top = tk.Toplevel(parent)
        self.top.title(f"Project Details: {project.name}")
        self.top.geometry("1200x800")
        self.top.grab_set()
        
        # Apply styling
        self.top.configure(bg="#F3F3F3")
        
        # Main container with scrolling
        self._build_ui()
    
    def _build_ui(self):
        """Build the complete UI."""
        # ══════════════════════════════════════════════════════════════════
        # HEADER
        # ══════════════════════════════════════════════════════════════════
        header_frame = tk.Frame(self.top, bg="#0067C0", padx=20, pady=15)
        header_frame.pack(fill="x")
        
        # Project name and status
        title_frame = tk.Frame(header_frame, bg="#0067C0")
        title_frame.pack(fill="x")
        
        tk.Label(title_frame, text=f"📁 {self.project.name}",
                 font=("Segoe UI", 18, "bold"), bg="#0067C0", fg="white").pack(side="left")
        
        # Status badge
        status = self.project.computed_status(self.today)
        status_colors = {
            "on-track": ("#16A34A", "ON TRACK"),
            "at-risk": ("#EAB308", "AT RISK"),
            "overdue": ("#DC2626", "OVERDUE"),
        }
        badge_color, badge_text = status_colors.get(status, ("#666666", status.upper()))
        
        status_badge = tk.Label(title_frame, text=f"  {badge_text}  ",
                                 font=("Segoe UI", 10, "bold"),
                                 bg=badge_color, fg="white", padx=10, pady=2)
        status_badge.pack(side="left", padx=15)
        
        # Progress percentage
        progress = self.project.progress(self.today)
        tk.Label(title_frame, text=f"{int(progress * 100)}% Complete",
                 font=("Segoe UI", 11), bg="#0067C0", fg="white").pack(side="right")
        
        # Date range
        tk.Label(header_frame, 
                 text=f"📅 {self.project.start_date.strftime('%B %d, %Y')} → {self.project.end_date.strftime('%B %d, %Y')}",
                 font=("Segoe UI", 10), bg="#0067C0", fg="#E0E0E0").pack(anchor="w", pady=(5, 0))
        
        # ══════════════════════════════════════════════════════════════════
        # MAIN CONTENT - Scrollable
        # ══════════════════════════════════════════════════════════════════
        # Create canvas for scrolling
        canvas = tk.Canvas(self.top, bg="#F3F3F3", highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.top, orient="vertical", command=canvas.yview)
        self.scrollable_frame = ttk.Frame(canvas)
        
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Enable mouse wheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # Unbind on close
        def on_close():
            canvas.unbind_all("<MouseWheel>")
            self.top.destroy()
        self.top.protocol("WM_DELETE_WINDOW", on_close)
        
        # Content padding
        content = ttk.Frame(self.scrollable_frame, padding=20)
        content.pack(fill="both", expand=True)
        
        # ══════════════════════════════════════════════════════════════════
        # ROW 1: Overview Cards
        # ═════════════════════��════════════════════════════════════════════
        self._build_overview_section(content)
        
        # ══════════════════════════════════════════════════════════════════
        # ROW 2: Timeline & Activities (side by side)
        # ══════════════════════════════════════════════════════════════════
        row2 = ttk.Frame(content)
        row2.pack(fill="x", pady=(15, 0))
        
        # Left: Mini Timeline
        self._build_timeline_section(row2)
        
        # Right: Key Activities This Week
        self._build_activities_section(row2)
        
        # ══════════════════════════════════════════════════════════════════
        # ROW 3: QCTP & Team (side by side)
        # ══════════════════════════════════════════════════════════════════
        row3 = ttk.Frame(content)
        row3.pack(fill="x", pady=(15, 0))
        
        # Left: QCTP
        self._build_qctp_section(row3)
        
        # Right: Team Members
        self._build_team_section(row3)
        
        # ══════════════════════════════════════════════════════════════════
        # ROW 4: Milestones
        # ══════════════════════════════════════════════════════════════════
        self._build_milestones_section(content)
        
        # ══════════════════════════════════════════════════════════════════
        # FOOTER BUTTONS
        # ══════════════════════════════════════════════════════════════════
        btn_frame = ttk.Frame(content)
        btn_frame.pack(fill="x", pady=(20, 0))
        
        ttk.Button(btn_frame, text="Close", command=on_close).pack(side="right", padx=5)
        if self.can_edit:
            ttk.Button(btn_frame, text="✏️ Edit Project", 
                      command=self._edit_project).pack(side="right", padx=5)

    def _build_overview_section(self, parent):
        """Build overview cards section."""
        frame = ttk.Frame(parent)
        frame.pack(fill="x")
        
        # Calculate metrics
        total_milestones = len(self.project.milestones)
        completed_milestones = sum(1 for ms in self.project.milestones if ms.is_complete())
        total_phases = len(self.project.phases)
        days_remaining = (self.project.end_date - self.today).days
        total_days = (self.project.end_date - self.project.start_date).days
        
        # Get team count
        team_count = 0
        try:
            with db._connect(self.db_path) as conn:
                row = conn.execute(
                    "SELECT COUNT(*) as cnt FROM project_assignments WHERE project_id = ?",
                    (self.project.project_id,)
                ).fetchone()
                team_count = row["cnt"] if row else 0
        except:
            pass
        
        cards_data = [
            ("📊", "Progress", f"{int(self.project.progress(self.today) * 100)}%", "#0067C0"),
            ("📅", "Days Remaining", str(max(0, days_remaining)), "#6B7280"),
            ("🎯", "Milestones", f"{completed_milestones}/{total_milestones}", "#16A34A"),
            ("📐", "Phases", str(total_phases), "#8B5CF6"),
            ("👥", "Team Members", str(team_count), "#EC4899"),
            ("⏱️", "Duration", f"{total_days} days", "#F59E0B"),
        ]
        
        for icon, label, value, color in cards_data:
            card = tk.Frame(frame, bg="white", bd=1,
                           highlightbackground="#E0E0E0", highlightthickness=1)
            card.pack(side="left", fill="x", expand=True, padx=3, ipady=10, ipadx=15)
            
            tk.Label(card, text=icon, font=("Segoe UI", 20), bg="white").pack(side="left", padx=(10, 8))
            text_frame = tk.Frame(card, bg="white")
            text_frame.pack(side="left", fill="x", expand=True)
            tk.Label(text_frame, text=value, font=("Segoe UI", 16, "bold"),
                    bg="white", fg=color).pack(anchor="w")
            tk.Label(text_frame, text=label, font=("Segoe UI", 9),
                    bg="white", fg="#666666").pack(anchor="w")

    def _build_timeline_section(self, parent):
        """Build mini timeline visualization."""
        frame = ttk.LabelFrame(parent, text="  📈 Project Timeline  ", padding=10)
        frame.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        # Create mini Gantt chart
        fig, ax = plt.subplots(figsize=(6, 3))
        fig.patch.set_facecolor('#FFFFFF')
        ax.set_facecolor('#FAFAFA')
        
        project = self.project
        y = 1
        
        # Main project bar
        ax.barh(y, (project.end_date - project.start_date).days,
                left=project.start_date, height=0.4,
                color="#0067C0", alpha=0.3)
        
        # Progress bar
        progress = project.progress(self.today)
        if progress > 0:
            elapsed_days = (project.end_date - project.start_date).days * progress
            ax.barh(y, elapsed_days, left=project.start_date,
                    height=0.4, color="#0067C0", alpha=0.8)
        
        # Phases
        for p_idx, phase in enumerate(project.phases):
            phase_color = PHASE_PALETTE[p_idx % len(PHASE_PALETTE)]
            ax.barh(y - 0.5, (phase.end_date - phase.start_date).days,
                    left=phase.start_date, height=0.25,
                    color=phase_color, alpha=0.8, label=phase.name)
        
        # Milestones
        for ms in project.milestones:
            color = "#27AE60" if ms.is_complete() else "#E74C3C"
            ax.scatter(ms.date, y, marker="D", s=60, color=color, zorder=5)
            ax.annotate(ms.name, (ms.date, y), xytext=(0, 15),
                       textcoords="offset points", fontsize=7,
                       ha="center", rotation=45)
        
        # Today line
        ax.axvline(self.today, color="#DC2626", linewidth=1.5, linestyle="--", label="Today")
        
        ax.set_ylim(0, 2)
        ax.set_yticks([])
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%b %Y'))
        ax.xaxis.set_major_locator(mdates.MonthLocator())
        fig.autofmt_xdate(rotation=30)
        ax.grid(axis="x", linestyle=":", alpha=0.5)
        
        # Legend
        ax.legend(loc="upper right", fontsize=7, framealpha=0.9)
        
        fig.tight_layout()
        
        canvas = FigureCanvasTkAgg(fig, master=frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)
        
        # Store figure for cleanup
        self._timeline_fig = fig

    def _build_activities_section(self, parent):
        """Build key activities this week section."""
        frame = ttk.LabelFrame(parent, text="  📋 Key Activities This Week  ", padding=10)
        frame.pack(side="right", fill="both", expand=True, padx=(10, 0))
        
        # Calculate week range
        week_start = self.today - datetime.timedelta(days=self.today.weekday())
        week_end = week_start + datetime.timedelta(days=6)
        
        activities = []
        
        # Find milestones this week
        for ms in self.project.milestones:
            if week_start <= ms.date <= week_end:
                status = "✅ Complete" if ms.is_complete() else "⏳ Pending"
                activities.append({
                    "type": "Milestone",
                    "name": ms.name,
                    "date": ms.date,
                    "status": status,
                    "icon": "🎯"
                })
        
        # Find tasks due this week
        try:
            with db._connect(self.db_path) as conn:
                for ms in self.project.milestones:
                    if week_start <= ms.date <= week_end:
                        tasks = conn.execute(
                            """SELECT task_name, status FROM milestone_tasks 
                               WHERE milestone_id = ? AND status != 'Completed'""",
                            (ms.milestone_id,)
                        ).fetchall()
                        for task in tasks:
                            activities.append({
                                "type": "Task",
                                "name": task["task_name"][:40] + "..." if len(task["task_name"]) > 40 else task["task_name"],
                                "date": ms.date,
                                "status": task["status"],
                                "icon": "📝"
                            })
        except:
            pass
        
        # Find phases starting/ending this week
        for phase in self.project.phases:
            if week_start <= phase.start_date <= week_end:
                activities.append({
                    "type": "Phase Start",
                    "name": phase.name,
                    "date": phase.start_date,
                    "status": "Starting",
                    "icon": "🚀"
                })
            if week_start <= phase.end_date <= week_end:
                activities.append({
                    "type": "Phase End",
                    "name": phase.name,
                    "date": phase.end_date,
                    "status": "Ending",
                    "icon": "🏁"
                })
        
        # Sort by date
        activities.sort(key=lambda x: x["date"])
        
        if activities:
            # Create treeview
            cols = ("icon", "type", "name", "date", "status")
            tree = ttk.Treeview(frame, columns=cols, show="headings", height=6)
            tree.heading("icon", text="")
            tree.heading("type", text="Type")
            tree.heading("name", text="Activity")
            tree.heading("date", text="Date")
            tree.heading("status", text="Status")
            tree.column("icon", width=30)
            tree.column("type", width=80)
            tree.column("name", width=150)
            tree.column("date", width=80)
            tree.column("status", width=80)
            
            for act in activities:
                tree.insert("", "end", values=(
                    act["icon"],
                    act["type"],
                    act["name"],
                    act["date"].strftime("%b %d"),
                    act["status"]
                ))
            
            tree.pack(fill="both", expand=True)
        else:
            ttk.Label(frame, text="No activities scheduled for this week",
                     font=("Segoe UI", 10), foreground="#999999").pack(expand=True)

    def _build_qctp_section(self, parent):
        """Build QCTP points section."""
        frame = ttk.LabelFrame(parent, text="  📊 QCTP Points  ", padding=10)
        frame.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        # Load QCTP data
        qctp_data = {"quality": "", "cost": "", "time": "", "performance": ""}
        try:
            with db._connect(self.db_path) as conn:
                row = conn.execute(
                    "SELECT quality, cost, time, performance FROM qctp WHERE project_id = ?",
                    (self.project.project_id,)
                ).fetchone()
                if row:
                    qctp_data = {
                        "quality": row["quality"] or "",
                        "cost": row["cost"] or "",
                        "time": row["time"] or "",
                        "performance": row["performance"] or ""
                    }
        except:
            pass
        
        # Create 2x2 grid
        grid = ttk.Frame(frame)
        grid.pack(fill="both", expand=True)
        
        quadrants = [
            ("Quality", "🎯", qctp_data["quality"], 0, 0),
            ("Cost", "💰", qctp_data["cost"], 0, 1),
            ("Time", "⏱️", qctp_data["time"], 1, 0),
            ("Performance", "📈", qctp_data["performance"], 1, 1),
        ]
        
        grid.grid_columnconfigure(0, weight=1)
        grid.grid_columnconfigure(1, weight=1)
        grid.grid_rowconfigure(0, weight=1)
        grid.grid_rowconfigure(1, weight=1)
        
        for name, icon, value, row, col in quadrants:
            card = tk.Frame(grid, bg="white", bd=1,
                           highlightbackground="#E0E0E0", highlightthickness=1)
            card.grid(row=row, column=col, padx=3, pady=3, sticky="nsew")
            
            header = tk.Frame(card, bg="#F5F5F5")
            header.pack(fill="x")
            tk.Label(header, text=f"{icon} {name}", font=("Segoe UI", 9, "bold"),
                    bg="#F5F5F5", fg="#333333", anchor="w", padx=8, pady=4).pack(fill="x")
            
            # Content
            content_text = value[:150] + "..." if len(value) > 150 else value
            if not content_text:
                content_text = "No data entered"
            tk.Label(card, text=content_text, font=("Segoe UI", 9),
                    bg="white", fg="#666666" if value else "#999999",
                    wraplength=180, justify="left", anchor="nw",
                    padx=8, pady=5).pack(fill="both", expand=True)

    def _build_team_section(self, parent):
        """Build team members section."""
        frame = ttk.LabelFrame(parent, text="  👥 Team Members  ", padding=10)
        frame.pack(side="right", fill="both", expand=True, padx=(10, 0))
        
        # Load team assignments
        team_members = []
        try:
            assignments = get_project_assignments(self.project.project_id, self.db_path)
            team_members = assignments
        except:
            pass
        
        if team_members:
            cols = ("name", "role", "allocation")
            tree = ttk.Treeview(frame, columns=cols, show="headings", height=5)
            tree.heading("name", text="Name")
            tree.heading("role", text="Role")
            tree.heading("allocation", text="Allocation")
            tree.column("name", width=150)
            tree.column("role", width=100)
            tree.column("allocation", width=80)
            
            for member in team_members:
                tree.insert("", "end", values=(
                    member["name"],
                    member["role_in_project"] or member["role"],
                    f"{member['allocation_pct']:.0f}%"
                ))
            
            tree.pack(fill="both", expand=True)
        else:
            ttk.Label(frame, text="No team members assigned",
                     font=("Segoe UI", 10), foreground="#999999").pack(expand=True)
            
            if self.can_edit:
                ttk.Button(frame, text="➕ Assign Team Members",
                          command=self._go_to_assignments).pack(pady=10)

    def _build_milestones_section(self, parent):
        """Build milestones summary section."""
        frame = ttk.LabelFrame(parent, text="  🎯 Milestones Summary  ", padding=10)
        frame.pack(fill="x", pady=(15, 0))
        
        if not self.project.milestones:
            ttk.Label(frame, text="No milestones defined for this project",
                     font=("Segoe UI", 10), foreground="#999999").pack(pady=20)
            return
        
        # Create treeview
        cols = ("name", "date", "status", "tasks", "progress")
        tree = ttk.Treeview(frame, columns=cols, show="headings", height=min(8, len(self.project.milestones)))
        tree.heading("name", text="Milestone")
        tree.heading("date", text="Date")
        tree.heading("status", text="Status")
        tree.heading("tasks", text="Tasks")
        tree.heading("progress", text="Progress")
        tree.column("name", width=150)
        tree.column("date", width=100)
        tree.column("status", width=100)
        tree.column("tasks", width=80)
        tree.column("progress", width=100)
        
        for ms in sorted(self.project.milestones, key=lambda m: m.date):
            # Calculate task stats
            total_tasks = sum(ms.task_statuses.values()) if ms.task_statuses else 0
            completed_tasks = ms.task_statuses.get("Completed", 0) if ms.task_statuses else 0
            
            # Determine status
            if ms.is_complete():
                status = "✅ Complete"
                tag = "complete"
            elif ms.date < self.today:
                status = "🔴 Overdue"
                tag = "overdue"
            elif ms.date <= self.today + datetime.timedelta(days=7):
                status = "🟡 Due Soon"
                tag = "soon"
            else:
                status = "⏳ Upcoming"
                tag = "upcoming"
            
            # Progress bar text
            if total_tasks > 0:
                pct = int((completed_tasks / total_tasks) * 100)
                progress = f"{'█' * (pct // 10)}{'░' * (10 - pct // 10)} {pct}%"
            else:
                progress = "No tasks"
            
            tree.insert("", "end", values=(
                ms.name,
                ms.date.strftime("%Y-%m-%d"),
                status,
                f"{completed_tasks}/{total_tasks}",
                progress
            ), tags=(tag,))
        
        # Configure tags
        tree.tag_configure("complete", foreground="#15803D")
        tree.tag_configure("overdue", foreground="#DC2626")
        tree.tag_configure("soon", foreground="#B45309")
        tree.tag_configure("upcoming", foreground="#6B7280")
        
        tree.pack(fill="x")

    def _edit_project(self):
        """Open project edit dialog."""
        self.top.destroy()
        # Note: This would need to trigger the edit dialog in MainApp
        # For now, just close this dialog

    def _go_to_assignments(self):
        """Navigate to assignments tab."""
        self.top.destroy()
        # Note: This would need to switch to Resources tab in MainApp


class _ActivityDialog:
    """Dialog for adding/editing weekly activities."""

    STATUS_OPTIONS = ["WIP", "Completed"]

    def __init__(self, parent, title, db_path, week=None, year=None, initial=None):
        self.result = None
        self.db_path = db_path
        self.top = tk.Toplevel(parent)
        self.top.title(title)
        self.top.geometry("550x580")
        self.top.grab_set()

        # ── Try to import tkcalendar ────────────────────────────────────
        self._has_calendar = False
        try:
            from tkcalendar import DateEntry
            self._has_calendar = True
        except ImportError:
            self._has_calendar = False

        f = ttk.Frame(self.top, padding=20)
        f.pack(fill="both", expand=True)

        # ── Activity Name ────────────────────────────────────────────────
        ttk.Label(f, text="Activity Name: *").pack(anchor="w")
        self.name_entry = ttk.Entry(f, width=60)
        self.name_entry.pack(pady=(0, 8), ipady=2)

        # ── Date row ────────────────────────────────────────────────────
        date_frame = ttk.Frame(f)
        date_frame.pack(fill="x", pady=(0, 8))

        ttk.Label(date_frame, text="Start Date: *").pack(side="left")

        if self._has_calendar:
            from tkcalendar import DateEntry
            self.start_date_picker = DateEntry(
                date_frame, width=14, font=("Segoe UI", 9),
                background="#0067C0", foreground="white", headersbackground="#0067C0",
                headersforeground="white", selectbackground="#005A9E",
                date_pattern="yyyy-mm-dd", state="readonly"
            )
            self.start_date_picker.pack(side="left", padx=(4, 15))
        else:
            # Fallback: 3 Comboboxes for Year / Month / Day
            self._start_year_var = tk.StringVar()
            self._start_month_var = tk.StringVar()
            self._start_day_var = tk.StringVar()

            current_year = datetime.date.today().year
            years = [str(y) for y in range(current_year - 2, current_year + 5)]
            months = [f"{m:02d}" for m in range(1, 13)]
            days = [f"{d:02d}" for d in range(1, 32)]

            self._start_year_cb = ttk.Combobox(date_frame, textvariable=self._start_year_var,
                                                values=years, width=5, state="readonly",
                                                font=("Segoe UI", 9))
            self._start_year_cb.pack(side="left", padx=(4, 1))
            self._start_year_cb.set(str(current_year))

            ttk.Label(date_frame, text="-", font=("Segoe UI", 9)).pack(side="left")

            self._start_month_cb = ttk.Combobox(date_frame, textvariable=self._start_month_var,
                                                 values=months, width=3, state="readonly",
                                                 font=("Segoe UI", 9))
            self._start_month_cb.pack(side="left", padx=1)
            self._start_month_cb.set(f"{datetime.date.today().month:02d}")
            self._start_month_cb.bind("<<ComboboxSelected>>",
                                       lambda e: self._update_days("start"))

            ttk.Label(date_frame, text="-", font=("Segoe UI", 9)).pack(side="left")

            self._start_day_cb = ttk.Combobox(date_frame, textvariable=self._start_day_var,
                                               values=days, width=3, state="readonly",
                                               font=("Segoe UI", 9))
            self._start_day_cb.pack(side="left", padx=(1, 15))
            self._start_day_cb.set(f"{datetime.date.today().day:02d}")

            self._start_year_cb.bind("<<ComboboxSelected>>",
                                      lambda e: self._update_days("start"))

        ttk.Label(date_frame, text="End Date: *").pack(side="left")

        if self._has_calendar:
            self.end_date_picker = DateEntry(
                date_frame, width=14, font=("Segoe UI", 9),
                background="#0067C0", foreground="white", headersbackground="#0067C0",
                headersforeground="white", selectbackground="#005A9E",
                date_pattern="yyyy-mm-dd", state="readonly"
            )
            self.end_date_picker.pack(side="left", padx=(4, 0))
        else:
            self._end_year_var = tk.StringVar()
            self._end_month_var = tk.StringVar()
            self._end_day_var = tk.StringVar()

            current_year = datetime.date.today().year
            years = [str(y) for y in range(current_year - 2, current_year + 5)]
            months = [f"{m:02d}" for m in range(1, 13)]
            days = [f"{d:02d}" for d in range(1, 32)]

            self._end_year_cb = ttk.Combobox(date_frame, textvariable=self._end_year_var,
                                              values=years, width=5, state="readonly",
                                              font=("Segoe UI", 9))
            self._end_year_cb.pack(side="left", padx=(4, 1))
            self._end_year_cb.set(str(current_year))

            ttk.Label(date_frame, text="-", font=("Segoe UI", 9)).pack(side="left")

            self._end_month_cb = ttk.Combobox(date_frame, textvariable=self._end_month_var,
                                               values=months, width=3, state="readonly",
                                               font=("Segoe UI", 9))
            self._end_month_cb.pack(side="left", padx=1)
            self._end_month_cb.set(f"{datetime.date.today().month:02d}")
            self._end_month_cb.bind("<<ComboboxSelected>>",
                                     lambda e: self._update_days("end"))

            ttk.Label(date_frame, text="-", font=("Segoe UI", 9)).pack(side="left")

            self._end_day_cb = ttk.Combobox(date_frame, textvariable=self._end_day_var,
                                             values=days, width=3, state="readonly",
                                             font=("Segoe UI", 9))
            self._end_day_cb.pack(side="left", padx=(1, 0))
            self._end_day_cb.set(f"{datetime.date.today().day:02d}")

            self._end_year_cb.bind("<<ComboboxSelected>>",
                                    lambda e: self._update_days("end"))

        # ── Time Taken (Hours & Minutes dropdowns) ──────────────────────
        ttk.Label(f, text="Time Taken:").pack(anchor="w")
        time_frame = ttk.Frame(f)
        time_frame.pack(fill="x", pady=(0, 8))

        # Hours dropdown (0–99)
        self._time_hours_var = tk.StringVar(value="0")
        ttk.Label(time_frame, text="Hours:", font=("Segoe UI", 9)).pack(side="left")
        self.time_hours_cb = ttk.Combobox(
            time_frame, textvariable=self._time_hours_var,
            values=[str(h) for h in range(0, 100)],
            width=4, state="readonly", font=("Segoe UI", 9)
        )
        self.time_hours_cb.pack(side="left", padx=(4, 12))
        self.time_hours_cb.set("0")

        # Minutes dropdown (0, 5, 10, ... 55)
        self._time_mins_var = tk.StringVar(value="0")
        ttk.Label(time_frame, text="Minutes:", font=("Segoe UI", 9)).pack(side="left")
        self.time_mins_cb = ttk.Combobox(
            time_frame, textvariable=self._time_mins_var,
            values=[str(m) for m in range(0, 60, 5)],
            width=4, state="readonly", font=("Segoe UI", 9)
        )
        self.time_mins_cb.pack(side="left", padx=(4, 0))
        self.time_mins_cb.set("0")

        # ── Members Involved (multi-select from resources) ──────────────
        ttk.Label(f, text="Members Involved:").pack(anchor="w")
        members_frame = ttk.Frame(f)
        members_frame.pack(fill="x", pady=(0, 8))

        self.members_listbox = tk.Listbox(members_frame, selectmode="multiple",
                                           height=4, font=("Segoe UI", 9),
                                           exportselection=False)
        members_scroll = ttk.Scrollbar(members_frame, orient="vertical",
                                        command=self.members_listbox.yview)
        self.members_listbox.configure(yscrollcommand=members_scroll.set)
        self.members_listbox.pack(side="left", fill="x", expand=True)
        members_scroll.pack(side="right", fill="y")

        # Populate members from resources table
        self._resources = []
        try:
            from timeline_tool.resources import get_all_resources
            self._resources = get_all_resources(db_path)
            for res in self._resources:
                self.members_listbox.insert("end", f"{res.name} ({res.role})")
        except Exception:
            self.members_listbox.insert("end", "(No team members found)")

        # ── Hard Points ─────────────────────────────────────────────────
        ttk.Label(f, text="Hard Points / Challenges:").pack(anchor="w")
        self.hard_text = tk.Text(f, height=3, font=("Segoe UI", 9), wrap="word")
        self.hard_text.pack(fill="x", pady=(0, 8))

        # ── Status ──────────────────────────────────────────────────────
        status_frame = ttk.Frame(f)
        status_frame.pack(fill="x", pady=(0, 8))

        ttk.Label(status_frame, text="Status: *").pack(side="left")
        self.status_combo = ttk.Combobox(status_frame, values=self.STATUS_OPTIONS,
                                          width=15, state="readonly")
        self.status_combo.pack(side="left", padx=(8, 0))
        self.status_combo.set("WIP")

        # ── Attachment ──────────────────────────────────────────────────
        attach_frame = ttk.Frame(f)
        attach_frame.pack(fill="x", pady=(0, 12))

        ttk.Label(attach_frame, text="Attachment:").pack(side="left")
        self.attach_path_var = tk.StringVar(value="")
        self.attach_label = ttk.Label(attach_frame, textvariable=self.attach_path_var,
                                       font=("Segoe UI", 8), foreground="#666666")
        self.attach_label.pack(side="left", padx=(8, 8), fill="x", expand=True)
        ttk.Button(attach_frame, text="📎 Browse",
                   command=self._browse_attachment).pack(side="right")

        # ── Buttons ─────────────────────────────────────────────────────
        btn_frame = ttk.Frame(f)
        btn_frame.pack(fill="x")

        save_btn = tk.Button(btn_frame, text="💾 Save", font=("Segoe UI", 10, "bold"),
                              bg="#0067C0", fg="white", relief="flat", cursor="hand2",
                              activebackground="#005A9E", activeforeground="white",
                              command=self._save, padx=20, pady=5)
        save_btn.pack(side="right", padx=5)

        cancel_btn = tk.Button(btn_frame, text="Cancel", font=("Segoe UI", 10),
                                bg="#E0E0E0", relief="flat", cursor="hand2",
                                command=self.top.destroy, padx=20, pady=5)
        cancel_btn.pack(side="right", padx=5)

        # ── Pre-fill if editing ─────────────────────────────────────────
        if initial:
            self.name_entry.insert(0, initial.get("activity_name", ""))
            self._set_date("start", initial.get("start_date", ""))
            self._set_date("end", initial.get("end_date", ""))
            self._set_time_taken(initial.get("time_taken", ""))
            self.hard_text.insert("1.0", initial.get("hard_points", ""))
            self.status_combo.set(initial.get("status", "WIP"))
            self.attach_path_var.set(initial.get("attachment_path", ""))

            # Pre-select members
            saved_members = initial.get("members", "")
            if saved_members:
                saved_list = [m.strip() for m in saved_members.split(",")]
                for i, res in enumerate(self._resources):
                    display = f"{res.name} ({res.role})"
                    if res.name in saved_list or display in saved_list:
                        self.members_listbox.selection_set(i)

    # ── Date helper methods ─────────────────────────────────────────────

    def _update_days(self, which: str):
        """Update the day dropdown to match the selected month/year
           (handles 28/29/30/31 correctly)."""
        import calendar

        if which == "start":
            year_var, month_var, day_cb, day_var = (
                self._start_year_var, self._start_month_var,
                self._start_day_cb, self._start_day_var)
        else:
            year_var, month_var, day_cb, day_var = (
                self._end_year_var, self._end_month_var,
                self._end_day_cb, self._end_day_var)

        try:
            y = int(year_var.get())
            m = int(month_var.get())
            max_day = calendar.monthrange(y, m)[1]
        except (ValueError, KeyError):
            max_day = 31

        new_days = [f"{d:02d}" for d in range(1, max_day + 1)]
        day_cb["values"] = new_days

        # Clamp current selection if it exceeds the new max
        try:
            current_day = int(day_var.get())
            if current_day > max_day:
                day_var.set(f"{max_day:02d}")
        except ValueError:
            day_var.set("01")

    def _get_date_string(self, which: str) -> str:
        """Return the date as a YYYY-MM-DD string from either DateEntry or combo boxes."""
        if self._has_calendar:
            picker = self.start_date_picker if which == "start" else self.end_date_picker
            return picker.get_date().strftime("%Y-%m-%d")
        else:
            if which == "start":
                y, m, d = (self._start_year_var.get(),
                           self._start_month_var.get(),
                           self._start_day_var.get())
            else:
                y, m, d = (self._end_year_var.get(),
                           self._end_month_var.get(),
                           self._end_day_var.get())
            return f"{y}-{m}-{d}"

    def _set_date(self, which: str, date_str: str):
        """Set the date widgets from a YYYY-MM-DD string (used for pre-fill on edit)."""
        if not date_str:
            return

        try:
            dt = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return

        if self._has_calendar:
            picker = self.start_date_picker if which == "start" else self.end_date_picker
            picker.set_date(dt)
        else:
            if which == "start":
                self._start_year_var.set(str(dt.year))
                self._start_month_var.set(f"{dt.month:02d}")
                self._start_day_var.set(f"{dt.day:02d}")
                self._update_days("start")
            else:
                self._end_year_var.set(str(dt.year))
                self._end_month_var.set(f"{dt.month:02d}")
                self._end_day_var.set(f"{dt.day:02d}")
                self._update_days("end")

    # ── Time helper methods ─────────────────────────────────────────────

    def _get_time_taken(self) -> str:
        """Build a human-readable time string from the hours/minutes dropdowns."""
        try:
            hours = int(self._time_hours_var.get())
        except ValueError:
            hours = 0
        try:
            mins = int(self._time_mins_var.get())
        except ValueError:
            mins = 0

        if hours == 0 and mins == 0:
            return ""

        parts = []
        if hours > 0:
            parts.append(f"{hours}h")
        if mins > 0:
            parts.append(f"{mins}m")
        return " ".join(parts)

    def _set_time_taken(self, time_str: str):
        """Parse a time string like '2h 30m' or '45m' back into the dropdowns."""
        if not time_str:
            self.time_hours_cb.set("0")
            self.time_mins_cb.set("0")
            return

        import re
        hours = 0
        mins = 0

        # Match patterns like "2h", "30m", "2h 30m", "2h30m"
        h_match = re.search(r'(\d+)\s*h', time_str, re.IGNORECASE)
        m_match = re.search(r'(\d+)\s*m', time_str, re.IGNORECASE)

        if h_match:
            hours = int(h_match.group(1))
        if m_match:
            mins = int(m_match.group(1))

        # Snap minutes to nearest 5
        mins = round(mins / 5) * 5
        if mins >= 60:
            hours += mins // 60
            mins = mins % 60

        self.time_hours_cb.set(str(min(hours, 99)))
        self.time_mins_cb.set(str(mins))    
    
    def _browse_attachment(self):
        """Open file dialog to select an attachment."""
        filepath = filedialog.askopenfilename(
            title="Select Attachment",
            filetypes=[
                ("All files", "*.*"),
                ("Documents", "*.pdf *.docx *.xlsx *.pptx"),
                ("Images", "*.png *.jpg *.jpeg *.gif"),
            ]
        )
        if filepath:
            self.attach_path_var.set(filepath)

    def _save(self):
        """Validate and save the activity."""
        from tkinter import messagebox

        activity_name = self.name_entry.get().strip()
        status = self.status_combo.get()

        if not activity_name:
            messagebox.showwarning("Validation", "Activity name is required.")
            return
        if not status:
            messagebox.showwarning("Validation", "Please select a status.")
            return

        # Get dates from the pickers / combo boxes
        start_date = self._get_date_string("start")
        end_date = self._get_date_string("end")

        # Validate the assembled dates
        try:
            sd = datetime.datetime.strptime(start_date, "%Y-%m-%d").date()
            ed = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            messagebox.showwarning("Validation", "Invalid date selected.")
            return

        if ed < sd:
            messagebox.showwarning("Validation",
                                    "End date cannot be before the start date.")
            return

        # Collect selected members
        selected_indices = self.members_listbox.curselection()
        selected_members = []
        for idx in selected_indices:
            if idx < len(self._resources):
                selected_members.append(self._resources[idx].name)
        members_str = ", ".join(selected_members)

        self.result = {
            "activity_name": activity_name,
            "start_date": start_date,
            "end_date": end_date,
            "time_taken": self._get_time_taken(),
            "members": members_str,
            "hard_points": self.hard_text.get("1.0", "end-1c").strip(),
            "status": status,
            "attachment_path": self.attach_path_var.get(),
        }
        self.top.destroy()

# ─────────────────────────────────────────────────────────────────────────
# Launch function
# ─────────────────────────────────────────────────────────────────────────

def launch(db_path: pathlib.Path | None = None):
    """Open login → then main app with dashboard as first tab."""
    # Login
    login_root = tk.Tk()
    login_app = LoginWindow(login_root, db_path)
    login_root.mainloop()

    if login_app.user is None:
        print("❌ Login cancelled.")
        return

    print(f"✅ Logged in as {login_app.user['username']} ({login_app.user['role']})")

    # Main app
    app_root = tk.Tk()
    MainApp(app_root, login_app.user, db_path)
    app_root.mainloop()