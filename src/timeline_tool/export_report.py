"""
PDF/Excel report generation module for Project Management Tool.
Generates professional reports with Gantt chart, summary tables, and KPI dashboard.
"""

from __future__ import annotations

import datetime
import io
import pathlib
import tempfile
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from timeline_tool.models import Project, Milestone

# PDF generation with ReportLab
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        Image, PageBreak, KeepTogether
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Excel generation with openpyxl
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

import matplotlib.pyplot as plt
import matplotlib.dates as mdates


def check_dependencies():
    """Check if required dependencies are available."""
    missing = []
    if not REPORTLAB_AVAILABLE:
        missing.append("reportlab")
    if not OPENPYXL_AVAILABLE:
        missing.append("openpyxl")
    return missing


def generate_pdf_report(
    projects: list,
    output_path: str | pathlib.Path,
    title: str = "Project Status Report",
    include_gantt: bool = True,
    include_milestones: bool = True,
    include_kpis: bool = True,
    gantt_image_path: str | pathlib.Path | None = None,
) -> bool:
    """
    Generate a professional PDF report.
    
    Args:
        projects: List of Project objects
        output_path: Path to save the PDF
        title: Report title
        include_gantt: Whether to include Gantt chart image
        include_milestones: Whether to include milestone status table
        include_kpis: Whether to include KPI dashboard
        gantt_image_path: Path to pre-rendered Gantt chart image
    
    Returns:
        True if successful, False otherwise
    """
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab is required for PDF export. Install with: pip install reportlab")
    
    output_path = pathlib.Path(output_path)
    
    # Create document
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=1*cm,
        bottomMargin=1*cm,
    )
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#0067C0"),
    )
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceBefore=20,
        spaceAfter=10,
        textColor=colors.HexColor("#333333"),
    )
    normal_style = styles['Normal']
    
    # Build content
    story = []
    
    # Title
    story.append(Paragraph(title, title_style))
    story.append(Paragraph(
        f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}",
        ParagraphStyle('Date', parent=normal_style, alignment=TA_CENTER, textColor=colors.gray)
    ))
    story.append(Spacer(1, 0.5*inch))
    
    # Executive Summary
    story.append(Paragraph("📊 Executive Summary", heading_style))
    summary_data = _build_summary_table(projects)
    summary_table = Table(summary_data, colWidths=[2*inch, 1.5*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0067C0")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#F5F5F5")),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor("#DDDDDD")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9F9F9")]),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Gantt Chart Image
    if include_gantt and gantt_image_path:
        gantt_path = pathlib.Path(gantt_image_path)
        if gantt_path.exists():
            story.append(Paragraph("📈 Project Timeline (Gantt Chart)", heading_style))
            img = Image(str(gantt_path))
            # Scale to fit page width
            img.drawWidth = 10*inch
            img.drawHeight = 4*inch
            story.append(img)
            story.append(Spacer(1, 0.3*inch))
    
    # Project Details Table
    story.append(Paragraph("📁 Project Details", heading_style))
    project_data = _build_project_table(projects)
    project_table = Table(project_data, colWidths=[2.5*inch, 1*inch, 1*inch, 1*inch, 1.2*inch, 1*inch, 1*inch])
    project_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0067C0")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9F9F9")]),
    ]))
    story.append(project_table)
    
    # Milestone Status
    if include_milestones:
        story.append(PageBreak())
        story.append(Paragraph("🎯 Milestone Completion Status", heading_style))
        milestone_data = _build_milestone_table(projects)
        if len(milestone_data) > 1:
            milestone_table = Table(milestone_data, colWidths=[2*inch, 2*inch, 1*inch, 1.5*inch, 2*inch])
            milestone_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#16A34A")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9F9F9")]),
            ]))
            story.append(milestone_table)
        else:
            story.append(Paragraph("No milestones defined.", normal_style))
    
    # KPI Dashboard
    if include_kpis:
        story.append(Spacer(1, 0.3*inch))
        story.append(Paragraph("📊 KPI Dashboard", heading_style))
        kpi_data = _build_kpi_table(projects)
        kpi_table = Table(kpi_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#9333EA")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9F9F9")]),
        ]))
        story.append(kpi_table)
    
    # Build PDF
    doc.build(story)
    return True


def generate_excel_report(
    projects: list,
    output_path: str | pathlib.Path,
    title: str = "Project Status Report",
) -> bool:
    """
    Generate an Excel report with multiple sheets.
    
    Args:
        projects: List of Project objects
        output_path: Path to save the Excel file
        title: Report title
    
    Returns:
        True if successful, False otherwise
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    output_path = pathlib.Path(output_path)
    wb = openpyxl.Workbook()
    
    # Header style
    header_fill = PatternFill(start_color="0067C0", end_color="0067C0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary_sheet(ws_summary, projects, header_fill, header_font, header_alignment, thin_border)
    
    # Sheet 2: Projects
    ws_projects = wb.create_sheet("Projects")
    _write_projects_sheet(ws_projects, projects, header_fill, header_font, header_alignment, thin_border)
    
    # Sheet 3: Milestones
    ws_milestones = wb.create_sheet("Milestones")
    _write_milestones_sheet(ws_milestones, projects, header_fill, header_font, header_alignment, thin_border)
    
    # Sheet 4: KPIs
    ws_kpis = wb.create_sheet("KPIs")
    _write_kpis_sheet(ws_kpis, projects, header_fill, header_font, header_alignment, thin_border)
    
    wb.save(str(output_path))
    return True


def _build_summary_table(projects: list) -> list:
    """Build summary statistics table data."""
    total = len(projects)
    on_track = sum(1 for p in projects if p.computed_status() == "on-track")
    at_risk = sum(1 for p in projects if p.computed_status() == "at-risk")
    overdue = sum(1 for p in projects if p.computed_status() == "overdue")
    
    return [
        ["Metric", "Value"],
        ["Total Projects", str(total)],
        ["On Track", f"{on_track} ({100*on_track/total:.0f}%)" if total > 0 else "0"],
        ["At Risk", f"{at_risk} ({100*at_risk/total:.0f}%)" if total > 0 else "0"],
        ["Overdue", f"{overdue} ({100*overdue/total:.0f}%)" if total > 0 else "0"],
        ["Report Date", datetime.date.today().strftime("%Y-%m-%d")],
    ]


def _build_project_table(projects: list) -> list:
    """Build project details table data."""
    headers = ["Project Name", "Start", "End", "Duration", "Status", "Dev Region", "Sales Region"]
    rows = [headers]
    
    for p in projects:
        rows.append([
            p.name,
            p.start_date.strftime("%Y-%m-%d"),
            p.end_date.strftime("%Y-%m-%d"),
            f"{p.duration_days()} days",
            p.computed_status().upper(),
            p.dev_region or "-",
            p.sales_region or "-",
        ])
    
    return rows


def _build_milestone_table(projects: list) -> list:
    """Build milestone status table data."""
    headers = ["Project", "Milestone", "Date", "Status", "Tasks"]
    rows = [headers]
    
    for p in projects:
        for ms in p.milestones:
            task_summary = ""
            if ms.task_statuses:
                completed = ms.task_statuses.get("Completed", 0)
                total = sum(ms.task_statuses.values())
                task_summary = f"{completed}/{total} complete"
            
            status = "✅ Complete" if ms.is_complete() else "⏳ In Progress"
            if ms.date < datetime.date.today() and not ms.is_complete():
                status = "❌ Overdue"
            
            rows.append([
                p.name,
                ms.name,
                ms.date.strftime("%Y-%m-%d"),
                status,
                task_summary or "-",
            ])
    
    return rows


def _build_kpi_table(projects: list) -> list:
    """Build KPI dashboard table data."""
    headers = ["KPI", "Value", "Target", "Status"]
    rows = [headers]
    
    total = len(projects)
    on_track_pct = (sum(1 for p in projects if p.computed_status() == "on-track") / total * 100) if total > 0 else 0
    
    # Calculate milestone completion rate
    total_milestones = sum(len(p.milestones) for p in projects)
    completed_milestones = sum(
        1 for p in projects for ms in p.milestones if ms.is_complete()
    )
    milestone_pct = (completed_milestones / total_milestones * 100) if total_milestones > 0 else 0
    
    rows.append(["Project Health Rate", f"{on_track_pct:.1f}%", "≥80%", "✅" if on_track_pct >= 80 else "⚠️"])
    rows.append(["Milestone Completion", f"{milestone_pct:.1f}%", "≥70%", "✅" if milestone_pct >= 70 else "⚠️"])
    rows.append(["Total Active Projects", str(total), "-", "-"])
    rows.append(["Total Milestones", str(total_milestones), "-", "-"])
    
    return rows


def _write_summary_sheet(ws, projects, header_fill, header_font, header_alignment, border):
    """Write summary sheet to Excel workbook."""
    ws.cell(row=1, column=1, value="Project Status Summary")
    ws.cell(row=1, column=1).font = Font(bold=True, size=16)
    ws.merge_cells('A1:B1')
    
    data = _build_summary_table(projects)
    for row_idx, row_data in enumerate(data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if row_idx == 3:  # Header row
                cell.fill = header_fill
                cell.font = header_font
            cell.alignment = header_alignment
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20


def _write_projects_sheet(ws, projects, header_fill, header_font, header_alignment, border):
    """Write projects sheet to Excel workbook."""
    data = _build_project_table(projects)
    for row_idx, row_data in enumerate(data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            cell.alignment = header_alignment
    
    for i in range(1, 8):
        ws.column_dimensions[get_column_letter(i)].width = 15


def _write_milestones_sheet(ws, projects, header_fill, header_font, header_alignment, border):
    """Write milestones sheet to Excel workbook."""
    data = _build_milestone_table(projects)
    for row_idx, row_data in enumerate(data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            cell.alignment = header_alignment
    
    for i in range(1, 6):
        ws.column_dimensions[get_column_letter(i)].width = 18


def _write_kpis_sheet(ws, projects, header_fill, header_font, header_alignment, border):
    """Write KPIs sheet to Excel workbook."""
    data = _build_kpi_table(projects)
    for row_idx, row_data in enumerate(data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            cell.alignment = header_alignment
    
    for i in range(1, 5):
        ws.column_dimensions[get_column_letter(i)].width = 20